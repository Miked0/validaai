import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# ABA 1: Principal (instruções)
# ============================================================
ws_principal = wb.active
ws_principal.title = 'Principal'
ws_principal['A1'] = 'Guia'
ws_principal['A3'] = 'Verificar na aba Produtos a cadastrar, que os produtos estejam cadastrados com codigo de barras e precos corretos.'
ws_principal['A4'] = 'Executar testes da aba Testes, considerando os seguintes pontos:'
ws_principal['A5'] = 'PREENCHIMENTO SCANNTECH'
ws_principal['K5'] = 'Numero Loja:'
ws_principal['M5'] = 'Numero de Caixa:'
ws_principal.merge_cells('A3:N3')
ws_principal.merge_cells('A4:N4')

# ============================================================
# ABA 2: Produtos a cadastrar
# ============================================================
ws_pro = wb.create_sheet('Produtos a cadastrar')
ws_pro['A1'] = 'Ao realizar o cadastro não deve ser repetido o código de barras no código interno do item'
ws_pro.merge_cells('A1:J1')
ws_pro['A2'] = 'CODIGO INTERNO'
ws_pro['B2'] = 'CODIGO DE BARRA'
ws_pro['C2'] = 'DESCRIPCIÓN'
ws_pro['D2'] = 'PREÇO'
ws_pro['E2'] = 'Obs'

produtos = [
    ('', '7891000029329', 'CAFE SOLUVEL NESCAFE FORTE TRADICAO GRAN VD 100G', 11.9, ''),
    ('', '7891000010860', 'CAFE SOLUVEL NESCAFE TRADICAO REFIL 50G', 11.9, ''),
    ('', '7894904500383', 'HAMBURGUER BOVINO/AVES TEXAS BURGER 56G', 1.37, ''),
    ('', '7894904578207', 'PRODUTO 4', 1.50, ''),
    ('', '7894904003495', 'PRODUTO 5', 1.20, ''),
    ('', '7891149440801', 'PRODUTO 6', 2.50, ''),
    ('', '7891149102808', 'PRODUTO 7', 2.30, ''),
    ('', '7891991001359', 'PRODUTO 8', 1.80, ''),
    ('', '7891149105533', 'PRODUTO 9', 25.00, ''),
    ('', '7891149103119', 'PRODUTO 10', 20.25, ''),
    ('', '7896079500175', 'PRODUTO 11', 3.69, ''),
    ('', '7897511400237', 'PRODUTO 12', 3.50, ''),
    ('', '7897511400244', 'PRODUTO 13', 3.50, ''),
    ('', '7891024132906', 'PRODUTO 14', 3.50, ''),
    ('', '7891150024588', 'PRODUTO 15', 50.00, ''),
    ('', '5000329002537', 'PRODUTO 16', 100.00, ''),
    ('', '7894904573394', 'PRODUTO 17', 2.50, ''),
    ('', '7894904573387', 'PRODUTO 18', 2.50, ''),
    ('', '7891991294959', 'PRODUTO 19', 5.00, ''),
    ('', '7891991294942', 'PRODUTO 20', 5.00, ''),
    ('', '7891000315507', 'CAFE SOLUVEL NESCAFE MATINAL NORMAL V 100G', 15.00, ''),
    ('', '7891000306703', 'CAFE SOLUVEL NESCAFE ORIGINAL SACHET 50G', 12.00, ''),
    ('', '7891000029329', 'CAFE SOLUVEL NESCAFE FORTE TRADICAO GRAN VD 100G', 11.9, ''),
    ('', '7894000033730', 'PRODUTO 25', 10.00, ''),
    ('', '7891700202510', 'PRODUTO 26', 15.00, ''),
    ('', '7894000000299', 'PRODUTO 27', 12.00, ''),
    ('', '1003607622300391065', 'EAN INVALIDO 19 DIGITOS', 0.79, ''),
    ('', '7891999144485', 'PRODUTO 29', 6.75, ''),
]

for i, (cod_int, cod_bar, desc, preco, obs) in enumerate(produtos, 3):
    ws_pro[f'A{i}'] = cod_int
    ws_pro[f'B{i}'] = cod_bar
    ws_pro[f'C{i}'] = desc
    ws_pro[f'D{i}'] = preco
    ws_pro[f'E{i}'] = obs

# ============================================================
# ABA 3: Promocoes
# ============================================================
ws_promo = wb.create_sheet('Promocoes')
ws_promo['A1'] = '1 – HOMOLOGACION PACK DESCUENTO FIJO '
ws_promo['A2'] = 'Levando 3 unidades de qualquer um dos seguintes itens'
ws_promo['B2'] = 'Suporta'
ws_promo['C2'] = 'Sem limite'
ws_promo['A3'] = 'DESCUENTO FIJO'
ws_promo['B3'] = '7891000315507 - CAFE SOLUVEL NESCAFE MATINAL NORMAL V 100G'
ws_promo['A4'] = ''
ws_promo['B4'] = '7891000029329 - CAFE SOLUVEL NESCAFE FORTE TRADICAO GRAN VD 100G'
ws_promo['A5'] = ''
ws_promo['B5'] = '7891000010860 - CAFE SOLUVEL NESCAFE TRADICAO REFIL 50G'
ws_promo['A6'] = ''
ws_promo['B6'] = '7891000306703 - CAFE SOLUVEL NESCAFE ORIGINAL SACHET 50G'

wb.save('biblioteca/TEMPLATE_COM_BIN_NOVO.xlsx')
print('Base template saved')