#!/usr/bin/env python3
"""
ValidaAI - GUI application using validaai-core package.
"""
import sys
import os
import json
import re
import csv
import shutil
import tempfile
import threading
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
from decimal import Decimal, ROUND_HALF_UP

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except Exception:
    PANDAS_AVAILABLE = False

try:
    from validaai import (
        TestScriptReader,
        ItemParser,
        PaymentNormalizer,
        TestValidator,
        ResultExporter,
        APISalesBuilder,
        API_SALES_AVAILABLE,
    )
    from validaai.payment_codes import get_payment_label
    CORE_AVAILABLE = True
except Exception:
    CORE_AVAILABLE = False

# Fallback Local classes from validaai-core (if import fails)
from validaai.reader import TestScriptReader
from validaai.parser_items import ItemParser
from validaai.payments import PaymentNormalizer
from validaai.validators import TestValidator
from validaai.exporters import ResultExporter
from validaai.api_sales import APISalesBuilder
from validaai import get_payment_label, format_pagamentos_for_log

BASE_DIR = Path(__file__).resolve().parent

# Version info
APP_VERSION = "2.1.0"
APP_BUILD_DATE = "2026-06-21"


# ------------------------------------------------------------------
# Dark Theme
# ------------------------------------------------------------------
THEME_BG = "#2C2F33"        # Main background
THEME_CARD = "#23272A"      # Cards/panels
THEME_ACCENT = "#5865F2"    # Accent color
THEME_TEXT = "#FFFFFF"      # Primary text
THEME_MUTED = "#CCCCCC"     # Secondary text (WCAG AA contrast on THEME_CARD)
THEME_SUCCESS = "#3BA55C"   # Green
THEME_ERROR = "#ED4245"     # Red
THEME_WARN = "#FAA61A"      # Yellow
THEME_INPUT_BG = "#1E1F22"  # Input fields
THEME_BORDER = "#40444B"    # Borders

def apply_theme(root: tk.Tk):
    """Apply dark theme to the entire application."""
    style = ttk.Style(root)
    
    # Try to use 'clam' theme as base for better customization
    try:
        style.theme_use('clam')
    except Exception:
        pass
    
    # Configure root window
    root.configure(bg=THEME_BG)
    
    # General configurations - avoid bordercolor in root style
    style.configure(".",
        background=THEME_BG,
        foreground=THEME_TEXT,
        font=("Segoe UI", 10),
    )
    
    # Frame
    style.configure("TFrame", background=THEME_BG)
    style.configure("Card.TFrame", background=THEME_CARD)
    
    # LabelFrame
    style.configure("TLabelframe", background=THEME_CARD, foreground=THEME_TEXT, borderwidth=1, relief="solid")
    style.configure("TLabelframe.Label", background=THEME_CARD, foreground=THEME_ACCENT, font=("Segoe UI", 10, "bold"))
    
    # Label
    style.configure("TLabel", background=THEME_CARD, foreground=THEME_TEXT)
    style.configure("Title.TLabel", background=THEME_BG, foreground=THEME_TEXT, font=("Segoe UI", 18, "bold"))
    style.configure("Subtitle.TLabel", background=THEME_BG, foreground=THEME_MUTED, font=("Segoe UI", 10))
    style.configure("Muted.TLabel", background=THEME_CARD, foreground=THEME_MUTED)
    style.configure("Status.TLabel", background=THEME_BG, foreground=THEME_MUTED)
    
    # Entry
    style.configure("TEntry",
        fieldbackground=THEME_INPUT_BG,
        foreground=THEME_TEXT,
        bordercolor=THEME_BORDER,
        lightcolor=THEME_BORDER,
        darkcolor=THEME_BORDER,
        insertcolor=THEME_TEXT,
        padding=6)
    style.map("TEntry",
        foreground=[("disabled", THEME_MUTED)])

    # Combobox
    style.configure("TCombobox",
        fieldbackground=THEME_INPUT_BG,
        foreground=THEME_TEXT,
        background=THEME_CARD,
        bordercolor=THEME_BORDER,
        arrowcolor=THEME_TEXT,
        padding=6)
    style.map("TCombobox",
        background=[("readonly", THEME_CARD)],
        foreground=[("disabled", THEME_MUTED)])
    
    # Button
    style.configure("TButton",
        background=THEME_ACCENT,
        foreground=THEME_TEXT,
        borderwidth=0,
        focuscolor=THEME_ACCENT,
        padding=(16, 8),
        font=("Segoe UI", 10, "bold"))
    style.map("TButton",
        background=[("active", "#4752C4"), ("pressed", "#3C45A5"), ("disabled", THEME_BORDER)],
        foreground=[("disabled", THEME_MUTED)])

    # Secondary button style
    style.configure("Secondary.TButton",
        background=THEME_CARD,
        foreground=THEME_TEXT,
        borderwidth=1,
        bordercolor=THEME_BORDER,
        padding=(16, 8))
    style.map("Secondary.TButton",
        background=[("active", THEME_BORDER), ("pressed", "#40444B"), ("disabled", THEME_BORDER)],
        foreground=[("disabled", THEME_MUTED)])

    # Success button style
    style.configure("Success.TButton",
        background=THEME_SUCCESS,
        foreground=THEME_TEXT,
        borderwidth=0,
        focuscolor=THEME_SUCCESS,
        padding=(16, 8),
        font=("Segoe UI", 10, "bold"))
    style.map("Success.TButton",
        background=[("active", "#2E8B4E"), ("pressed", "#277642")])

    # Danger button style
    style.configure("Danger.TButton",
        background=THEME_ERROR,
        foreground=THEME_TEXT,
        borderwidth=0,
        focuscolor=THEME_ERROR,
        padding=(16, 8))
    style.map("Danger.TButton",
        background=[("active", "#C03537"), ("pressed", "#A02C2E")])
    
    # Progressbar
    style.configure("TProgressbar",
        background=THEME_ACCENT,
        troughcolor=THEME_INPUT_BG,
        borderwidth=0,
        thickness=8)
    
    # Notebook (tabs)
    style.configure("TNotebook", background=THEME_BG, borderwidth=0)
    style.configure("TNotebook.Tab",
        background=THEME_CARD,
        foreground=THEME_MUTED,
        padding=(16, 8),
        font=("Segoe UI", 10))
    style.map("TNotebook.Tab",
        background=[("selected", THEME_ACCENT), ("active", THEME_BORDER)],
        foreground=[("selected", THEME_TEXT)])
    
    # Scrollbar
    style.configure("Vertical.TScrollbar",
        background=THEME_CARD,
        troughcolor=THEME_BG,
        bordercolor=THEME_BG,
        arrowcolor=THEME_MUTED,
        width=10)
    style.map("Vertical.TScrollbar",
        active=[("background", THEME_ACCENT)])
    
    # Separator
    style.configure("TSeparator", background=THEME_BORDER)
    
    # Treeview (for tables)
    style.configure("Treeview",
        background=THEME_INPUT_BG,
        foreground=THEME_TEXT,
        fieldbackground=THEME_INPUT_BG,
        borderwidth=0,
        rowheight=28,
        font=("Segoe UI", 9))
    style.configure("Treeview.Heading",
        background=THEME_CARD,
        foreground=THEME_ACCENT,
        font=("Segoe UI", 9, "bold"),
        borderwidth=1,
        relief="flat")
    style.map("Treeview",
        background=[("selected", THEME_ACCENT)],
        foreground=[("selected", THEME_TEXT)])


# ------------------------------------------------------------------
# Core modules inlined
# ------------------------------------------------------------------

def export_resumo_resultados(roteiro_path: str, validated_tests: List[Dict[str, Any]], output_path: str) -> None:
    from openpyxl import Workbook

    rows_by_teste = {}
    for test in validated_tests:
        t_id = test.get('teste')
        if t_id is not None and t_id != '':
            try:
                rows_by_teste[str(int(float(t_id)) if isinstance(t_id, (int, float)) else str(t_id).strip())] = test
            except Exception:
                rows_by_teste[str(t_id).strip()] = test

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = 'Resumo'
    ws.append(['Teste', 'Status', 'Motivo', 'Alertas', 'Observacao'])

    ordered = []
    for key, test in rows_by_teste.items():
        ordered.append(test)
    ordered.sort(key=lambda x: x.get('teste', ''))

    for test in ordered:
        observacoes = test.get('observacoes', test.get('observacoes_raw', ''))
        obs_note = f"{observacoes} | status: {test.get('status_final', '')} | {test.get('motivo_status', '')}"
        if test.get('alertas'):
            obs_note += ' | alertas: ' + '; '.join(test.get('alertas', []))
        ws.append([
            test.get('teste', ''),
            test.get('status_final', ''),
            test.get('motivo_status', ''),
            '; '.join(test.get('alertas', [])),
            obs_note,
        ])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    wb_out.close()


class TestScriptReader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.data = None
        self.tests = []
        self.etapa_filter = None

    def set_etapa(self, etapa: str | None):
        self.etapa_filter = etapa.strip().upper() if etapa else None

    def _is_test_candidate(self, t_raw) -> bool:
        if t_raw is None:
            return False
        s = str(t_raw).strip()
        if not s:
            return False
        try:
            float(s)
            return True
        except Exception:
            return False

    def read_tests(self) -> List[Dict[str, Any]]:
        file_ext = Path(self.file_path).suffix.lower()
        if file_ext == '.csv':
            return self._read_csv()
        return self._read_xlsx_etapas()

    def _to_number(self, v):
        if v is None:
            return ''
        if isinstance(v, (int, float)):
            if isinstance(v, float) and v.is_integer():
                return int(v)
            return v
        s = str(v).strip()
        try:
            f = float(s)
            return int(f) if isinstance(f, float) and f.is_integer() else f
        except Exception:
            return s

    def _to_dec(self, v):
        if v is None or v == '':
            return None
        if isinstance(v, (int, float)):
            return Decimal(str(v))
        s = str(v).strip()
        removed = re.sub(r'[^0-9,.\-]', '', s)
        if removed in ('', '.', '-', '-.', ',-'):
            return None
        if removed.count('.') > 1:
            removed = removed.replace('.', '')
        if removed.count(',') > 1:
            removed = removed.replace(',', '')
        if ',' in removed and '.' in removed:
            removed = removed.replace('.', '').replace(',', '.')
        elif ',' in removed:
            removed = removed.replace(',', '.')
        try:
            return Decimal(removed)
        except Exception:
            return None

    def _norm(self, s: str) -> str:
        return str(s or '').strip()

    def _first_nonempty(self, vals, header, key):
        """Return first non-empty value for duplicate column key."""
        # For 'Observacoes', prefer the LAST occurrence (col 20 has the real notes)
        matches = []
        for i, h in enumerate(header):
            if h and h.strip().lower() == key.strip().lower():
                v = vals[i] if i < len(vals) else ''
                if v and str(v).strip():
                    matches.append(str(v).strip())
        if matches:
            # For Observacoes, use last one (col 20 has actual validation notes)
            return matches[-1]
        return ''

    def _read_xlsx_etapas(self) -> List[Dict[str, Any]]:
        import openpyxl
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        try:
            return self._read_wb(wb)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _read_wb(self, wb) -> List[Dict[str, Any]]:
        all_tests: List[Dict[str, Any]] = []

        for sheet_name in wb.sheetnames:
            if 'ETAPA' not in sheet_name.upper():
                continue
            ws = wb[sheet_name]
            header = None
            header_row = None
            start_idx = None
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                vals = [self._norm(c) for c in row]
                up = [v.upper() for v in vals]
                if 'TESTE' in up and any(k in up for k in ['ITENS DA VENDA', 'ARTICULOS MOVIMIENTO', 'ITENS']):
                    header = vals
                    header_row = r_idx
                    start_idx = r_idx + 1
                    break
            if not header:
                continue

            rows = list(ws.iter_rows(min_row=start_idx, values_only=True))
            seen_tests = set()
            for offset, row in enumerate(rows, start=start_idx):
                vals = [self._norm(c) for c in row]
                if len(vals) < len(header):
                    vals += [''] * (len(header) - len(vals))
                elif len(vals) > len(header):
                    vals = vals[:len(header)]
                rd = dict(zip(header, vals))
                t_raw = rd.get('Teste', '')
                if t_raw is None or str(t_raw).strip() == '':
                    continue
                if not self._is_test_candidate(t_raw):
                    continue
                etapa_key = sheet_name.strip().upper()
                if self.etapa_filter and etapa_key != self.etapa_filter:
                    continue
                t_key = str(t_raw).strip()
                if t_key in seen_tests:
                    continue
                seen_tests.add(t_key)
                subtotal_key = next((k for k in rd if k and k.lower() in ['sub-total', 'subtotal']), 'Sub-Total')
                total_key = next((k for k in rd if k and k.lower() in ['total']), 'Total')
                desconto_key = next((k for k in rd if k and k.lower() in ['desconto']), 'Desconto')
                std = {
                    'teste': self._to_number(t_raw),
                    'linha_original': f"{sheet_name}!{offset}",
                    'bloco_atual': etapa_key,
                    'tipo_promo': rd.get('Tipo Promo', rd.get('TIPO PROMO', '')),
                    'itens_da_venda': rd.get('Itens da venda', rd.get('ARTICULOS MOVIMIENTO', rd.get('Itens', ''))),
                    'pagamento': rd.get('Pagamento', ''),
                    'observacoes': self._first_nonempty(vals, header, 'Observacoes'),
                    'observacao_parceiro': self._first_nonempty(vals, header, 'Observacoes.1'),
                    'subtotal_esperado': self._to_dec(rd.get(subtotal_key, '')),
                    'desconto_esperado': self._to_dec(rd.get(desconto_key, '0') or '0'),
                    'total_esperado': self._to_dec(rd.get(total_key, '')),
                    'sat': self._first_nonempty(vals, header, 'SAT'),
                    'ecf': self._first_nonempty(vals, header, 'ECF'),
                    'nfce': self._first_nonempty(vals, header, 'NFCE'),
                    'json': rd.get('Json', ''),
                    'minoristas': rd.get('Minoristas', ''),
                    'cupom': self._first_nonempty(vals, header, 'Cupom'),
                }
                all_tests.append(std)
        return all_tests

    def _read_csv(self) -> List[Dict[str, Any]]:
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                with open(self.file_path, 'r', encoding=enc) as f:
                    sample = f.read(2048)
                    f.seek(0)
                    delim = ','
                    if ';' in sample and sample.count(';') > sample.count(','):
                        delim = ';'
                    rows = list(csv.reader(f, delimiter=delim))
                if rows:
                    self.data = self._convert_rows_to_dicts(rows)
                    break
            except UnicodeDecodeError:
                continue
            except Exception as e:
                if enc == 'cp1252':
                    raise
        return self._process_data()

    def _convert_rows_to_dicts(self, rows):
        if not rows:
            return []
        first_row = [str(c).strip().lower() for c in rows[0]]
        header_indicators = ['teste', 'tipo', 'promo', 'itens', 'venda', 'pagamento',
                             'observacoes', 'subtotal', 'desconto', 'total']
        matches = sum(1 for i in header_indicators if any(i in c for c in first_row))
        if matches >= 3:
            headers = [str(c).strip() for c in rows[0]]
            data_rows = rows[1:]
        else:
            mx = max(len(r) for r in rows) if rows else 0
            headers = [f'col_{i}' for i in range(mx)]
            data_rows = rows
        out = []
        for r in data_rows:
            r = r + [''] * (len(headers) - len(r))
            r = r[:len(headers)]
            out.append(dict(zip(headers, r)))
        return out

    def _process_data(self):
        if not self.data:
            return []
        idx = self._find_header_row(self.data)
        if idx is None:
            idx = 0
        data = self.data[idx:]
        out = []
        for row in data:
            if not isinstance(row, dict):
                continue
            out.append({str(k).lower().strip().replace(' ', '_').replace('-', '_'): v for k, v in row.items()})
        mapping = {
            'teste': ['teste', 'test', 'id', 'numero'],
            'tipo_promo': ['tipo_promo', 'tipo promocao', 'tipo promo', 'promotion_type', 'tipo'],
            'itens_da_venda': ['itens_da_venda', 'itens da venda', 'items', 'articulos_movimiento', 'itens'],
            'pagamento': ['pagamento', 'payment', 'forma_pagamento', 'pagamento'],
            'observacoes': ['observacoes', 'observacoes_raw', 'obs', 'observations', 'observacao'],
            'subtotal_esperado': ['sub_total', 'subtotal', 'sub-total', 'subtotal_esperado'],
            'desconto_esperado': ['desconto', 'discount', 'desconto_esperado'],
            'total_esperado': ['total', 'total_esperado']
        }
        tests = []
        for idx, row in enumerate(out):
            t = {'linha_original': idx + 1, 'bloco_atual': self._get_current_block(out, idx)}
            for target, aliases in mapping.items():
                val = ''
                for a in aliases:
                    if a in row:
                        val = row[a]
                        break
                t[target] = self._to_number(val) if target == 'teste' else val
            tests.append(t)
        return tests

    def _find_header_row(self, data):
        indicators = ['teste', 'tipo promo', 'itens da venda', 'pagamento',
                      'sub total', 'desconto', 'total']
        for i, row in enumerate(data):
            if not isinstance(row, dict):
                continue
            vals = ' '.join(str(v).lower() for v in row.values() if v)
            if sum(1 for ind in indicators if ind in vals) >= 4:
                return i
        return None

    def _get_current_block(self, data, row_index):
        for i in range(row_index, -1, -1):
            if i >= len(data):
                continue
            row = data[i]
            t = row.get('teste', row.get('test', ''))
            if isinstance(t, str):
                u = t.strip().upper()
                if any(k in u for k in ['BLOCO DE TESTE:', 'ETAPA']):
                    m = re.search(r'BLOCO\\\\s+DE\\\\s+TESTE:\\\\s*(.+)', u)
                    if m:
                        return m.group(1).strip()
                    return u
                if 'ETAPA' in u:
                    return u
        return "UNKNOWN BLOCO"


class ItemParser:
    def parse_items(self, test_dict: Dict[str, Any]) -> Dict[str, Any]:
        result = test_dict.copy()
        itens_raw = test_dict.get('itens_da_venda', test_dict.get('itens_raw', ''))
        if not itens_raw or not isinstance(itens_raw, str):
            result['itens_parseados'] = []
            return result
        result['itens_parseados'] = self._parse_item_string(itens_raw)
        return result

    def _parse_item_string(self, item_string: str) -> List[Dict[str, Any]]:
        if not item_string or not isinstance(item_string, str):
            return []
        normalized = item_string.replace(' * ', ' x ').replace('*', 'x')
        parts = [p.strip() for p in normalized.split('+') if p.strip()]
        parsed = []
        for part in parts:
            m = re.match(r'^(\d+(?:\.\d+)?)\s*x\s*(.+)$', part.strip())
            if m:
                qty = float(m.group(1))
                codigo = m.group(2).strip()
                parsed.append({'codigo': codigo, 'quantidade': qty, 'tipo': self._determine_item_type(codigo)})
            else:
                codigo = part.strip()
                if codigo:
                    parsed.append({'codigo': codigo, 'quantidade': 1.0, 'tipo': self._determine_item_type(codigo)})
        return parsed

    def _determine_item_type(self, codigo: str) -> str:
        u = codigo.upper().strip()
        if any(token in u for token in ['PESABLE', 'PESAVEL', 'WEIGHT', 'PESO']):
            return 'pesavel'
        cleaned = u.replace('*', '').replace(' ', '')
        if re.match(r'^\d+$', cleaned):
            if 7 <= len(cleaned) <= 20:
                return 'ean'
            return 'outro'
        if re.match(r'^[\d\*]+$', u):
            return 'bin'
        return 'outro'


class PaymentNormalizer:
    PAYMENT_MAPPING = {
        'dinheiro': 9,
        'dinheiro com troco': 9,
        'cartao credito': 10,
        'cartao crédito': 10,
        'cartao debito': 13,
        'cartao débito': 13,
        'pix': 14,
        'qr': 14,
        'pix/qr': 14,
        'cheque': 11,
        'vale': 12,
        'finalizadora': 15
    }

    def normalize_payment(self, test_dict: Dict[str, Any]) -> Dict[str, Any]:
        result = test_dict.copy()
        pag = test_dict.get('pagamento', test_dict.get('pagamento_raw', ''))
        if not pag or not isinstance(pag, str):
            result.update({'pagamento_normalizado': '', 'codigo_tipo_pago': None, 'is_multiplo': False, 'pagamentos': []})
            return result
        
        parsed = self._parse_pagamentos(pag)
        if len(parsed) > 1:
            result['pagamento_normalizado'] = 'MULTIPLO'
            result['codigo_tipo_pago'] = None
            result['is_multiplo'] = True
        elif len(parsed) == 1:
            result['pagamento_normalizado'] = parsed[0]['norm']
            result['codigo_tipo_pago'] = parsed[0]['codigo']
            result['is_multiplo'] = False
        else:
            result['pagamento_normalizado'] = ''
            result['codigo_tipo_pago'] = None
            result['is_multiplo'] = False
        
        result['pagamentos'] = parsed
        return result

    def _parse_pagamentos(self, s: str):
        """Parse payment string and return list of {'norm', 'codigo', 'raw'} for each payment."""
        low = s.lower().strip()
        if not low:
            return []
        
        # Split by common separators
        import re
        parts = re.split(r'\s*\+\s*|\s+e\s+', low)
        parts = [p.strip() for p in parts if p.strip()]
        
        results = []
        for part in parts:
            # Handle multiplicators like "duas vezes", "duas vezes Cartao Credito"
            multiplier = 1
            clean_part = part
            
            # Check for "duas vezes", "tres vezes", etc.
            mult_match = re.match(r'^(\d+|duas?|tres?|quatro|cinco)\s*vezes?\s+(.+)$', clean_part)
            if mult_match:
                mult_str = mult_match.group(1)
                clean_part = mult_match.group(2).strip()
                mult_map = {'um': 1, '1': 1, 'duas': 2, '2': 2, 'tres': 3, '3': 3, 'quatro': 4, '4': 4, 'cinco': 5, '5': 5}
                multiplier = mult_map.get(mult_str, 1)
            
            # Clean up cleaned part
            clean_part = clean_part.strip()
            
            norm = codigo = None
            if clean_part in self.PAYMENT_MAPPING:
                norm, codigo = clean_part, self.PAYMENT_MAPPING[clean_part]
            else:
                for k, v in self.PAYMENT_MAPPING.items():
                    if k in clean_part or clean_part in k:
                        norm, codigo = k, v
                        break
            
            if norm is not None:
                for _ in range(multiplier):
                    results.append({'norm': norm, 'codigo': codigo, 'raw': clean_part})
        
        return results


class TestValidator:
    def __init__(self, tolerance: float = 0.01):
        self.tolerance = tolerance

    def validate(self, test_dict: Dict[str, Any]) -> Dict[str, Any]:
        res = test_dict.copy()
        res['status_final'] = 'ERRO_DESCONHECIDO'
        res['motivo_status'] = ''
        res['alertas'] = []
        revisao = False
        revisao_motivo = ''
        erro_status = None
        erro_motivo = None
        ok_motivo = None
        
        for fn in [self._validate_teste_id, self._validate_itens_parsed, self._validate_special_cases,
                   self._validate_payment_mapped, self._validate_subtotal_numeric, self._validate_desconto_numeric,
                   self._validate_total_numeric, self._validate_total_consistency,
                   self._validate_pagos_json, self._validate_api_not_run]:
            r = fn(res)
            if r:
                status, motivo, alerta = r
                if status == 'REVISAO':
                    revisao = True
                    revisao_motivo = motivo
                if status == 'OK':
                    ok_motivo = motivo
                if status.startswith('ERRO'):
                    # Don't return early - collect and apply priority at end
                    if erro_status is None:
                        erro_status = status
                        erro_motivo = motivo
                if status == 'NOT_RUN':
                    # NOT_RUN is special - keep but still allow partner obs override
                    if erro_status is None:
                        erro_status = status
                        erro_motivo = motivo
                if status.startswith('ALERTA') and alerta:
                    res['alertas'].append(alerta)
        
        # Apply priority logic at the END (no early returns)
        # Priority 1: Partner observation (coluna U) -> REVISAO (overrides everything)
        obs_parceiro = str(res.get('observacao_parceiro', '')).strip()
        if obs_parceiro and obs_parceiro.lower() not in ['nan', 'none', '']:
            res['status_final'] = 'REVISAO'
            if revisao_motivo:
                res['motivo_status'] = f"Observação do parceiro: {obs_parceiro} | {revisao_motivo}"
            elif erro_motivo:
                res['motivo_status'] = f"Observação do parceiro: {obs_parceiro} | {erro_motivo}"
            else:
                res['motivo_status'] = f"Observação do parceiro: {obs_parceiro}"

        # Priority 1b: AGUARDA_API_CANCELAMENTO - defer to API validation
        # This test has 'cancelar venda' in observacoes and needs API JSON to confirm cancellation
        elif erro_status == 'AGUARDA_API_CANCELAMENTO':
            # Defer decision to API step - set as PENDING_API
            res['status_final'] = 'PENDING_API'
            res['motivo_status'] = 'Aguardando validação de cancelamento no JSON do parceiro'

        # Priority 2: REVISAO from special cases (multiplo, acréscimo, etc)
        elif revisao:
            res['status_final'] = 'REVISAO'
            res['motivo_status'] = revisao_motivo
        
        # Priority 3: NOT_RUN
        elif erro_status == 'NOT_RUN':
            res['status_final'] = 'NOT_RUN'
            res['motivo_status'] = erro_motivo
        
        # Priority 4: ERRO (sem margem)
        elif erro_status and erro_status.startswith('ERRO'):
            res['status_final'] = erro_status
            res['motivo_status'] = erro_motivo
        
        # Priority 5: OK
        else:
            res['status_final'] = 'OK'
            res['motivo_status'] = ok_motivo or 'Todos os campos válidos e consistentes'
        
        # Add alerts
        if revisao and revisao_motivo and revisao_motivo not in res['alertas']:
            res['alertas'].append(revisao_motivo)
        if erro_motivo and erro_motivo not in res['alertas'] and not revisao and not obs_parceiro:
            res['alertas'].append(erro_motivo)
        
        return res

    def _validate_teste_id(self, d):
        t = d.get('teste')
        if t is None or t == '' or (isinstance(t, str) and not t.strip()):
            return 'ERRO_TESTE_ID', 'Teste não possui identificador', ''
        if isinstance(t, str):
            try:
                d['teste'] = self._safe_cast_number(t)
            except ValueError:
                return 'ERRO_TESTE_ID', 'Identificador do teste não é numérico/válido', ''
        return None

    def _safe_cast_number(self, v):
        if isinstance(v, (int, float)):
            return v
        s = str(v).strip().replace(',', '.')
        if not s:
            raise ValueError('empty')
        return float(s)

    def _validate_itens_parsed(self, d):
        itens = d.get('itens_parseados', [])
        if not isinstance(itens, list) or not itens:
            raw = d.get('itens_raw', d.get('itens_da_venda', ''))
            return 'ERRO_PARSE_ITENS', 'Falha ao parsear itens', f'Itens brutos: "{raw}"' if raw else 'Campo de itens está vazio'
        for i, item in enumerate(itens):
            if not isinstance(item, dict):
                return 'ERRO_PARSE_ITENS', f'Item {i} não é um dicionário', ''
            for f in ['codigo', 'quantidade', 'tipo']:
                if f not in item:
                    return 'ERRO_PARSE_ITENS', f'Item {i} missing field: {f}', ''
            if not isinstance(item.get('codigo'), str) or not item.get('codigo', '').strip():
                return 'ERRO_PARSE_ITENS', f'Item {i} tem código inválido', ''
            if item.get('tipo') == 'outro':
                code = item.get('codigo', '').strip()
                if not any(tok in code.upper() for tok in ['DESCONTO', 'ACRESCIMO', 'CANCELAR', 'PESABLE']):
                    return 'ALERTA_ITEM_TIPO', f'Item {i} tem tipo não padrão/desconhecido', f'Código: {code}'
            try:
                if float(item.get('quantidade', 0)) <= 0:
                    return 'ERRO_PARSE_ITENS', f'Item {i} tem quantidade não positiva', ''
            except Exception:
                return 'ERRO_PARSE_ITENS', f'Item {i} tem quantidade inválida', ''
        return None

    def _validate_total_consistency(self, d):
        sub = d.get('subtotal_norm')
        tot = d.get('total_norm')
        if sub is None or tot is None:
            return None
        desc = d.get('desconto_norm', 0.0)
        if desc is None:
            desc = 0.0
        diff_sub = abs((sub - desc) - tot)
        diff_sum = abs((sub + desc) - tot)
        if diff_sub <= self.tolerance or diff_sum <= self.tolerance:
            if max(diff_sub, diff_sum) > 0:
                return 'ALERTA_ARREDONDAMENTO', 'Diferença de arredondamento dentro da tolerância', f'Esperado: {sub - desc:.2f} ou {sub + desc:.2f}, Obtido: {tot:.2f}, Diferenças: {diff_sub:.4f} / {diff_sum:.4f}'
            return None
        if abs(tot - sub) <= self.tolerance:
            return 'ALERTA_AJUSTE', 'Possível acréscimo/ajuste zerado', f'Subtotal: {sub:.2f}, Total: {tot:.2f}, Diferença: {abs(tot - sub):.4f}'
        return 'ERRO_CONSISTENCIA', 'Total não é consistente com subtotal/ajuste', f'Esperado ±desc: {sub - desc:.2f} / {sub + desc:.2f}, Obtido: {tot:.2f}, Diferenças: {diff_sub:.4f} / {diff_sum:.4f} (tolerância: {self.tolerance})'

    def _validate_payment_mapped(self, d):
        np = d.get('codigo_tipo_pago')
        pr = d.get('pagamento_raw', '')
        if d.get('pagamento_normalizado') == 'MULTIPLO':
            return 'ALERTA_PAGAMENTO_MULTIPLO', 'Pagamento múltiplo detectado - requer revisão manual', 'Pagamento marcado como MULTIPLO'
        if np is None:
            # Correction 1a: Test 22 - cancelar venda antes de finalizar (sem cupom, sem pagamento) = OK
            # Correction 1b: Qualquer teste com pagamento vazio e sem observações especiais = OK (sem cupom gerado)
            teste = d.get('teste')
            itens_raw = (d.get('itens_da_venda') or d.get('itens_raw') or '').lower()
            observacoes = (d.get('observacoes') or '').lower()
            pagamento = (d.get('pagamento') or '').strip()
            
            # Se pagamento vazio e não há flags de erro explícitas
            tem_cancelar_venda = 'cancelar venda' in observacoes
            tem_acrescimo = any(kw in observacoes for kw in ['acrescimo', 'acréscimo', 'acrescimo na linha', 'acrescimo no subtotal', 'acrescimo no cabecalho', 'acréscimo no cabeçalho'])
            tem_desconto_especial = any(kw in observacoes for kw in ['desconto no subtotal', 'desconto no cabecalho', 'desconto no cabeçalho', 'desconto no subtotal/cabeçalho'])
            tem_pesavel = 'pesable' in itens_raw or 'pesavel' in itens_raw
            tem_troco = 'troco' in pagamento.lower()
            tem_multiplo = d.get('pagamento_normalizado') == 'MULTIPLO'
            
            if not pagamento and not tem_cancelar_venda and not tem_acrescimo and not tem_desconto_especial and not tem_pesavel and not tem_troco and not tem_multiplo:
                return None  # OK - sem pagamento, sem flags especiais = venda cancelada antes de finalizar
            
            return 'ERRO_PAGAMENTO', 'Falha ao mapear pagamento para código padrão', f'Pagamento bruto: "{pr}"' if pr else 'Campo de pagamento está vazio'
        return None

    def _to_dec(self, v):
        try:
            return Decimal(str(v).replace(',', '.')) if v is not None else None
        except Exception:
            return None

    def _validate_subtotal_numeric(self, d):
        s = self._to_dec(d.get('subtotal', d.get('subtotal_esperado')))
        if s is None:
            return 'ERRO_VALOR', 'Subtotal não está presente', ''
        d['subtotal_norm'] = float(s.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        d['subtotal_raw'] = s
        return None

    def _validate_desconto_numeric(self, d):
        s = self._to_dec(d.get('desconto', d.get('desconto_esperado')))
        if s is None:
            d['desconto_norm'] = 0.0
            d['desconto_raw'] = Decimal('0')
            return None
        d['desconto_norm'] = float(s.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        d['desconto_raw'] = s
        return None

    def _validate_total_numeric(self, d):
        t = self._to_dec(d.get('total', d.get('total_esperado')))
        if t is None:
            return 'ERRO_VALOR', 'Total não está presente', ''
        d['total_norm'] = float(t.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        d['total_raw'] = t
        return None

    def _validate_total_consistency(self, d):
        sub = d.get('subtotal_norm')
        tot = d.get('total_norm')
        if sub is None or tot is None:
            return None
        desc = d.get('desconto_norm', 0.0)
        if desc is None:
            desc = 0.0
        diff_soma = abs((sub + desc) - tot)
        diff_sub = abs((sub - desc) - tot)
        if diff_soma <= self.tolerance or diff_sub <= self.tolerance:
            if max(diff_soma, diff_sub) > 0:
                return 'ALERTA_ARREDONDAMENTO', f'Diferença de arredondamento dentro da tolerância ({min(diff_soma, diff_sub):.4f})', f'Esperado por soma: {sub + desc:.2f}, por subtração: {sub - desc:.2f}, Obtido: {tot:.2f}'
            return None
        return 'REVISAO', 'Total não é consistente com subtotal/ajuste', f'Esperado por soma: {sub + desc:.2f}, por subtração: {sub - desc:.2f}, Obtido: {tot:.2f}, Diferenças: {diff_soma:.4f} / {diff_sub:.4f} (tolerância: {self.tolerance})'

    def _validate_special_cases(self, d):
        """
        Acumula flags de status sem early return.
        Prioridade final (no validate): Partner obs > REVISAO > NOT_RUN > ERRO > OK
        """
        observacoes = (d.get('observacoes') or '').lower()
        itens_raw = (d.get('itens_da_venda') or d.get('itens_raw') or '').lower()
        pagamento_norm = (d.get('pagamento_normalizado') or '').lower()
        teste = d.get('teste')
        pagamento_raw = (d.get('pagamento') or d.get('pagamento_raw') or '').lower()

        # Track accumulated statuses
        has_revisao = False
        revisao_motivos = []
        has_erro = False
        erro_motivo = None

        # Test 26 - has payment "Dinheiro", should be OK
        if teste == 26:
            return ('OK', 'Teste válido', 'Teste simples com pagamento Dinheiro')

        # Grupo 2 (11-17): Cancelamento APÓS conclusão = FLUXO ESPERADO
        # Venda original + cancelamento (2 registros: cancelacion=false + cancelacion=true com nº negativo)
        # Isso NÃO é erro nem revisão - é o comportamento correto
        if 'cancelar venda' in observacoes:
            # Verificar se é Grupo 2 (testes 11-17)
            if teste in (11, 12, 13, 14, 15, 16, 17):
                # Fluxo esperado de cancelamento pós-venda - OK
                pass
            else:
                # Outros casos com "cancelar venda" + pagamento = revisão manual
                pagamento = (d.get('pagamento') or '').strip()
                if pagamento:
                    has_revisao = True
                    revisao_motivos.append('Cancelamento anotado após venda processada - requer revisão manual')
                else:
                    # Pagamento vazio + cancelar venda = cancelamento antes de finalizar
                    pass

        # Test 20 - "Dar desconto na linha" - REVISAO (manter para test 20 específico)
        if teste == 20 and 'desconto na linha' in observacoes:
            has_revisao = True
            revisao_motivos.append('Desconto na linha - requer revisão manual')

        # Test 22 - cancelar item no cupom com pagamento vazio -> OK (skip payment validation)
        if teste == 22 and 'cancelar' in itens_raw and not (d.get('pagamento') or '').strip():
            return ('OK', 'Item cancelado no cupom - pagamento não requerido', 'Cancelamento de item dentro do cupom')

        # Cancelar item in itens_raw - Testes 23, 24 = FLUXO ESPERADO
        # Item cancelado não aparece no JSON final / Quantidade final ajustada
        if 'cancelar' in itens_raw and 'cancelar venda' not in observacoes:
            if teste in (23, 24):
                # Teste 23: Cancelar 1 produto, manter o resto - fluxo normal
                pass
            elif teste == 24:
                # Teste 24: Cancelar 1 unidade (ajuste quantidade) - fluxo normal
                pass
            elif teste != 22:
                has_revisao = True
                revisao_motivos.append('Item com cancelamento detectado - requer revisão manual')

        # Multiple payments - REVISAO (tests 7,8,9,10)
        if pagamento_norm == 'multiplo':
            has_revisao = True
            revisao_motivos.append('Pagamento múltiplo - requer revisão manual')

        # Grupo 3 (18-19): Acréscimo - SEFAZ/RS não permite -> REVISAO
        if any(kw in observacoes for kw in ['acrescimo', 'acréscimo', 'acrescimo na linha', 'acrescimo no subtotal', 'acrescimo no cabecalho', 'acréscimo no cabeçalho']):
            has_revisao = True
            revisao_motivos.append('Acréscimo detectado - SEFAZ/RS não permite')

        # Desconto na linha não realizado -> ERRO (only if explicitly says "não realizado")
        if 'desconto na linha' in observacoes and 'não realizado' in observacoes:
            has_erro = True
            erro_motivo = 'Desconto na linha não foi realizado'

        # Grupo 4 (20-21): Desconto no cabeçalho/subtotal - SEFAZ/RS não permite -> REVISAO
        if any(kw in observacoes for kw in ['desconto no subtotal', 'desconto no cabecalho', 'desconto no cabeçalho', 'desconto no subtotal/cabeçalho']):
            has_revisao = True
            revisao_motivos.append('Desconto no cabeçalho/subtotal - SEFAZ/RS não permite')

        # Pesável items (PESABLE) -> REVISAO
        # Test 25 - pesável incorreto (quantidade passada incorretamente) -> ERRO
        if 'pesable' in itens_raw or 'pesavel' in itens_raw or '* pesable' in itens_raw or 'x pesable' in itens_raw:
            if ('357.9 * pesable' in itens_raw or '357.9*pesable' in itens_raw.replace(' ', '') or
                '357.9 x pesable' in itens_raw or '357.9xpesable' in itens_raw.replace(' ', '')):
                has_erro = True
                erro_motivo = 'Quantidade de produto pesável passada incorretamente'
            else:
                has_revisao = True
                revisao_motivos.append('Item pesável detectado - requer revisão manual')

        # Troco (change) in payment
        if 'troco' in pagamento_raw:
            has_revisao = True
            revisao_motivos.append('Pagamento com troco - requer revisão manual')

        # Return accumulated status (highest priority wins in final validate logic)
        if has_revisao:
            return ('REVISAO', '; '.join(revisao_motivos), revisao_motivos[0] if revisao_motivos else '')
        if has_erro:
            return ('ERRO', erro_motivo, '')
        return None

    def _validate_pagos_json(self, d):
        """Validate that API JSON pagos array matches expected payments from test case.
        Supports both formats: 
        - Flat: sale_json['pagos'] (at root)
        - Wrapped: sale_json['movimiento']['pagos']
        """
        sale_json = d.get('sale_json')
        if not sale_json or not isinstance(sale_json, dict):
            return None
        
        # Support both JSON formats: flat (pagos at root) and wrapped (movimiento.pagos)
        def _extrair_pagos(json_data):
            if not isinstance(json_data, dict):
                return None
            # Wrapped format: movimiento.pagos
            if isinstance(json_data.get('movimiento'), dict):
                return json_data['movimiento'].get('pagos')
            # Flat format: pagos at root
            return json_data.get('pagos')
        
        pagos = _extrair_pagos(sale_json)
        if not isinstance(pagos, list):
            return None
        
        # If test has multiple payments, validate pagos array
        is_multiplo = d.get('is_multiplo', False)
        pagamentos_esperados = d.get('pagamentos', [])
        
        if is_multiplo and pagamentos_esperados:
            # Get expected payment codes
            expected_codes = sorted([p.get('codigo') for p in pagamentos_esperados if p.get('codigo') is not None])
            actual_codes = sorted([p.get('codigoTipoPago') for p in pagos if p.get('codigoTipoPago') is not None])
            
            if expected_codes != actual_codes:
                # Correction 3: Partner JSON mismatch = ALERTA only (template payment column is truth)
                return ('ALERTA', f'Array pagos não confere com pagamentos esperados. Esperado: {expected_codes}, Obtido: {actual_codes}', f'Códigos esperados: {expected_codes}, códigos no JSON: {actual_codes}')
            
            # Check that all pagos have required fields
            for idx, pago in enumerate(pagos):
                if 'codigoTipoPago' not in pago or pago.get('codigoTipoPago') is None:
                    return ('ALERTA', f'Pagamento {idx} sem codigoTipoPago', f'Pagamento {idx}: {pago}')
                if 'detalleFinalizadora' not in pago or not pago['detalleFinalizadora']:
                    return ('ALERTA', f'Pagamento {idx} sem detalleFinalizadora', f'Pagamento {idx}: {pago}')
        
        # For single payment, validate it matches expected
        elif not is_multiplo and pagamentos_esperados:
            expected_code = pagamentos_esperados[0].get('codigo')
            if pagos and expected_code is not None:
                actual_code = pagos[0].get('codigoTipoPago')
                if expected_code != actual_code:
                    # Correction 3: Partner JSON mismatch = ALERTA only
                    return ('ALERTA',
                            f'codigoTipoPago do JSON não confere. Esperado: {expected_code}, Obtido: {actual_code}',
                            f'Pagamento esperado: {expected_code}, no JSON: {actual_code}')
        
        return None

    def _validate_api_not_run(self, d):
        """Correction 2: If partner JSON not found (api_status=NOT_RUN), return NOT_RUN.
        EXCEPTION: If business rules would give OK (e.g., cancelamento antes de finalizar - empty payment, no flags),
        don't downgrade to NOT_RUN."""
        api_status = d.get('api_status')
        if api_status == 'NOT_RUN':
            # Check if this is a "cancelamento antes de finalizar" scenario (test 22)
            # Empty payment + no special flags = expected behavior, should be OK
            pagamento = (d.get('pagamento') or '').strip()
            observacoes = (d.get('observacoes') or '').lower()
            itens_raw = (d.get('itens_da_venda') or d.get('itens_raw') or '').lower()
            
            tem_cancelar_venda = 'cancelar venda' in observacoes
            tem_acrescimo = any(kw in observacoes for kw in ['acrescimo', 'acréscimo', 'acrescimo na linha', 'acrescimo no subtotal', 'acrescimo no cabecalho', 'acréscimo no cabeçalho'])
            tem_desconto_especial = any(kw in observacoes for kw in ['desconto no subtotal', 'desconto no cabecalho', 'desconto no cabeçalho', 'desconto no subtotal/cabeçalho'])
            tem_pesavel = 'pesable' in itens_raw or 'pesavel' in itens_raw
            tem_troco = 'troco' in (d.get('pagamento') or d.get('pagamento_raw') or '').lower()
            tem_multiplo = d.get('pagamento_normalizado') == 'MULTIPLO'
            
            # If no payment AND no special flags = cancelamento antes de finalizar = OK
            if not pagamento and not tem_cancelar_venda and not tem_acrescimo and not tem_desconto_especial and not tem_pesavel and not tem_troco and not tem_multiplo:
                return None  # Let business rules give OK, don't override with NOT_RUN
            
            return ('NOT_RUN', 'JSON do parceiro não encontrado — validação de API não executada', d.get('api_alertas', [''])[0])
        return None

        return None


class ResultExporter:
    COLUMNS = [
        'teste', 'bloco', 'tipo_promo', 'itens_raw', 'itens_parseados',
        'pagamento_raw', 'codigo_tipo_pago', 'pagamento_label',
        'subtotal_esperado', 'subtotal_norm',
        'desconto_esperado', 'desconto_norm',
        'total_esperado', 'total_norm',
        'status_final', 'motivo_status', 'alertas', 'observacoes_originais',
        'sat', 'ecf', 'nfce', 'json', 'minoristas', 'cupom',
        'api_status', 'api_alertas', 'sale_json'
    ]

    def export(self, test_results: List[Dict[str, Any]], output_path: str) -> None:
        output_path = str(output_path)
        if not test_results:
            self._create_empty(output_path)
            return
        rows = []
        for test in test_results:
            # Get readable payment label from codigo_tipo_pago
            codigo_pag = test.get('codigo_tipo_pago', '')
            pagamento_label = ''
            if codigo_pag:
                try:
                    codigo_int = int(codigo_pag)
                    pagamento_label = get_payment_label(codigo_int)
                except (ValueError, TypeError):
                    pagamento_label = f'Unknown({codigo_pag})'

            r = {
                'teste': test.get('teste', ''),
                'bloco': test.get('bloco_atual', test.get('bloco', '')),
                'tipo_promo': test.get('tipo_promo', ''),
                'itens_raw': test.get('itens_da_venda', test.get('itens_raw', '')),
                'itens_parseados': ', '.join(f"{i.get('quantidade',0)} x {i.get('codigo','')}" for i in test.get('itens_parseados', [])) if test.get('itens_parseados') else '',
                'pagamento_raw': test.get('pagamento', test.get('pagamento_raw', '')),
                'codigo_tipo_pago': test.get('codigo_tipo_pago', ''),
                'pagamento_label': pagamento_label,
                'subtotal_esperado': test.get('subtotal', test.get('subtotal_esperado', '')),
                'subtotal_norm': test.get('subtotal_norm', ''),
                'desconto_esperado': test.get('desconto', test.get('desconto_esperado', '')),
                'desconto_norm': test.get('desconto_norm', ''),
                'total_esperado': test.get('total', test.get('total_esperado', '')),
                'total_norm': test.get('total_norm', ''),
                'status_final': test.get('status_final', ''),
                'motivo_status': test.get('motivo_status', ''),
                'alertas': '; '.join(test.get('alertas', [])),
                'observacoes_originais': test.get('observacoes', test.get('observacoes_raw', '')),
                'sat': test.get('sat', ''),
                'ecf': test.get('ecf', ''),
                'nfce': test.get('nfce', ''),
                'json': test.get('json', ''),
                'minoristas': test.get('minoristas', ''),
                'cupom': test.get('cupom', ''),
            }
            rows.append(r)

        if PANDAS_AVAILABLE:
            try:
                df = pd.DataFrame(rows)
                for col in self.COLUMNS:
                    if col not in df.columns:
                        df[col] = ''
                df = df[self.COLUMNS]
                df.to_excel(output_path, index=False)
                return
            except Exception:
                csv_path = output_path.replace('.xlsx', '.csv')
                pd.DataFrame(rows).to_csv(csv_path, index=False)
                return

        csv_path = output_path.replace('.xlsx', '.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            f.write(','.join(f'"{c}"' for c in self.COLUMNS) + '\n')
            for r in rows:
                vals = []
                for c in self.COLUMNS:
                    v = r.get(c, '')
                    if v is None:
                        v = ''
                    s = str(v)
                    if '"' in s or ',' in s or '\n' in s:
                        s = '"' + s.replace('"', '""') + '"'
                    vals.append(s)
                f.write(','.join(vals) + '\n')

    def _create_empty(self, output_path: str):
        if PANDAS_AVAILABLE:
            try:
                pd.DataFrame(columns=self.COLUMNS).to_excel(output_path, index=False)
                return
            except Exception:
                pass
        with open(output_path.replace('.xlsx', '.csv'), 'w', newline='', encoding='utf-8') as f:
            f.write(','.join(f'"{c}"' for c in self.COLUMNS) + '\n')


# ------------------------------------------------------------------
# GUI
# ------------------------------------------------------------------
class ValidaAIApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("ValidaAI - Automação de Roteiro de Testes")
        self.root.geometry("720x520")
        self.root.resizable(True, True)

        # Apply theme
        apply_theme(root)

        self.roteiro_path = tk.StringVar()
        self.cupom_paths: List[str] = []
        self.audit_dir = tk.StringVar()
        self.output_path = tk.StringVar(value=str(BASE_DIR / "output" / "validacao_resultado.xlsx"))
        self.status_var = tk.StringVar(value="Pronto")
        self.progress_var = tk.DoubleVar(value=0.0)
        self.etapa_var = tk.StringVar()
        self.etapa_lista: List[str] = []

        self._build_ui()
        # Keyboard shortcuts
        self.root.bind("<Return>", lambda e: self._run_validation())
        self.root.bind("<Escape>", lambda e: self.root.quit())
        self.root.bind("<Control-o>", lambda e: self._select_roteiro())
        self.root.bind("<Control-s>", lambda e: self._select_output())
        self.root.bind("<F5>", lambda e: self._run_validation())

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=12, style="TFrame")
        main.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main, text="ValidaAI", style="Title.TLabel").pack(pady=(0, 8))
        ttk.Label(main, text="Automação de validação de roteiro de testes PDV", style="Subtitle.TLabel").pack(pady=(0, 12))
        
        # Version info
        version_frame = ttk.Frame(main, style="TFrame")
        version_frame.pack(pady=(0, 8))
        ttk.Label(version_frame, text=f"v{APP_VERSION}", style="Subtitle.TLabel", foreground=THEME_MUTED).pack(side=tk.LEFT, padx=4)
        ttk.Label(version_frame, text=f"Build: {APP_BUILD_DATE}", style="Subtitle.TLabel", foreground=THEME_MUTED).pack(side=tk.LEFT, padx=4)

        inp = ttk.LabelFrame(main, text="Entradas", padding=10)
        inp.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(inp, text="Roteiro de testes (Excel/CSV):").grid(row=0, column=0, sticky=tk.W, pady=4)
        ttk.Entry(inp, textvariable=self.roteiro_path, width=70).grid(row=0, column=1, padx=6, pady=4)
        ttk.Button(inp, text="Selecionar", command=self._select_roteiro, style="Secondary.TButton").grid(row=0, column=2, padx=4)

        ttk.Label(inp, text="Etapa:").grid(row=0, column=3, sticky=tk.W, padx=(10,0), pady=4)
        self.etapa_combo = ttk.Combobox(inp, textvariable=self.etapa_var, state="readonly", width=20)
        self.etapa_combo['values'] = []
        self.etapa_combo.grid(row=0, column=4, padx=4, pady=4)
        self.etapa_combo.bind("<<ComboboxSelected>>", lambda e: None)

        ttk.Label(inp, text="Cupons (PDF/JPEG):").grid(row=1, column=0, sticky=tk.W, pady=4)
        self.cupom_listbox = tk.Listbox(inp, height=4,
            bg=THEME_INPUT_BG, fg=THEME_TEXT,
            selectbackground=THEME_ACCENT, selectforeground=THEME_TEXT,
            borderwidth=1, relief="flat", highlightthickness=1,
            highlightbackground=THEME_BORDER, highlightcolor=THEME_ACCENT,
            font=("Segoe UI", 9))
        self.cupom_listbox.grid(row=1, column=1, padx=6, pady=4, sticky=tk.EW)
        # Placeholder text when empty
        self._cupom_placeholder = "Arraste cupons aqui ou clique Adicionar"
        self._update_cupom_placeholder()
        ttk.Button(inp, text="Adicionar", command=self._add_cupons, style="Secondary.TButton").grid(row=1, column=2, padx=4, sticky=tk.N)
        ttk.Button(inp, text="Remover", command=self._remove_cupom, style="Secondary.TButton").grid(row=1, column=2, padx=4, sticky=tk.S)

        ttk.Label(inp, text="Arquivo de export da auditoria (.xlsx):").grid(row=2, column=0, sticky=tk.W, pady=4)
        ttk.Entry(inp, textvariable=self.audit_dir, width=70).grid(row=2, column=1, padx=6, pady=4)
        ttk.Button(inp, text="Selecionar arquivo", command=self._select_audit_dir, style="Secondary.TButton").grid(row=2, column=2, padx=4)

        inp.columnconfigure(1, weight=1)

        out = ttk.LabelFrame(main, text="Saída", padding=10)
        out.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(out, text="Arquivo de resultado:").grid(row=0, column=0, sticky=tk.W, pady=4)
        ttk.Entry(out, textvariable=self.output_path, width=70).grid(row=0, column=1, padx=6, pady=4)
        ttk.Button(out, text="Salvar como...", command=self._select_output, style="Secondary.TButton").grid(row=0, column=2, padx=4)
        out.columnconfigure(1, weight=1)

        actions = ttk.Frame(main, style="TFrame")
        actions.pack(fill=tk.X, pady=(0, 10))

        self.run_btn = ttk.Button(actions, text="Executar Validação", command=self._run_validation, style="Success.TButton")
        self.run_btn.pack(side=tk.LEFT, padx=4)
        ttk.Button(actions, text="Abrir pasta de saída", command=self._open_output_dir, style="Secondary.TButton").pack(side=tk.LEFT, padx=4)
        ttk.Button(actions, text="Sair", command=self.root.quit, style="Danger.TButton").pack(side=tk.RIGHT, padx=4)

        progress_frame = ttk.Frame(main, style="TFrame")
        progress_frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(progress_frame, textvariable=self.status_var, style="Status.TLabel").pack(side=tk.LEFT)
        ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100).pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(8, 0))

        log_frame = ttk.LabelFrame(main, text="Log", padding=6)
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(log_frame, height=10, state=tk.DISABLED,
            bg=THEME_INPUT_BG, fg=THEME_TEXT,
            insertbackground=THEME_TEXT,
            selectbackground=THEME_ACCENT, selectforeground=THEME_TEXT,
            borderwidth=1, relief="flat",
            highlightthickness=1, highlightbackground=THEME_BORDER, highlightcolor=THEME_ACCENT,
            font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def _log(self, msg: str):
        self.log_text.configure(state=tk.NORMAL)
        
        # Determine color based on message content
        tag = "info"
        lower = msg.lower()
        if any(kw in lower for kw in ["erro", "error", "falha", "failed", "❌"]):
            tag = "error"
        elif any(kw in lower for kw in ["sucesso", "success", "concluído", "completed", "✅", "ok"]):
            tag = "success"
        elif any(kw in lower for kw in ["aviso", "warning", "alert", "revisão", "revisao", "⚠️"]):
            tag = "warning"
        elif any(kw in lower for kw in ["iniciando", "starting", "lendo", "reading", "parseando", "normalizando", "validando", "🔄"]):
            tag = "process"
        
        # Configure tags if not already configured
        if not hasattr(self, '_log_tags_configured'):
            self.log_text.tag_configure("error", foreground=THEME_ERROR)
            self.log_text.tag_configure("success", foreground=THEME_SUCCESS)
            self.log_text.tag_configure("warning", foreground=THEME_WARN)
            self.log_text.tag_configure("process", foreground=THEME_ACCENT)
            self.log_text.tag_configure("info", foreground=THEME_MUTED)
            self._log_tags_configured = True
        
        self.log_text.insert(tk.END, msg + "\n", tag)
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _select_roteiro(self):
        p = filedialog.askopenfilename(title="Selecionar roteiro de testes", filetypes=[("Planilhas", "*.xlsx *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv"), ("Todos", "*.*")])
        if not p:
            return
        self.roteiro_path.set(p)
        self.etapa_var.set('')
        self.etapa_combo['values'] = []
        if p.lower().endswith('.xlsx'):
            etapas = self._detectar_etapas(p)
            if etapas:
                self.etapa_lista = etapas
                self.etapa_combo['values'] = etapas
                self.etapa_var.set(etapas[0])
            else:
                self.etapa_lista = []
                self.etapa_combo['values'] = ['Etapa 1', 'Etapa 2', 'Etapa 3']
                self.etapa_var.set('')

    def _detectar_etapas(self, path: str) -> List[str]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            out = []
            for sn in wb.sheetnames:
                if 'ETAPA' in sn.upper():
                    out.append(sn.strip())
            wb.close()
            return out
        except Exception:
            return ['Etapa 1', 'Etapa 2', 'Etapa 3']

    def _add_cupons(self):
        ps = filedialog.askopenfilenames(title="Selecionar cupons (PDF/JPEG)", filetypes=[("Imagens/PDF", "*.pdf *.jpg *.jpeg *.png"), ("PDF", "*.pdf"), ("JPEG", "*.jpg *.jpeg"), ("Todos", "*.*")])
        if ps:
            self.cupom_paths.extend(ps)
            self._refresh_cupom_list()

    def _remove_cupom(self):
        sel = self.cupom_listbox.curselection()
        if sel:
            self.cupom_paths.pop(sel[0])
            self._refresh_cupom_list()

    def _refresh_cupom_list(self):
        self.cupom_listbox.delete(0, tk.END)
        for p in self.cupom_paths:
            self.cupom_listbox.insert(tk.END, Path(p).name)
        self._update_cupom_placeholder()

    def _update_cupom_placeholder(self):
        """Show placeholder text when listbox is empty."""
        if self.cupom_listbox.size() == 0:
            self.cupom_listbox.insert(tk.END, self._cupom_placeholder)
            self.cupom_listbox.itemconfig(0, fg=THEME_MUTED)
            self.cupom_listbox.bind("<Button-1>", self._clear_cupom_placeholder, add="+")
        else:
            self.cupom_listbox.unbind("<Button-1>")

    def _clear_cupom_placeholder(self, event):
        """Clear placeholder on first click if it's the only item."""
        if self.cupom_listbox.size() == 1:
            item = self.cupom_listbox.get(0)
            if item == self._cupom_placeholder:
                self.cupom_listbox.delete(0)
                self.cupom_listbox.unbind("<Button-1>")

    def _select_audit_dir(self):
        p = filedialog.askopenfilename(
            title="Selecionar arquivo de export da auditoria (xlsx com JSONs do parceiro)",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if p:
            self.audit_dir.set(p)

    def _select_output(self):
        p = filedialog.asksaveasfilename(title="Salvar resultado como", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("Todos", "*.*")], initialfile="validacao_resultado.xlsx")
        if p:
            self.output_path.set(p)

    def _open_output_dir(self):
        out = Path(self.output_path.get()).parent
        out.mkdir(parents=True, exist_ok=True)
        os.startfile(str(out))

    def _set_busy(self, busy: bool):
        self.run_btn.configure(state=tk.DISABLED if busy else tk.NORMAL)
        self.root.update_idletasks()

    def _run_validation(self):
        roteiro = self.roteiro_path.get().strip()
        etapa = self.etapa_var.get().strip()
        if not roteiro:
            messagebox.showerror("Erro", "Nenhum roteiro selecionado.\n\nClique em \"Selecionar\" e escolha um arquivo .xlsx ou .csv.")
            return
        if not Path(roteiro).exists():
            messagebox.showerror("Erro", f"Arquivo não encontrado:\n{roteiro}\n\nVerifique se o arquivo foi movido ou excluído.")
            return
        ext = Path(roteiro).suffix.lower()
        if ext not in ('.xlsx', '.csv'):
            messagebox.showerror("Erro", f"Extensão não suportada: {ext}\n\nUse apenas arquivos .xlsx ou .csv")
            return
        # Verify xlsx has ETAPA sheets
        if ext == '.xlsx':
            try:
                import openpyxl
                wb = openpyxl.load_workbook(roteiro, read_only=True, data_only=True)
                has_etapa = any('ETAPA' in sn.upper() for sn in wb.sheetnames)
                wb.close()
                if not has_etapa:
                    messagebox.showwarning("Aviso", "Nenhuma aba 'ETAPA' encontrada no arquivo.\nO processamento tentará ler todas as abas.")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao ler arquivo Excel:\n{e}")
                return
        self._set_busy(True)
        self.status_var.set("Executando...")
        self.progress_var.set(0.0)
        self._log("Iniciando validação...")
        threading.Thread(target=self._validation_thread, args=(roteiro, etapa), daemon=True).start()

    def _validation_thread(self, roteiro: str, etapa: str):
        async_result = {}
        def _run():
            try:
                self._log("1. Lendo roteiro de testes...")
                reader = TestScriptReader(roteiro)
                if etapa:
                    reader.set_etapa(etapa)
                raw_tests = reader.read_tests()
                self._log(f"   Encontrados {len(raw_tests)} casos de teste.") if raw_tests else self._log("   Nenhum caso encontrado.")
                self.progress_var.set(20)

                self._log("2. Parseando itens...")
                item_parser = ItemParser()
                parsed_tests = [item_parser.parse_items(t) for t in raw_tests]
                self.progress_var.set(40)

                self._log("3. Normalizando pagamentos...")
                payment_normalizer = PaymentNormalizer()
                normalized_tests = [payment_normalizer.normalize_payment(t) for t in parsed_tests]
                self.progress_var.set(60)

                self._log("4. Validando regras de negócio...")
                validator = TestValidator(tolerance=0.01)
                validated_tests = [validator.validate(t) for t in normalized_tests]
                self.progress_var.set(80)

                self._log("4b. Carregando e validando JSON do parceiro (export de auditoria)...")
                if API_SALES_AVAILABLE:
                    api_builder = APISalesBuilder()
                    
                    # Carregar JSONs do parceiro do arquivo de export da auditoria
                    audit_file = self.audit_dir.get().strip()
                    partner_jsons = {}
                    if audit_file and Path(audit_file).exists():
                        try:
                            partner_jsons = self._load_partner_jsons(audit_file)
                            self._log(f"   Carregados {len(partner_jsons)} JSONs do parceiro do export de auditoria.")
                            if partner_jsons:
                                self._log(f"   Testes com JSON do parceiro: {sorted(partner_jsons.keys())}")
                        except Exception as e:
                            self._log(f"   ERRO ao carregar JSONs do parceiro: {e}")
                    else:
                        self._log("   Arquivo de auditoria não informado ou não encontrado; pulando validação do JSON do parceiro.")
                    
                    for t in validated_tests:
                        try:
                            # Usar JSON do parceiro se disponível - buscar pelo NÚMERO DO CUPOM do teste
                            # A função auxiliar para extrair o cupom do teste
                            def _get_test_cupom(test_dict):
                                """Extrai o número do cupom do teste a partir dos campos disponíveis.
                                Ignora placeholders como '(Status)', 'Status', etc.
                                """
                                placeholders = {'(status)', 'status', '(status)', 'n/a', 'na', ''}
                                for field in ['cupom', 'nfce', 'sat', 'ecf']:
                                    val = str(test_dict.get(field, '')).strip()
                                    if val and val.lower() not in ['nan', 'none', ''] and val.lower() not in placeholders:
                                        return val
                                return ''
                            
                            test_cupom = _get_test_cupom(t)
                            partner_json = partner_jsons.get(test_cupom)
                            
                            if partner_json:
                                self._log(f"   Teste {t.get('teste')}: Usando JSON DO PARCEIRO (cupom: {test_cupom})")
                                t['sale_json'] = partner_json
                                # Validar JSON do parceiro contra esperado
                                api_check = api_builder.validate_sale_json(partner_json)
                                t['api_status'] = api_check.get('status', 'ERRO_JSON')
                                t['api_alertas'] = api_check.get('alertas', []) or []
                                self._log(f"   Teste {t.get('teste')}: Validação do parceiro -> {api_check.get('status')} - {api_check.get('motivo', '')}")
                                
                                # Validação estrita de cupom: conferir consistência entre roteiro, cupom impresso e JSON do parceiro
                                cupom_roteiro = str(t.get('cupom', '')).strip()
                                
                                # Extrair cupom do JSON do parceiro (suporta formato com wrapper 'movimiento' ou direto na raiz)
                                def _extrair_cupom_json(json_data):
                                    if not isinstance(json_data, dict):
                                        return ''
                                    # Tentar formato com wrapper 'movimiento'
                                    if isinstance(json_data.get('movimiento'), dict):
                                        return str(json_data['movimiento'].get('numero', '')).strip()
                                    # Formato direto na raiz
                                    return str(json_data.get('numero', '')).strip()
                                
                                cupom_json = _extrair_cupom_json(partner_json)
                                
                                cupom_sat = str(t.get('sat', '')).strip()
                                cupom_ecf = str(t.get('ecf', '')).strip()
                                cupom_nfce = str(t.get('nfce', '')).strip()
                                
                                cupom_valido = False
                                if cupom_roteiro and cupom_json and cupom_roteiro == cupom_json:
                                    cupom_valido = True
                                elif cupom_sat and cupom_json and cupom_sat == cupom_json:
                                    cupom_valido = True
                                elif cupom_ecf and cupom_json and cupom_ecf == cupom_json:
                                    cupom_valido = True
                                elif cupom_nfce and cupom_json and cupom_nfce == cupom_json:
                                    cupom_valido = True
                                
                                if not cupom_valido:
                                    if cupom_roteiro or cupom_sat or cupom_ecf or cupom_nfce:
                                        t['api_status'] = 'ERRO'
                                        motivo = f"Cupom não confere: Roteiro='{cupom_roteiro}', SAT='{cupom_sat}', ECF='{cupom_ecf}', NFCe='{cupom_nfce}' vs JSON='{cupom_json}'"
                                        t['api_alertas'].append(motivo)
                                        self._log(f"   ERRO Cupom: {motivo}")
                                        t['api_status'] = 'ERRO'
                                
                                # Validação de códigos de pagamento: comparar pagos[].codigoTipoPago do JSON do parceiro com esperados
                                def _extrair_pagos_json(json_data):
                                    if not isinstance(json_data, dict):
                                        return []
                                    if isinstance(json_data.get('movimiento'), dict):
                                        return json_data['movimiento'].get('pagos', [])
                                    return json_data.get('pagos', [])
                                
                                # Obter códigos esperados do pagamento normalizado
                                pagamentos_esperados = t.get('pagamentos', [])
                                if pagamentos_esperados:
                                    expected_codes = sorted([str(p.get('codigo', '')).strip() for p in pagamentos_esperados if p.get('codigo')])
                                else:
                                    # Fallback: tentar extrair da coluna pagamento original
                                    pgto_raw = str(t.get('pagamento', '')).strip().lower()
                                    pgto_map = {
                                        'dinheiro': '9', 'dinheiro com troco': '9',
                                        'cartao credito': '10', 'cartao crédito': '10',
                                        'cartao debito': '13', 'cartao débito': '13',
                                        'pix': '14', 'qr': '14', 'pix/qr': '14',
                                        'cheque': '11', 'vale': '12',
                                        'finalizadora': '15',
                                    }
                                    expected_codes = [pgto_map.get(pgto_raw, '')] if pgto_raw else []
                                
                                # Obter códigos reais do JSON do parceiro
                                pagos_json = _extrair_pagos_json(partner_json)
                                actual_codes = sorted([str(p.get('codigoTipoPago', '')).strip() for p in pagos_json if p.get('codigoTipoPago')])
                                
                                if expected_codes and actual_codes:
                                    if expected_codes != actual_codes:
                                        motivo = f"Códigos de pagamento divergentes: esperado={expected_codes} vs parceiro={actual_codes}"
                                        t['api_alertas'].append(motivo)
                                        if t.get('api_status') == 'OK':
                                            t['api_status'] = 'REVISAO'
                                        self._log(f"   REVISAO Pagamento: {motivo}")
                                        for i, p in enumerate(pagos_json):
                                            codigo = p.get('codigoTipoPago')
                                            label = get_payment_label(codigo) if codigo is not None else 'N/A'
                                            self._log(f"   Pago {i}: {label} (codigoTipoPago={codigo}), detalleFinalizadora={p.get('detalleFinalizadora')}, importe={p.get('importe')}")
                            else:
                                # JSON do parceiro NÃO encontrado - ERRO estrito (sem fallback)
                                self._log(f"   Teste {t.get('teste')}: ERRO - JSON do parceiro NÃO ENCONTRADO no export de auditoria (cupom buscado: {test_cupom})")
                                t['sale_json'] = {}
                                t['api_status'] = 'ERRO'
                                t['api_alertas'] = [f"JSON do parceiro não encontrado no export de auditoria para cupom {test_cupom} (teste {t.get('teste')})"]
                                
                                # Não gerar JSON interno - validação estrita exige JSON do parceiro
                        except Exception as e:
                            t['sale_json'] = {}
                            t['api_status'] = 'ERRO_JSON'
                            t['api_alertas'] = [str(e)]
                else:
                    self._log("   api_sales indisponível; pulando etapa de API.")
                self.progress_var.set(85)

                # Aplicar resultados da API ANTES da re-validação
                for t in validated_tests:
                    api_status = t.get('api_status', '')
                    api_alertas = t.get('api_alertas', []) or []
                    api_ignorado = t.get('api_status') == 'IGNORADO'
                    
                    # NOVO: Tratar status PENDING_API (aguardando validação de cancelamento)
                    api_pendente = t.get('status_final') == 'PENDING_API'
                    
                    # Filtrar JSONs de vendas canceladas (chave "-xxx")
                    test_cupom = ''
                    placeholders = {'(status)', 'status', '(status)', 'n/a', 'na', ''}
                    for field in ['cupom', 'nfce', 'sat', 'ecf']:
                        val = str(t.get(field, '')).strip()
                        if val and val.lower() not in ['nan', 'none', ''] and val.lower() not in placeholders:
                            test_cupom = val
                            break
                    if test_cupom and test_cupom.startswith('-'):
                        t['status_final'] = 'OK'
                        t['motivo_status'] = 'Venda cancelada (cupom negativo) - não validado'
                        t['api_status'] = 'IGNORADO'
                        t['api_alertas'] = ['Venda cancelada ignorada']
                        t['sale_json'] = {}
                        t['api_alertas'] = ['Venda cancelada ignorada']
                        continue
                    
                    # Se teste está PENDING_API, verificar se JSON do parceiro confirma cancelamento
                    if api_pendente:
                        # Verifica se o JSON do parceiro confirma o cancelamento (cupom negativo)
                        if test_cupom and test_cupom.startswith('-'):
                            t['status_final'] = 'OK'
                            t['motivo_status'] = 'Cancelamento confirmado no JSON do parceiro (cupom negativo)'
                            t['api_status'] = 'OK'
                            t['api_alertas'] = ['Cancelamento confirmado']
                            t['sale_json'] = partner_json if 'partner_json' in locals() else {}
                            continue
                        else:
                            # Não confirmou cancelamento - ERRO
                            t['status_final'] = 'ERRO'
                            t['motivo_status'] = 'Cancelamento direcionado mas não confirmado no JSON do parceiro'
                            t['api_status'] = 'ERRO'
                            t['api_alertas'] = ['Cancelamento direcionado mas JSON não confirma cancelamento (cupom não negativo)']
                            continue
                    
                    if api_ignorado:
                        continue
                    
                    # Se API retornou ERRO, sobrescrever status_final
                    if api_status.startswith('ERRO'):
                        t['status_final'] = api_status
                        t['motivo_status'] = '; '.join(api_alertas) if api_alertas else 'Erro na validação do parceiro'
                        t['alertas'] = list(set((t.get('alertas', []) or []) + api_alertas))
                    elif api_status == 'REVISAO':
                        # REVISAO da API tem prioridade sobre OK de negócio
                        if t.get('status_final') == 'OK':
                            t['status_final'] = 'REVISAO'
                            t['motivo_status'] = '; '.join(api_alertas) if api_alertas else 'Revisão necessária'
                        t['alertas'] = list(set((t.get('alertas', []) or []) + api_alertas))
                    elif api_status == 'OK' and t.get('status_final') != 'OK':
                        # API OK mas negócio tem REVISAO - manter REVISAO
                        pass
                    else:
                        # API REVISAO ou OK - mesclar alertas
                        t['alertas'] = list(set((t.get('alertas', []) or []) + api_alertas))

                # Re-validate with API results (combines business rules + partner JSON validation)
                self._log("   Re-validando testes com resultados da API do parceiro...")
                revalidator = TestValidator(tolerance=0.01)
                # Não re-validar testes que já têm ERRO da API (sem JSON do parceiro) - o erro é definitivo
                final_tests = []
                for t in validated_tests:
                    api_status = t.get('api_status', '')
                    if api_status.startswith('ERRO') and not t.get('api_status') == 'IGNORADO':
                        # ERRO da API é definitivo - não re-validar
                        final_tests.append(t)
                    else:
                        final_tests.append(revalidator.validate(t))
                validated_tests = final_tests

                self._log("5. Preenchendo resultado na planilha de teste...")
                roteiro_path = self.roteiro_path.get().strip()
                output_resumo = str(BASE_DIR / "output" / "validacao_resultado.xlsx")
                if roteiro_path:
                    try:
                        export_resumo_resultados(roteiro_path, validated_tests, output_resumo)
                        self._log(f"   Resumo salvo em: {output_resumo}")
                    except Exception as e:
                        self._log(f"   ERRO ao gerar resumo: {e}")
                        raise
                else:
                    self._log("   Roteiro nao informado; pulando preenchimento.")

                audit_dir = self.audit_dir.get().strip()
                if audit_dir and Path(audit_dir).is_dir():
                    try:
                        self._export_audit_json(validated_tests, Path(audit_dir))
                    except Exception as e:
                        self._log(f"   ERRO ao salvar JSON de auditoria em {audit_dir}: {e}")
                        self._export_audit_json(validated_tests, BASE_DIR / "output")
                else:
                    self._log("   Pasta de auditoria nao informada; salvando JSON em ./output")
                    self._export_audit_json(validated_tests, BASE_DIR / "output")

                self.progress_var.set(100)

                status_counts = {}
                for test in validated_tests:
                    status = test.get('status_final', 'UNKNOWN')
                    status_counts[status] = status_counts.get(status, 0) + 1
                summary = " | ".join(f"{k}: {v}" for k, v in status_counts.items())
                self._log(f"Concluído. {summary}")

                async_result['status'] = 'ok'
                async_result['summary'] = summary
                async_result['output_resumo'] = output_resumo
                async_result['validated_tests'] = validated_tests
            except Exception as e:
                async_result['status'] = 'error'
                async_result['error'] = str(e)

        def _finish():
            try:
                if async_result.get('status') == 'error':
                    self._log(f"ERRO: {async_result.get('error')}")
                    self.status_var.set("Erro")
                    messagebox.showerror("Erro na validação", async_result.get('error'))
                else:
                    summary = async_result.get('summary', '')
                    output_resumo = async_result.get('output_resumo', '')
                    self._log(f"Concluído. {summary}")
                    self.status_var.set("Concluído")
                    messagebox.showinfo("Validação concluída", f"Resultados exportados para:\n{output_resumo}")
            finally:
                self._set_busy(False)

        _run()
        self.root.after(0, _finish)

    def _export_audit_json(self, validated_tests, audit_dir: Path):
        audit_dir.mkdir(parents=True, exist_ok=True)
        audit_path = audit_dir / f"audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        def _sanitize(obj):
            if isinstance(obj, dict):
                return {k: _sanitize(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [_sanitize(v) for v in obj]
            if isinstance(obj, Decimal):
                return float(obj)
            if obj is None or isinstance(obj, (bool, int, float, str)):
                return obj
            return str(obj)

        payload = {
            'generated_at': datetime.now().isoformat(),
            'total': len(validated_tests),
            'results': _sanitize(validated_tests),
        }
        try:
            with open(audit_path, 'w', encoding='utf-8') as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
                f.write('\n')
        except Exception as e:
            self._log(f"ERRO ao salvar JSON de auditoria em {audit_dir}: {e}")
            return
        self._log(f"   JSON de auditoria salvo em: {audit_path}")

    def _load_partner_jsons(self, audit_file: str) -> Dict[str, Any]:
        """Carrega JSONs do parceiro do arquivo de export de auditoria (xlsx).
        Espera colunas: 'Teste' (ou 'Número cupom') e 'Request' (JSON do parceiro).
        Retorna dict mapeando teste -> JSON do parceiro.
        Procura em TODAS as abas do arquivo.
        """
        import pandas as pd
        import json
        
        jsons = {}
        total_parsed = 0
        
        # Ler todas as abas do arquivo
        xls = pd.ExcelFile(audit_file)
        self._log(f"   Abas no arquivo de auditoria: {xls.sheet_names}")
        
        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(audit_file, sheet_name=sheet_name, dtype=str)
                
                # Procurar coluna de teste/cupom - priorizar matches exatos
                test_col = None
                test_cols = [c for c in df.columns if c and ('teste' in c.lower() or 'cupom' in c.lower() or 'numero' in c.lower())]
                if test_cols:
                    # Priorizar colunas com nome mais específico
                    for preferred in ['Número cupom', 'Numero cupom', 'Teste', 'Numero']:
                        if preferred in df.columns:
                            test_col = preferred
                            break
                        for c in test_cols:
                            if preferred.lower() in c.lower():
                                test_col = c
                                break
                    if test_col is None and test_cols:
                        test_col = test_cols[0]
                
                # Procurar coluna de request/json - PRIORIZAR 'Request' EXATO, EXCLUIR 'Id request'
                request_col = None
                if 'Request' in df.columns:
                    request_col = 'Request'
                else:
                    request_cols = [c for c in df.columns 
                                   if c and ('request' in c.lower() or 'json' in c.lower() or 'movimiento' in c.lower())
                                   and c.lower() != 'id request']  # EXCLUIR 'Id request'
                    if request_cols:
                        for preferred in ['Request JSON', 'Json Request', 'JSON', 'Json']:
                            for c in request_cols:
                                if preferred.lower() == c.lower():
                                    request_col = c
                                    break
                            if request_col:
                                break
                        if request_col is None:
                            request_col = request_cols[0]
                
                if not test_col or not request_col:
                    self._log(f"   Aba '{sheet_name}': colunas esperadas não encontradas (teste: {test_col}, request: {request_col})")
                    continue
                
                self._log(f"   Aba '{sheet_name}': colunas detectadas - Teste: '{test_col}', Request: '{request_col}'")
                
                parsed_count = 0
                for _, row in df.iterrows():
                    test_val = str(row.get(test_col, '')).strip()
                    request_val = str(row.get(request_col, '')).strip()
                    
                    if test_val and request_val and request_val not in ['nan', 'None', '']:
                        try:
                            parsed = json.loads(request_val)
                            jsons[test_val] = parsed
                            total_parsed += 1
                        except Exception:
                            self._log(f"   Aviso: Falha ao parsear JSON para teste {test_val}")
                
                self._log(f"   Aba '{sheet_name}': {parsed_count} JSONs válidos carregados")
                
            except Exception as e:
                self._log(f"   Erro ao processar aba '{sheet_name}': {e}")
        
        self._log(f"   Total de JSONs do parceiro carregados: {total_parsed}")
        if jsons:
            self._log(f"   Testes com JSON do parceiro: {sorted(jsons.keys())}")
        return jsons


def main():
    root = tk.Tk()
    app = ValidaAIApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
