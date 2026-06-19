"""
Export Pipeline - Pluggable exporters for validation results.
Supports Excel, CSV, JSON Audit, and HTML Report formats.
"""
import json
import csv
from abc import ABC, abstractmethod
from pathlib import Path
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional, Protocol
from dataclasses import dataclass, field, asdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


# ════════════════════════════════════════════════════════════
# Data Models
# ════════════════════════════════════════════════════════════

@dataclass
class ExportResult:
    """Result of an export operation."""
    exporter_name: str
    output_path: Path
    success: bool
    error: Optional[str] = None
    rows_exported: int = 0
    duration_ms: int = 0


@dataclass
class ExportConfig:
    """Configuration for export pipeline."""
    exporters: List[str] = field(default_factory=lambda: ["excel", "csv", "json_audit", "html_report"])
    excel: Dict[str, Any] = field(default_factory=dict)
    csv: Dict[str, Any] = field(default_factory=dict)
    json_audit: Dict[str, Any] = field(default_factory=dict)
    html_report: Dict[str, Any] = field(default_factory=dict)
    output: Dict[str, Any] = field(default_factory=dict)
    notifications: Dict[str, Any] = field(default_factory=dict)
    scheduler: Dict[str, Any] = field(default_factory=dict)
    testing: Dict[str, Any] = field(default_factory=dict)
    
    # Legacy flat fields for backward compatibility
    output_dir: str = "output"
    filename_template: str = "validacao_{timestamp}_{etapa}_v{version}"
    
    def __post_init__(self):
        # Sync legacy fields from nested output config
        if self.output:
            self.output_dir = self.output.get("directory", self.output_dir)
            self.filename_template = self.output.get("filename_template", self.filename_template)
    
    @classmethod
    def from_file(cls, path: Path) -> "ExportConfig":
        """Load config from JSON file."""
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return cls(**data)
        return cls()
    
    def to_file(self, path: Path) -> None:
        """Save config to JSON file."""
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(asdict(self), f, indent=2, ensure_ascii=False)
    
    def get_exporter_config(self, exporter_name: str) -> Dict[str, Any]:
        """Get configuration for a specific exporter."""
        return getattr(self, exporter_name, {}) or self.output.get(exporter_name, {})


# ════════════════════════════════════════════════════════════
# Base Exporter Interface
# ════════════════════════════════════════════════════════════

class Exporter(ABC):
    """Abstract base class for all exporters."""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        self.config = config or {}
        # Only set default name if subclass doesn't define its own name property
        if not hasattr(self.__class__, 'name') or isinstance(getattr(self.__class__, 'name', None), property):
            # Check if the name property is defined in this class or inherited
            for cls in self.__class__.__mro__:
                if 'name' in cls.__dict__ and isinstance(cls.__dict__['name'], property):
                    # Subclass has its own name property, don't override
                    break
            else:
                # No custom name property found, use default
                self.name = self.__class__.__name__.replace("Exporter", "").lower()
    
    @abstractmethod
    def export(
        self, 
        results: List[Dict[str, Any]], 
        output_path: Path,
        metadata: Optional[Dict[str, Any]] = None
    ) -> ExportResult:
        """Export validation results to output path."""
        pass
    
    @property
    @abstractmethod
    def extension(self) -> str:
        """File extension for this exporter (e.g., '.xlsx')."""
        pass
    
    def _sanitize(self, obj: Any) -> Any:
        """Convert non-serializable types to JSON-safe values."""
        if isinstance(obj, dict):
            return {k: self._sanitize(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [self._sanitize(v) for v in obj]
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, datetime):
            return obj.isoformat()
        if obj is None or isinstance(obj, (bool, int, float, str)):
            return obj
        return str(obj)
    
    def _prepare_row(self, test: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare a single test result row for export."""
        # Standard columns order
        cols = [
            "teste", "bloco", "tipo_promo",
            "itens_raw", "itens_parsed",
            "pagamento_raw", "codigo_tipo_pago", "pagamento_label",
            "subtotal_esperado", "subtotal_norm",
            "desconto_esperado", "desconto_norm",
            "total_esperado", "total_norm",
            "status_final", "motivo_status", "alertas",
            "observacoes", "observacao_parceiro",
            "sat", "ecf", "nfce", "cupom",
            "api_status", "api_alertas"
        ]
        
        row = {}
        for col in cols:
            val = test.get(col, "")
            if val is None:
                val = ""
            elif isinstance(val, (list, tuple)):
                val = "; ".join(str(v) for v in val)
            elif isinstance(val, (dict, Decimal, datetime)):
                val = str(val)
            row[col] = val
        
        # Add payment label if missing
        if not row.get("pagamento_label") and row.get("codigo_tipo_pago"):
            try:
                from validaai import get_payment_label
                row["pagamento_label"] = get_payment_label(int(row["codigo_tipo_pago"]))
            except Exception:
                row["pagamento_label"] = f"Unknown({row['codigo_tipo_pago']})"
        
        return row


# ════════════════════════════════════════════════════════════
# Excel Exporter
# ════════════════════════════════════════════════════════════

class ExcelExporter(Exporter):
    """Export to Excel (.xlsx) with formatting."""
    
    @property
    def extension(self) -> str:
        return ".xlsx"
    
    def export(
        self, 
        results: List[Dict[str, Any]], 
        output_path: Path,
        metadata: Optional[Dict[str, Any]] = None
    ) -> ExportResult:
        import time
        start = time.perf_counter()
        
        if not OPENPYXL_AVAILABLE:
            return ExportResult(
                exporter_name=self.name,
                output_path=output_path,
                success=False,
                error="openpyxl not installed"
            )
        
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.config.get("sheet_name", "Resumo")
            
            # Prepare rows
            rows = [self._prepare_row(t) for t in results]
            if not rows:
                # Empty workbook with headers only
                cols = list(rows[0].keys()) if rows else [
                    "teste", "status_final", "motivo_status"
                ]
                ws.append(cols)
                ws.append([""] * len(cols))
            else:
                # Headers
                cols = list(rows[0].keys())
                ws.append(cols)
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")
                header_alignment = Alignment(horizontal="center", wrap_text=True)
                
                for col_idx, header in enumerate(cols, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Data rows
                status_colors = {
                    "OK": "3BA55C",
                    "REVISAO": "FAA61A", 
                    "ERRO": "ED4245",
                    "ERRO_PAGAMENTO": "ED4245",
                    "NOT_RUN": "CCCCCC"
                }
                
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, key in enumerate(cols, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
                        cell.alignment = Alignment(wrap_text=True)
                        
                        # Color status column
                        if key == "status_final":
                            color = status_colors.get(row[key], "000000")
                            cell.font = Font(color=color, bold=True)
                
                # Auto-filter
                ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
                
                # Freeze panes
                if self.config.get("freeze_header", True):
                    ws.freeze_panes = "A2"
                
                # Auto-fit columns (approx)
                for col_idx, key in enumerate(cols, 1):
                    max_len = max(
                        len(str(key)),
                        max((len(str(row[key])) for row in rows), default=0)
                    )
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
            
            # Metadata sheet
            if metadata:
                meta_ws = wb.create_sheet("Metadados")
                meta_ws.append(["Chave", "Valor"])
                for k, v in metadata.items():
                    meta_ws.append([k, str(v)])
            
            wb.save(output_path)
            wb.close()
            
            duration = int((time.perf_counter() - start) * 1000)
            return ExportResult(
                exporter_name=self.name,
                output_path=output_path,
                success=True,
                rows_exported=len(rows),
                duration_ms=duration
            )
            
        except Exception as e:
            return ExportResult(
                exporter_name=self.name,
                output_path=output_path,
                success=False,
                error=str(e)
            )


# ════════════════════════════════════════════════════════════
# CSV Exporter
# ════════════════════════════════════════════════════════════

class CSVExporter(Exporter):
    """Export to CSV with UTF-8 encoding."""
    
    @property
    def extension(self) -> str:
        return ".csv"
    
    def export(
        self, 
        results: List[Dict[str, Any]], 
        output_path: Path,
        metadata: Optional[Dict[str, Any]] = None
    ) -> ExportResult:
        import time
        start = time.perf_counter()
        
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            rows = [self._prepare_row(t) for t in results]
            if not rows:
                return ExportResult(
                    exporter_name=self.name,
                    output_path=output_path,
                    success=True,
                    rows_exported=0,
                    duration_ms=int((time.perf_counter() - start) * 1000)
                )
            
            cols = list(rows[0].keys())
            delimiter = self.config.get("delimiter", ",")
            encoding = self.config.get("encoding", "utf-8-sig")
            
            with open(output_path, "w", newline="", encoding=encoding) as f:
                writer = csv.DictWriter(f, fieldnames=cols, delimiter=delimiter, quoting=csv.QUOTE_ALL)
                writer.writeheader()
                writer.writerows(rows)
            
            duration = int((time.perf_counter() - start) * 1000)
            return ExportResult(
                exporter_name=self.name,
                output_path=output_path,
                success=True,
                rows_exported=len(rows),
                duration_ms=duration
            )
            
        except Exception as e:
            return ExportResult(
                exporter_name=self.name,
                output_path=output_path,
                success=False,
                error=str(e)
            )


# ════════════════════════════════════════════════════════════
# JSON Audit Exporter
# ════════════════════════════════════════════════════════════

class JSONAuditExporter(Exporter):
    """Export full validation results as structured JSON for audit trail."""
    
    @property
    def name(self) -> str:
        return "json_audit"
    
    @property
    def extension(self) -> str:
        return ".json"
    
    def export(
        self, 
        results: List[Dict[str, Any]], 
        output_path: Path,
        metadata: Optional[Dict[str, Any]] = None
    ) -> ExportResult:
        import time
        start = time.perf_counter()
        
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Sanitize results
            sanitized = [self._sanitize(t) for t in results]
            
            payload = {
                "generated_at": datetime.now().isoformat(),
                "app_version": metadata.get("app_version", "unknown") if metadata else "unknown",
                "total": len(sanitized),
                "summary": self._compute_summary(sanitized),
                "metadata": metadata or {},
                "results": sanitized
            }
            
            indent = self.config.get("indent", 2)
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=indent)
                f.write("\n")
            
            duration = int((time.perf_counter() - start) * 1000)
            return ExportResult(
                exporter_name=self.name,
                output_path=output_path,
                success=True,
                rows_exported=len(sanitized),
                duration_ms=duration
            )
            
        except Exception as e:
            return ExportResult(
                exporter_name=self.name,
                output_path=output_path,
                success=False,
                error=str(e)
            )
    
    def _compute_summary(self, results: List[Dict]) -> Dict[str, Any]:
        """Compute summary statistics."""
        from collections import Counter
        status_counts = Counter(r.get("status_final", "UNKNOWN") for r in results)
        
        by_etapa = {}
        for r in results:
            etapa = r.get("bloco", r.get("bloco_atual", "UNKNOWN"))
            if etapa not in by_etapa:
                by_etapa[etapa] = Counter()
            by_etapa[etapa][r.get("status_final", "UNKNOWN")] += 1
        
        return {
            "by_status": dict(status_counts),
            "by_etapa": {k: dict(v) for k, v in by_etapa.items()},
            "total_with_partner_json": sum(1 for r in results if r.get("sale_json")),
            "total_errors": sum(1 for r in results if r.get("status_final", "").startswith("ERRO")),
            "total_revisao": sum(1 for r in results if r.get("status_final") == "REVISAO"),
            "total_ok": sum(1 for r in results if r.get("status_final") == "OK"),
        }


# ════════════════════════════════════════════════════════════
# HTML Report Exporter
# ════════════════════════════════════════════════════════════

class HTMLReportExporter(Exporter):
    """Export interactive HTML dashboard report."""
    
    @property
    def name(self) -> str:
        return "html_report"
    
    @property
    def extension(self) -> str:
        return ".html"
    
    # Inline CSS (tokens-based dark theme)
    CSS = """
    :root {
        --bg-primary: #2C2F33;
        --bg-card: #23272A;
        --bg-input: #1E1F22;
        --fg-primary: #FFFFFF;
        --fg-muted: #CCCCCC;
        --accent: #5865F2;
        --accent-hover: #4752C4;
        --success: #3BA55C;
        --error: #ED4245;
        --warning: #FAA61A;
        --border: #40444B;
        --radius: 6px;
        --font: 'Segoe UI', system-ui, sans-serif;
    }
    * { box-sizing: border-box; }
    body { margin: 0; font-family: var(--font); background: var(--bg-primary); color: var(--fg-primary); line-height: 1.5; }
    .container { max-width: 1400px; margin: 0 auto; padding: 24px; }
    header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 24px; padding-bottom: 16px; border-bottom: 1px solid var(--border); }
    h1 { margin: 0; font-size: 28px; }
    .meta { display: flex; gap: 16px; font-size: 14px; color: var(--fg-muted); }
    .summary-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 16px; margin-bottom: 24px; }
    .stat-card { background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius); padding: 16px; text-align: center; }
    .stat-value { font-size: 32px; font-weight: 700; }
    .stat-label { font-size: 12px; color: var(--fg-muted); text-transform: uppercase; margin-top: 4px; }
    .stat-ok .stat-value { color: var(--success); }
    .stat-revisao .stat-value { color: var(--warning); }
    .stat-erro .stat-value { color: var(--error); }
    .stat-notrun .stat-value { color: var(--fg-muted); }
    
    .tabs { display: flex; gap: 4px; margin-bottom: 16px; border-bottom: 1px solid var(--border); }
    .tab { padding: 12px 20px; background: none; border: none; color: var(--fg-muted); cursor: pointer; font: inherit; border-bottom: 2px solid transparent; transition: all 0.2s; }
    .tab:hover { color: var(--fg-primary); }
    .tab.active { color: var(--accent); border-bottom-color: var(--accent); }
    
    .tab-panel { display: none; animation: fadeIn 0.2s; }
    .tab-panel.active { display: block; }
    @keyframes fadeIn { from { opacity: 0; transform: translateY(4px); } to { opacity: 1; transform: none; } }
    
    .table-wrapper { overflow-x: auto; background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius); }
    table { width: 100%; border-collapse: collapse; font-size: 13px; }
    th, td { padding: 10px 12px; text-align: left; border-bottom: 1px solid var(--border); }
    th { background: var(--bg-input); font-weight: 600; color: var(--accent); position: sticky; top: 0; z-index: 1; }
    tr:hover td { background: var(--bg-input); }
    .status-badge { display: inline-flex; align-items: center; gap: 4px; padding: 4px 10px; border-radius: 999px; font-size: 11px; font-weight: 600; text-transform: uppercase; }
    .status-badge.ok { background: rgba(59, 165, 92, 0.2); color: var(--success); }
    .status-badge.revisao { background: rgba(250, 166, 26, 0.2); color: var(--warning); }
    .status-badge.erro { background: rgba(237, 66, 69, 0.2); color: var(--error); }
    .status-badge.not_run { background: rgba(204, 204, 204, 0.2); color: var(--fg-muted); }
    
    .details-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
    .detail-section { background: var(--bg-card); border: 1px solid var(--border); border-radius: var(--radius); padding: 16px; }
    .detail-section h3 { margin: 0 0 12px; font-size: 14px; color: var(--accent); }
    .detail-row { display: flex; justify-content: space-between; padding: 6px 0; border-bottom: 1px solid var(--border); }
    .detail-row:last-child { border-bottom: none; }
    .detail-label { color: var(--fg-muted); }
    .detail-value { font-family: monospace; word-break: break-all; }
    
    .chart-container { height: 300px; }
    
    .search-box { width: 100%; padding: 10px 14px; border: 1px solid var(--border); border-radius: var(--radius); background: var(--bg-input); color: var(--fg-primary); font: inherit; margin-bottom: 12px; }
    .search-box:focus { outline: none; border-color: var(--accent); }
    
    .footer { margin-top: 32px; padding-top: 16px; border-top: 1px solid var(--border); color: var(--fg-muted); font-size: 12px; text-align: center; }
    """
    
    JS = """
    // Tab switching
    document.querySelectorAll('.tab').forEach(btn => {
        btn.addEventListener('click', () => {
            document.querySelectorAll('.tab').forEach(b => b.classList.remove('active'));
            document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
            btn.classList.add('active');
            document.getElementById('panel-' + btn.dataset.tab).classList.add('active');
        });
    });
    
    // Table search
    const searchInput = document.getElementById('table-search');
    if (searchInput) {
        searchInput.addEventListener('input', (e) => {
            const term = e.target.value.toLowerCase();
            document.querySelectorAll('#results-table tbody tr').forEach(row => {
                const text = row.textContent.toLowerCase();
                row.style.display = text.includes(term) ? '' : 'none';
            });
        });
    }
    
    // Sortable headers
    document.querySelectorAll('#results-table th').forEach(th => {
        th.style.cursor = 'pointer';
        th.addEventListener('click', () => {
            const table = th.closest('table');
            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const colIndex = Array.from(th.parentNode.children).indexOf(th);
            const isAsc = th.dataset.sort !== 'asc';
            
            document.querySelectorAll('#results-table th').forEach(h => h.dataset.sort = '');
            th.dataset.sort = isAsc ? 'asc' : 'desc';
            
            rows.sort((a, b) => {
                const aVal = a.children[colIndex].textContent.trim();
                const bVal = b.children[colIndex].textContent.trim();
                const cmp = aVal.localeCompare(bVal, undefined, {numeric: true});
                return isAsc ? cmp : -cmp;
            });
            
            rows.forEach(r => tbody.appendChild(r));
        });
    });
    
    // Status chart
    const chartCanvas = document.getElementById('status-chart');
    if (chartCanvas && typeof Chart !== 'undefined') {
        const data = JSON.parse(chartCanvas.dataset.counts || '{}');
        new Chart(chartCanvas, {
            type: 'doughnut',
            data: {
                labels: Object.keys(data),
                datasets: [{
                    data: Object.values(data),
                    backgroundColor: ['#3BA55C', '#FAA61A', '#ED4245', '#CCCCCC'],
                    borderWidth: 0
                }]
            },
            options: {
                responsive: true,
                maintainAspectRatio: true,
                plugins: { legend: { position: 'bottom', labels: { color: '#CCCCCC', font: { size: 12 } } } },
                cutout: '60%'
            }
        });
    }
    
    // Row click - show details in sidebar or modal
    document.querySelectorAll('#results-table tbody tr').forEach(row => {
        row.style.cursor = 'pointer';
        row.addEventListener('click', () => {
            document.querySelectorAll('#results-table tbody tr').forEach(r => r.classList.remove('selected'));
            row.classList.add('selected');
            // Could show details in a side panel
        });
    });
    """
    
    HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ValidaAI - Relatório de Validação {{generated_at}}</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
    <style>{{css}}</style>
</head>
<body>
    <div class="container">
        <header>
            <h1>ValidaAI — Relatório de Validação</h1>
            <div class="meta">
                <span>Gerado em: {{generated_at}}</span>
                <span>Versão: {{app_version}}</span>
                <span>Etapa: {{etapa}}</span>
            </div>
        </header>
        
        <div class="summary-grid">
            <div class="stat-card stat-ok">
                <div class="stat-value">{{summary.ok}}</div>
                <div class="stat-label">OK</div>
            </div>
            <div class="stat-card stat-revisao">
                <div class="stat-value">{{summary.revisao}}</div>
                <div class="stat-label">REVISAO</div>
            </div>
            <div class="stat-card stat-erro">
                <div class="stat-value">{{summary.erro}}</div>
                <div class="stat-label">ERRO</div>
            </div>
            <div class="stat-card stat-notrun">
                <div class="stat-value">{{summary.not_run}}</div>
                <div class="stat-label">NOT_RUN</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{{summary.total}}</div>
                <div class="stat-label">TOTAL</div>
            </div>
        </div>
        
        <div class="tabs">
            <button class="tab active" data-tab="tabela">📋 Tabela</button>
            <button class="tab" data-tab="grafico">📊 Gráfico</button>
            <button class="tab" data-tab="detalhes">🔍 Detalhes</button>
        </div>
        
        <!-- Tabela -->
        <div id="panel-tabela" class="tab-panel active">
            <input type="text" id="table-search" class="search-box" placeholder="🔍 Filtrar testes...">
            <div class="table-wrapper">
                <table id="results-table">
                    <thead>
                        <tr>
                            <th>Teste</th>
                            <th>Bloco</th>
                            <th>Itens</th>
                            <th>Pagamento</th>
                            <th>Subtotal</th>
                            <th>Desc.</th>
                            <th>Total</th>
                            <th>Status</th>
                            <th>Motivo</th>
                            <th>Alertas</th>
                            <th>API Status</th>
                        </tr>
                    </thead>
                    <tbody>
                        {{table_rows}}
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- Gráfico -->
        <div id="panel-grafico" class="tab-panel">
            <div class="chart-container">
                <canvas id="status-chart" data-counts='{{chart_data}}'></canvas>
            </div>
            <div class="details-grid">
                <div class="detail-section">
                    <h3>Por Etapa</h3>
                    {{by_etapa}}
                </div>
                <div class="detail-section">
                    <h3>Estatísticas</h3>
                    <div class="detail-row"><span class="detail-label">Com JSON do parceiro</span><span class="detail-value">{{summary.with_partner_json}}</span></div>
                    <div class="detail-row"><span class="detail-label">Taxa de OK</span><span class="detail-value">{{summary.ok_rate}}%</span></div>
                    <div class="detail-row"><span class="detail-label">Taxa de REVISAO</span><span class="detail-value">{{summary.revisao_rate}}%</span></div>
                </div>
            </div>
        </div>
        
        <!-- Detalhes -->
        <div id="panel-detalhes" class="tab-panel">
            <p style="color: var(--fg-muted);">Clique em uma linha na aba Tabela para ver detalhes.</p>
            <div id="test-detail" class="detail-section" style="display: none;">
                <h3 id="detail-title">Detalhes do Teste</h3>
                <div id="detail-content"></div>
            </div>
        </div>
        
        <footer class="footer">
            ValidaAI v{{app_version}} — Relatório gerado automaticamente em {{generated_at}}
        </footer>
    </div>
    <script>{{js}}</script>
</body>
</html>"""
    
    @property
    def extension(self) -> str:
        return ".html"
    
    def export(
        self, 
        results: List[Dict[str, Any]], 
        output_path: Path,
        metadata: Optional[Dict[str, Any]] = None
    ) -> ExportResult:
        import time
        start = time.perf_counter()
        
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Compute data for template
            summary = self._compute_summary(results)
            chart_data = json.dumps(summary["by_status"])
            
            # Generate table rows
            table_rows = []
            for r in results:
                status = r.get("status_final", "UNKNOWN")
                status_class = status.lower().replace("_pagamento", "").replace("pagamento", "")
                if status_class not in ["ok", "revisao", "erro", "not_run"]:
                    status_class = "unknown"
                
                # Truncate long fields
                motivo = r.get("motivo_status", "")[:80]
                alertas = r.get("alertas", "")
                if isinstance(alertas, list):
                    alertas = "; ".join(alertas)
                alertas = alertas[:80]
                
                row = f"""<tr>
    <td>{r.get('teste', '')}</td>
    <td>{r.get('bloco', r.get('bloco_atual', ''))}</td>
    <td>{str(r.get('itens_raw', r.get('itens_da_venda', '')))[:60]}</td>
    <td>{r.get('pagamento_label', '') or r.get('pagamento_raw', '')}</td>
    <td>{r.get('subtotal_norm', r.get('subtotal_esperado', ''))}</td>
    <td>{r.get('desconto_norm', r.get('desconto_esperado', ''))}</td>
    <td>{r.get('total_norm', r.get('total_esperado', ''))}</td>
    <td><span class="status-badge {status_class}">{status}</span></td>
    <td>{motivo}</td>
    <td>{alertas}</td>
    <td>{r.get('api_status', '')}</td>
</tr>"""
                table_rows.append(row)
            
            # By etapa HTML
            by_etapa_html = ""
            for etapa, counts in summary.get("by_etapa", {}).items():
                by_etapa_html += f"""
                <div class="detail-row">
                    <span class="detail-label">{etapa}</span>
                    <span class="detail-value">OK: {counts.get('OK', 0)} | REVISAO: {counts.get('REVISAO', 0)} | ERRO: {counts.get('ERRO', 0) + counts.get('ERRO_PAGAMENTO', 0)} | NOT_RUN: {counts.get('NOT_RUN', 0)}</span>
                </div>"""
            
            # Rates
            total = summary.get("total", 1)
            ok_rate = round(summary.get("ok", 0) / total * 100, 1)
            revisao_rate = round(summary.get("revisao", 0) / total * 100, 1)
            
            # Render template
            html = self.HTML_TEMPLATE.replace("{{css}}", self.CSS)\
                .replace("{{js}}", self.JS)\
                .replace("{{generated_at}}", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))\
                .replace("{{app_version}}", metadata.get("app_version", "2.0.0") if metadata else "2.0.0")\
                .replace("{{etapa}}", metadata.get("etapa", "Todas") if metadata else "Todas")\
                .replace("{{table_rows}}", "\n".join(table_rows))\
                .replace("{{chart_data}}", chart_data)\
                .replace("{{by_etapa}}", by_etapa_html)\
                .replace("{{summary.ok}}", str(summary.get("ok", 0)))\
                .replace("{{summary.revisao}}", str(summary.get("revisao", 0)))\
                .replace("{{summary.erro}}", str(summary.get("erro", 0) + summary.get("erro_pagamento", 0)))\
                .replace("{{summary.not_run}}", str(summary.get("not_run", 0)))\
                .replace("{{summary.total}}", str(summary.get("total", 0)))\
                .replace("{{summary.with_partner_json}}", str(summary.get("with_partner_json", 0)))\
                .replace("{{summary.ok_rate}}", str(ok_rate))\
                .replace("{{summary.revisao_rate}}", str(revisao_rate))
            
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(html)
            
            duration = int((time.perf_counter() - start) * 1000)
            return ExportResult(
                exporter_name=self.name,
                output_path=output_path,
                success=True,
                rows_exported=len(results),
                duration_ms=duration
            )
            
        except Exception as e:
            return ExportResult(
                exporter_name=self.name,
                output_path=output_path,
                success=False,
                error=str(e)
            )
    
    def _compute_summary(self, results: List[Dict]) -> Dict[str, Any]:
        from collections import Counter
        status_counts = Counter(r.get("status_final", "UNKNOWN") for r in results)
        
        by_etapa = {}
        for r in results:
            etapa = r.get("bloco", r.get("bloco_atual", "UNKNOWN"))
            if etapa not in by_etapa:
                by_etapa[etapa] = Counter()
            by_etapa[etapa][r.get("status_final", "UNKNOWN")] += 1
        
        return {
            "by_status": dict(status_counts),
            "by_etapa": {k: dict(v) for k, v in by_etapa.items()},
            "total": len(results),
            "ok": status_counts.get("OK", 0),
            "revisao": status_counts.get("REVISAO", 0),
            "erro": status_counts.get("ERRO", 0),
            "erro_pagamento": status_counts.get("ERRO_PAGAMENTO", 0),
            "not_run": status_counts.get("NOT_RUN", 0),
            "with_partner_json": sum(1 for r in results if r.get("sale_json")),
        }


# ═════════════════════════════════════════════════════════════
# Notification System
# ═════════════════════════════════════════════════════════════

class Notifier(ABC):
    """Abstract base class for notification channels."""
    
    @abstractmethod
    def send(
        self, 
        subject: str, 
        message: str, 
        attachments: Optional[List[Path]] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> bool:
        """Send notification. Returns True on success."""
        pass


class EmailNotifier(Notifier):
    """Email notification via SMTP."""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.enabled = config.get("enabled", False)
        self.smtp_host = config.get("smtp_host", "smtp.gmail.com")
        self.smtp_port = config.get("smtp_port", 587)
        self.username = config.get("username", "")
        self.password = config.get("password", "")
        self.from_email = config.get("from_email", self.username)
        self.to_emails = config.get("to_emails", [])
        self.subject_template = config.get("subject_template", "ValidaAI Report")
    
    def send(
        self, 
        subject: str, 
        message: str, 
        attachments: Optional[List[Path]] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> bool:
        if not self.enabled:
            return False
        if not self.to_emails:
            print("EmailNotifier: No recipients configured")
            return False
        
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            from email.mime.application import MIMEApplication
            
            msg = MIMEMultipart()
            msg["Subject"] = subject
            msg["From"] = self.from_email
            msg["To"] = ", ".join(self.to_emails)
            
            msg.attach(MIMEText(message, "plain", "utf-8"))
            
            if attachments:
                for filepath in attachments:
                    if filepath.exists():
                        with open(filepath, "rb") as f:
                            part = MIMEApplication(f.read(), Name=filepath.name)
                        part["Content-Disposition"] = f'attachment; filename="{filepath.name}"'
                        msg.attach(part)
            
            with smtplib.SMTP(self.smtp_host, self.smtp_port) as server:
                server.starttls()
                if self.username and self.password:
                    server.login(self.username, self.password)
                server.send_message(msg)
            
            return True
        except Exception as e:
            print(f"EmailNotifier error: {e}")
            return False


class SlackNotifier(Notifier):
    """Slack notification via webhook."""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.enabled = config.get("enabled", False)
        self.webhook_url = config.get("webhook_url", "")
        self.channel = config.get("channel", "#general")
        self.username = config.get("username", "ValidaAI Bot")
        self.icon_emoji = config.get("icon_emoji", ":robot_face:")
        self.message_template = config.get("message_template", "")
    
    def send(
        self, 
        subject: str, 
        message: str, 
        attachments: Optional[List[Path]] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> bool:
        if not self.enabled:
            return False
        if not self.webhook_url:
            print("SlackNotifier: No webhook URL configured")
            return False
        
        try:
            import urllib.request
            import json
            
            payload = {
                "channel": self.channel,
                "username": self.username,
                "icon_emoji": self.icon_emoji,
                "text": message or self._format_message(metadata)
            }
            
            req = urllib.request.Request(
                self.webhook_url,
                data=json.dumps(payload).encode("utf-8"),
                headers={"Content-Type": "application/json"}
            )
            urllib.request.urlopen(req, timeout=10)
            return True
        except Exception as e:
            print(f"SlackNotifier error: {e}")
            return False
    
    def _format_message(self, metadata: Optional[Dict[str, Any]]) -> str:
        if not metadata:
            return "ValidaAI validation completed."
        
        template = self.message_template or "*ValidaAI* validation completed."
        try:
            return template.format(
                etapa=metadata.get("etapa", "N/A"),
                timestamp=metadata.get("generated_at", "N/A"),
                total=metadata.get("total", 0),
                ok=metadata.get("summary", {}).get("ok", 0),
                revisao=metadata.get("summary", {}).get("revisao", 0),
                erro=metadata.get("summary", {}).get("erro", 0)
            )
        except Exception:
            return template


# ═════════════════════════════════════════════════════════════
# ═════════════════════════════════════════════════════════════
# Pipeline Orchestrator
# ══════════════════════════════════════════════════════════════

EXPORTER_REGISTRY = {
    "excel": ExcelExporter,
    "csv": CSVExporter,
    "json_audit": JSONAuditExporter,
    "html_report": HTMLReportExporter,
}


class ExportPipeline:
    """Orchestrates multiple exporters to run in sequence."""
    
    def __init__(self, config: ExportConfig):
        self.config = config
        self._exporters: List[Exporter] = []
        self._build_exporters()
    
    def _build_exporters(self) -> None:
        for name in self.config.exporters:
            exporter_class = EXPORTER_REGISTRY.get(name)
            if exporter_class:
                exporter_config = self.config.get_exporter_config(name)
                self._exporters.append(exporter_class(exporter_config))
            else:
                print(f"Warning: Unknown exporter '{name}'")
    
    def __init__(self, config: ExportConfig):
        self.config = config
        self._exporters: List[Exporter] = []
        self._notifiers: List[Notifier] = []
        self._build_exporters()
        self._build_notifiers()
    
    def _build_exporters(self) -> None:
        for name in self.config.exporters:
            exporter_class = EXPORTER_REGISTRY.get(name)
            if exporter_class:
                exporter_config = self.config.get_exporter_config(name)
                self._exporters.append(exporter_class(exporter_config))
            else:
                print(f"Warning: Unknown exporter '{name}'")
    
    def _build_notifiers(self) -> None:
        notif_config = self.config.notifications
        if not notif_config.get("enabled", False):
            return
        
        email_config = notif_config.get("email", {})
        if email_config.get("enabled", False):
            self._notifiers.append(EmailNotifier(email_config))
        
        slack_config = notif_config.get("slack", {})
        if slack_config.get("enabled", False):
            self._notifiers.append(SlackNotifier(slack_config))
    
    def _notify(
        self, 
        subject: str, 
        message: str, 
        attachments: Optional[List[Path]] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> None:
        for notifier in self._notifiers:
            try:
                notifier.send(subject, message, attachments, metadata)
            except Exception as e:
                print(f"Notifier error: {e}")
    
    def run(
        self, 
        results: List[Dict[str, Any]], 
        base_path: Path,
        metadata: Optional[Dict[str, Any]] = None
    ) -> List[ExportResult]:
        """Run all exporters and return results."""
        export_results = []
        output_files = []
        
        for exporter in self._exporters:
            output_path = base_path.with_suffix(exporter.extension)
            print(f"  Exporting {exporter.name} -> {output_path}")
            
            result = exporter.export(results, output_path, metadata)
            export_results.append(result)
            
            if result.success:
                print(f"    ✓ {result.rows_exported} rows in {result.duration_ms}ms")
                output_files.append(result.output_path)
            else:
                print(f"    ✗ ERROR: {result.error}")
        
        # Send notifications
        if self._notifiers:
            success_count = sum(1 for r in export_results if r.success)
            total_count = len(export_results)
            
            subject = f"ValidaAI Report - {success_count}/{total_count} exports successful"
            message = (
                f"Validation export completed.\n"
                f"Total: {total_count} exports\n"
                f"Successful: {success_count}\n"
                f"Failed: {total_count - success_count}\n"
            )
            
            # Add summary from metadata
            if metadata:
                summary = metadata.get("summary", {})
                if summary:
                    message += f"\nSummary: OK={summary.get('ok',0)} REVISAO={summary.get('revisao',0)} ERRO={summary.get('erro',0)}"
            
            output_paths = [f for f in output_files if f.exists()]
            self._notify(subject, message, output_paths, metadata)
        
        return export_results
    
    def add_exporter(self, exporter: Exporter) -> None:
        """Add a custom exporter to the pipeline."""
        self._exporters.append(exporter)
    
    @classmethod
    def from_config_file(cls, config_path: Path) -> "ExportPipeline":
        """Create pipeline from JSON config file."""
        config = ExportConfig.from_file(config_path)
        return cls(config)


# ════════════════════════════════════════════════════════════
# Convenience Function
# ════════════════════════════════════════════════════════════

def export_all(
    results: List[Dict[str, Any]],
    output_dir: Path,
    timestamp: Optional[str] = None,
    etapa: str = "Todas",
    version: str = "2.0.0",
    exporters: Optional[List[str]] = None,
    metadata: Optional[Dict[str, Any]] = None
) -> List[ExportResult]:
    """
    One-shot export all formats.
    
    Args:
        results: Validation results
        output_dir: Output directory
        timestamp: Timestamp string (default: now)
        etapa: Etapa name for filename
        version: App version
        exporters: List of exporter names (default: all)
        metadata: Additional metadata for exports
    
    Returns:
        List of ExportResult
    """
    if timestamp is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if exporters is None:
        exporters = ["excel", "csv", "json_audit", "html_report"]
    
    config = ExportConfig(exporters=exporters)
    pipeline = ExportPipeline(config)
    
    filename = f"validacao_{timestamp}_{etapa}_v{version}"
    base_path = output_dir / filename
    output_dir.mkdir(parents=True, exist_ok=True)
    
    meta = metadata or {}
    meta.update({
        "app_version": version,
        "etapa": etapa,
        "timestamp": timestamp,
        "generated_at": datetime.now().isoformat(),
    })
    
    return pipeline.run(results, base_path, meta)


if __name__ == "__main__":
    # Demo with sample data
    from pathlib import Path
    
    sample_results = [
        {
            "teste": 1, "bloco": "ETAPA 1", "tipo_promo": "",
            "itens_da_venda": "2 x 7891000010860 + 3.579 * PESABLE",
            "pagamento": "Dinheiro", "codigo_tipo_pago": 9,
            "subtotal_esperado": 149.065, "desconto_esperado": 0, "total_esperado": 149.065,
            "status_final": "REVISAO", "motivo_status": "Item pesável detectado",
            "alertas": ["Item pesável detectado - requer revisão manual"],
            "observacoes": "", "observacao_parceiro": "",
            "sat": "", "ecf": "1", "nfce": "", "cupom": "",
            "api_status": "OK", "api_alertas": [], "sale_json": {}
        },
        {
            "teste": 2, "bloco": "ETAPA 1", "tipo_promo": "",
            "itens_da_venda": "5 x 7894904500383", "pagamento": "Dinheiro com Troco",
            "codigo_tipo_pago": 9, "subtotal_esperado": 16.17, "desconto_esperado": 0, "total_esperado": 16.17,
            "status_final": "REVISAO", "motivo_status": "Pagamento com troco",
            "alertas": ["Pagamento com troco - requer revisão manual"],
            "observacoes": "", "sat": "", "ecf": "2", "nfce": "", "cupom": "",
            "api_status": "OK", "api_alertas": [], "sale_json": {}
        },
        {
            "teste": 3, "bloco": "ETAPA 1", "tipo_promo": "",
            "itens_da_venda": "1 x 7894904003495", "pagamento": "Cartao Credito",
            "codigo_tipo_pago": 10, "subtotal_esperado": 17.28, "desconto_esperado": 0, "total_esperado": 17.28,
            "status_final": "OK", "motivo_status": "Todos os campos válidos",
            "alertas": [], "observacoes": "", "sat": "", "ecf": "3", "nfce": "", "cupom": "",
            "api_status": "OK", "api_alertas": [], "sale_json": {}
        },
    ]
    
    output_dir = Path("output/demo")
    results = export_all(
        sample_results,
        output_dir,
        etapa="ETAPA1",
        version="2.0.0"
    )
    
    print("\nExport Results:")
    for r in results:
        status = "✓" if r.success else "✗"
        print(f"  {status} {r.exporter_name}: {r.output_path} ({r.duration_ms}ms)")