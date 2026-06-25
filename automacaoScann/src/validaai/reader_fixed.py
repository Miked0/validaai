#!/usr/bin/env python3
"""Fixed _read_excel_openpyxl method"""
import openpyxl
from typing import List, Dict, Any

def _read_excel_openpyxl(self) -> List[Dict[str, Any]]:
    """Read test script from Excel file using openpyxl.
    Supports multiple ETAPA sheets and Observacoes.1 column.
    """
    workbook = openpyxl.load_workbook(self.file_path, data_only=True)
    all_tests: List[Dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        # Filter sheets by ETAPA
        if 'ETAPA' not in sheet_name.upper():
            continue
        # Apply etapa_filter if set
        if self.etapa_filter and sheet_name.strip().upper() != self.etapa_filter:
            continue

        ws = workbook[sheet_name]
        header = None
        header_row = None
        start_idx = None
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [str(c).strip() if c is not None else '' for c in row]
            up = [v.upper() for v in vals]
            if 'TESTE' in up and any(k in up for k in ['ITENS DA VENDA', 'ARTICULOS MOVIMIENTO', 'ITENS']):
                header = vals
                header_row = r_idx
                start_idx = r_idx + 1
                break
        if not header:
            continue

        rows = list(ws.iter_rows(min_row=start_idx, values_only=True))
        seen_tests = {}  # t_key -> row data (keep best one with valid cupom)
        
        for offset, row in enumerate(rows, start=start_idx):
            vals = [str(c).strip() if c is not None else '' for c in row]
            if len(vals) < len(header):
                vals += [''] * (len(header) - len(vals))
            elif len(vals) > len(header):
                vals = vals[:len(header)]
            rd = dict(zip(header, vals))
            t_raw = rd.get('Teste', '')
            if t_raw is None or str(t_raw).strip() == '':
                continue
            
            # Handle both explicit numbers and Excel formulas (e.g., =A8+1)
            t_str = str(t_raw).strip()
            is_formula = t_str.startswith('=')
            try:
                t_num = float(t_str) if not is_formula else None
            except ValueError:
                t_num = None
            
            # Use explicit number if available, otherwise assign sequential
            if t_num is not None:
                t_key = str(int(t_num) if t_num == int(t_num) else t_num)
            else:
                # For formula rows, assign next sequential number
                last_num = max([int(k) for k in seen_tests if k.isdigit()], default=0)
                t_key = str(last_num + 1)
            
            # Check if we already have this test number - keep the one with valid cupom
            if t_key in seen_tests:
                # Check if existing one has valid cupom, if not replace with this one
                existing_rd = seen_tests[t_key]
                existing_nfce = _first_nonempty(list(existing_rd.values()), list(existing_rd.keys()), nfce_key)
                existing_sat = _first_nonempty(list(existing_rd.values()), list(existing_rd.keys()), sat_key)
                existing_ecf = _first_nonempty(list(existing_rd.values()), list(existing_rd.keys()), ecf_key)
                existing_cupom = _first_nonempty(list(existing_rd.values()), list(existing_rd.keys()), cupom_key)
                
                # Check if this row has better cupom data
                this_nfce = _first_nonempty(vals, header, nfce_key)
                this_sat = _first_nonempty(vals, header, sat_key)
                this_ecf = _first_nonempty(vals, header, ecf_key)
                this_cupom = _first_nonempty(vals, header, cupom_key)
                
                existing_has_cupom = any(v and v.strip() for v in [existing_nfce, existing_sat, existing_ecf, existing_cupom])
                this_has_cupom = any(v and v.strip() for v in [this_nfce, this_sat, this_ecf, this_cupom])
                
                if not existing_has_cupom and this_has_cupom:
                    seen_tests[t_key] = rd  # Replace with better row
                # else keep existing
                continue
            seen_tests[t_key] = rd
            
            # Find column keys (case-insensitive)
            subtotal_key = next((k for k in rd if k and k.lower() in ['sub-total', 'subtotal']), 'Sub-Total')
            total_key = next((k for k in rd if k and k.lower() in ['total']), 'Total')
            desconto_key = next((k for k in rd if k and k.lower() in ['desconto']), 'Desconto')
            
            # Also look for alternative column names for ETAPA 2
            itens_key = next((k for k in rd if k and k.lower() in ['itens da venda', 'articulos movimiento', 'itens']), 'Itens da venda')
            pagamento_key = next((k for k in rd if k and k.lower() in ['pagamento']), 'Pagamento')
            observacoes_key = next((k for k in rd if k and k.lower() in ['observacoes']), 'Observacoes')
            observacao_parceiro_key = next((k for k in rd if k and k.lower() in ['observacoes.1']), 'Observacoes.1')
            tipo_promo_key = next((k for k in rd if k and k.lower() in ['tipo promo', 'tipo promocao', 'tipo promo', 'promotion_type', 'tipo']), 'Tipo Promo')
            
            # Cupom fields - priority: NFCE (under "Numero de cupom") > SAT > ECF > explicit Cupom columns
            nfce_key = next((k for k in rd if k and k.lower() in ['nfce', 'nfc-e']), 'NFCE')
            sat_key = next((k for k in rd if k and k.lower() in ['sat']), 'SAT')
            ecf_key = next((k for k in rd if k and k.lower() in ['ecf']), 'ECF')
            cupom_key = next((k for k in rd if k and k.lower() in ['cupom', 'numero cupom', 'numero do cupom']), 'Cupom')
            json_key = next((k for k in rd if k and k.lower() in ['json']), 'Json')
            minoristas_key = next((k for k in rd if k and k.lower() in ['minoristas']), 'Minoristas')
            
            def _first_nonempty(vals_list, header_list, key):
                matches = []
                for i, h in enumerate(header_list):
                    if h and h.strip().lower() == key.strip().lower():
                        v = vals_list[i] if i < len(vals_list) else ''
                        if v and str(v).strip():
                            val = str(v).strip()
                            # Skip placeholder values like "(Status)", "(status)", "Status", etc.
                            if val.lower() in ['(status)', 'status', '(pendente)', 'pendente', '(aguardando)', 'aguardando']:
                                continue
                            matches.append(val)
                return matches[-1] if matches else ''

            # Cupom: priority NFCE (under "Numero de cupom") > SAT > ECF > explicit Cupom
            cupom_val = (_first_nonempty(vals, header, nfce_key) or 
                        _first_nonempty(vals, header, sat_key) or 
                        _first_nonempty(vals, header, ecf_key) or 
                        _first_nonempty(vals, header, cupom_key))
            
            std = {
                'teste': float(t_key) if '.' in t_key else int(t_key),
                'linha_original': f"{sheet_name}!{offset}",
                'bloco_atual': sheet_name.strip().upper(),
                'tipo_promo': rd.get(tipo_promo_key, ''),
                'itens_da_venda': rd.get(itens_key, ''),
                'pagamento': rd.get(pagamento_key, ''),
                'observacoes': _first_nonempty(vals, header, observacoes_key),
                'observacao_parceiro': _first_nonempty(vals, header, observacao_parceiro_key),
                'subtotal_esperado': rd.get(subtotal_key, ''),
                'desconto_esperado': rd.get(desconto_key, '0') or '0',
                'total_esperado': rd.get(total_key, ''),
                'sat': _first_nonempty(vals, header, sat_key),
                'ecf': _first_nonempty(vals, header, ecf_key),
                'nfce': _first_nonempty(vals, header, nfce_key),
                'json': rd.get(json_key, ''),
                'minoristas': rd.get(minoristas_key, ''),
                # Cupom: priority NFCE > SAT > ECF > explicit Cupom
                'cupom': cupom_val,
                # Include product catalog for downstream use
                'product_catalog': self.product_catalog,
            }
            all_tests.append(std)

    try:
        workbook.close()
    except Exception:
        pass

    return all_tests