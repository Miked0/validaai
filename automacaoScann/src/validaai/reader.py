#!/usr/bin/env python3
"""
Test Script Reader Module
Responsible for reading and parsing the test script spreadsheet.
Supports ETAPA sheets and partner observation column (Observacoes.1).
"""

import csv
import re
from typing import List, Dict, Any, Optional
from pathlib import Path

# Try to import pandas, but have a fallback
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("WARNING: Pandas not available. Using CSV-only mode for reading test scripts.")

class TestScriptReader:
    """Reads test script spreadsheets and extracts test cases."""

    def __init__(self, file_path: str):
        """
        Initialize the reader.

        Args:
            file_path: Path to the Excel or CSV file
        """
        self.file_path = file_path
        self.data = None
        self.tests = []
        self.etapa_filter: Optional[str] = None
        # Product catalog from "Produtos a cadastrar" sheet
        self.product_catalog: Dict[str, Dict[str, Any]] = {}

    def set_etapa(self, etapa: Optional[str]) -> None:
        """
        Set the ETAPA filter to read only a specific stage.
        
        Args:
            etapa: ETAPA name (e.g., 'ETAPA 1', 'ETAPA 2') or None for all
        """
        self.etapa_filter = etapa.strip().upper() if etapa else None

    def read_tests(self) -> List[Dict[str, Any]]:
        """
        Public method to read tests from the spreadsheet.
        """
        return self._read_excel_openpyxl()

    def _load_product_catalog(self) -> None:
        """
        Load product catalog from 'Produtos a cadastrar' sheet.
        Expected columns: EAN, Valor (preço), Descrição (opcional).
        Populates self.product_catalog with EAN as key.
        """
        import openpyxl
        workbook = openpyxl.load_workbook(self.file_path)
        
        for sheet_name in workbook.sheetnames:
            # Look for "Produtos a cadastrar" sheet (case insensitive)
            if 'PRODUTO' not in sheet_name.upper() and 'CADASTRAR' not in sheet_name.upper():
                continue
            
            ws = workbook[sheet_name]
            header = None
            header_row = None
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                vals = [str(c).strip() if c is not None else '' for c in row]
                up = [v.upper() for v in vals]
                # Detect header: look for EAN + Valor/Preço
                if 'EAN' in up and any(k in up for k in ['VALOR', 'PRECO', 'PREÇO', 'PRICE']):
                    header = vals
                    header_row = r_idx
                    break
            if not header:
                continue
            
            # Map column indices
            ean_idx = None
            valor_idx = None
            desc_idx = None
            for i, h in enumerate(header):
                h_up = h.upper()
                if h_up == 'EAN' or h_up == 'CODIGO' or h_up == 'CÓDIGO':
                    ean_idx = i
                elif any(k in h_up for k in ['VALOR', 'PRECO', 'PREÇO', 'PRICE', 'PREÇO UNIT', 'PRECO UNIT']):
                    valor_idx = i
                elif any(k in h_up for k in ['DESCRICAO', 'DESCRIÇÃO', 'DESC', 'PRODUTO', 'NOME']):
                    desc_idx = i
            
            if ean_idx is None or valor_idx is None:
                continue
            
            rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
            for row in rows:
                if len(row) <= max(ean_idx, valor_idx):
                    continue
                ean = str(row[ean_idx]).strip() if row[ean_idx] else ''
                valor_raw = row[valor_idx]
                desc = str(row[desc_idx]).strip() if desc_idx is not None and len(row) > desc_idx and row[desc_idx] else ''
                
                if not ean:
                    continue
                
                # Parse valor
                try:
                    if isinstance(valor_raw, (int, float)):
                        valor = float(valor_raw)
                    else:
                        valor_str = str(valor_raw).strip().replace(',', '.')
                        valor = float(valor_str) if valor_str else 0.0
                except (ValueError, TypeError):
                    valor = 0.0
                
                self.product_catalog[ean] = {
                    'preco': valor,
                    'descricao': desc
                }
        
        try:
            workbook.close()
        except Exception:
            pass

    def _read_excel(self) -> List[Dict[str, Any]]:
        """
        Read the test script and extract test cases.

        Returns:
            List of dictionaries representing raw test cases
        """
        # First, load product catalog from "Produtos a cadastrar" sheet
        self._load_product_catalog()
        
        # Read the file based on extension and available libraries
        file_ext = Path(self.file_path).suffix.lower()
        
        if file_ext == '.xlsx':
            # Use openpyxl directly (works without pandas/tkinter)
            return self._read_excel_openpyxl()
        elif file_ext == '.csv':
            return self._read_csv()
        else:
            raise ValueError(f"Unsupported file extension: {file_ext}")

    def _read_excel(self) -> List[Dict[str, Any]]:
        """Read test script from Excel file using pandas.
        Supports multiple ETAPA sheets and Observacoes.1 column.
        """
        import openpyxl
        workbook = openpyxl.load_workbook(self.file_path)
        all_tests: List[Dict[str, Any]] = []

        for sheet_name in workbook.sheetnames:
            # Filter sheets by ETAPA
            if 'ETAPA' not in sheet_name.upper():
                continue
            # Apply etapa_filter if set
            if self.etapa_filter and sheet_name.strip().upper() != self.etapa_filter:
                continue

            ws = workbook[sheet_name]
            header = None
            header_row = None
            start_idx = None
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                vals = [str(c).strip() if c is not None else '' for c in row]
                up = [v.upper() for v in vals]
                if 'TESTE' in up and any(k in up for k in ['ITENS DA VENDA', 'ARTICULOS MOVIMIENTO', 'ITENS']):
                    header = vals
                    header_row = r_idx
                    start_idx = r_idx + 1
                    break
            if not header:
                continue

            # Get column indices for key fields (use first occurrence of each)
            col_indices = {}
            for i, h in enumerate(header):
                h_lower = h.strip().lower() if h else ''
                if h_lower and h_lower not in col_indices:
                    col_indices[h_lower] = i
            
            # Map keys to indices
            def get_col_idx(*possible_names):
                for name in possible_names:
                    if name.lower() in col_indices:
                        return col_indices[name.lower()]
                return None
            
            nfce_idx = get_col_idx('nfce', 'nfc-e')
            sat_idx = get_col_idx('sat')
            ecf_idx = get_col_idx('ecf')
            cupom_idx = get_col_idx('cupom', 'numero cupom', 'numero do cupom')
            json_idx = get_col_idx('json')
            minoristas_idx = get_col_idx('minoristas')
            itens_idx = get_col_idx('itens da venda', 'articulos movimiento', 'itens')
            pagamento_idx = get_col_idx('pagamento')
            observacoes_idx = get_col_idx('observacoes')
            observacao_parceiro_idx = get_col_idx('observacoes.1')
            tipo_promo_idx = get_col_idx('tipo promo', 'tipo promocao', 'tipo promo', 'promotion_type', 'tipo')
            subtotal_idx = get_col_idx('sub-total', 'subtotal')
            total_idx = get_col_idx('total')
            desconto_idx = get_col_idx('desconto')
            
            rows = list(ws.iter_rows(min_row=start_idx, values_only=True))
            seen_tests = {}  # t_key -> row data (keep best one with valid cupom)

            def _get_val(vals_list, idx):
                """Get value from list by index, return '' if out of bounds."""
                if idx is not None and idx < len(vals_list):
                    v = vals_list[idx]
                    if v and str(v).strip():
                        val = str(v).strip()
                        # Skip placeholder values
                        if val.lower() in ['(status)', 'status', '(pendente)', 'pendente', '(aguardando)', 'aguardando', 'none', 'null']:
                            return ''
                        return val
                return ''

            # Reset rows iterator
            rows = list(ws.iter_rows(min_row=start_idx, values_only=True))

            for offset, row in enumerate(rows, start=start_idx):
                vals = [str(c).strip() if c is not None else '' for c in row]
                if len(vals) < len(header):
                    vals += [''] * (len(header) - len(vals))
                elif len(vals) > len(header):
                    vals = vals[:len(header)]
                rd = dict(zip(header, vals))
                t_raw = rd.get('Teste', '')
                if t_raw is None or str(t_raw).strip() == '':
                    continue
                
                # Handle both explicit numbers and Excel formulas (e.g., =A8+1)
                t_str = str(t_raw).strip()
                is_formula = t_str.startswith('=')
                try:
                    t_num = float(t_str) if not is_formula else None
                except ValueError:
                    t_num = None
                
                # Use explicit number if available, otherwise assign sequential
                if t_num is not None:
                    t_key = str(int(t_num) if t_num == int(t_num) else t_num)
                else:
                    # For formula rows, assign next sequential number
                    # Find the last assigned number
                    last_num = max([int(k) for k in seen_tests if k.isdigit()], default=0)
                    t_key = str(last_num + 1)
                
                if t_key in seen_tests:
                    # Check if existing one has valid cupom, if not replace with this one
                    existing_vals = seen_tests[t_key]
                    # Use the new index-based extraction
                    existing_nfce = _get_val(existing_vals, nfce_idx)
                    existing_sat = _get_val(existing_vals, sat_idx)
                    existing_ecf = _get_val(existing_vals, ecf_idx)
                    existing_cupom = _get_val(existing_vals, cupom_idx)
                    
                    # Check if this row has better cupom data
                    this_nfce = _get_val(vals, nfce_idx)
                    this_sat = _get_val(vals, sat_idx)
                    this_ecf = _get_val(vals, ecf_idx)
                    this_cupom = _get_val(vals, cupom_idx)
                    
                    existing_has_cupom = any(v and v.strip() for v in [existing_nfce, existing_sat, existing_ecf, existing_cupom])
                    this_has_cupom = any(v and v.strip() for v in [this_nfce, this_sat, this_ecf, this_cupom])
                    
                    if not existing_has_cupom and this_has_cupom:
                        seen_tests[t_key] = vals  # Replace with better row (store vals list)
                    # else keep existing
                    continue
                seen_tests[t_key] = vals  # Store vals list (maintains column order)
                # Find column keys (case-insensitive)
                subtotal_key = next((k for k in rd if k and k.lower() in ['sub-total', 'subtotal']), 'Sub-Total')
                total_key = next((k for k in rd if k and k.lower() in ['total']), 'Total')
                desconto_key = next((k for k in rd if k and k.lower() in ['desconto']), 'Desconto')
                
                # Also look for alternative column names for ETAPA 2
                itens_key = next((k for k in rd if k and k.lower() in ['itens da venda', 'articulos movimiento', 'itens']), 'Itens da venda')
                pagamento_key = next((k for k in rd if k and k.lower() in ['pagamento']), 'Pagamento')
                observacoes_key = next((k for k in rd if k and k.lower() in ['observacoes']), 'Observacoes')
                observacao_parceiro_key = next((k for k in rd if k and k.lower() in ['observacoes.1']), 'Observacoes.1')
                tipo_promo_key = next((k for k in rd if k and k.lower() in ['tipo promo', 'tipo promocao', 'tipo promo', 'promotion_type', 'tipo']), 'Tipo Promo')
                
                # Cupom fields - priority: NFCE (under "Numero de cupom") > SAT > ECF > explicit Cupom columns
                nfce_key = next((k for k in rd if k and k.lower() in ['nfce', 'nfc-e']), 'NFCE')
                sat_key = next((k for k in rd if k and k.lower() in ['sat']), 'SAT')
                ecf_key = next((k for k in rd if k and k.lower() in ['ecf']), 'ECF')
                cupom_key = next((k for k in rd if k and k.lower() in ['cupom', 'numero cupom', 'numero do cupom']), 'Cupom')
                json_key = next((k for k in rd if k and k.lower() in ['json']), 'Json')
                minoristas_key = next((k for k in rd if k and k.lower() in ['minoristas']), 'Minoristas')
                
                # Cupom: priority NFCE (under "Numero de cupom") > SAT > ECF > explicit Cupom
                cupom_val = (_get_val(vals, nfce_idx) or 
                            _get_val(vals, sat_idx) or 
                            _get_val(vals, ecf_idx) or 
                            _get_val(vals, cupom_idx))
                
                # Also get individual fields using index-based extraction for consistency
                sat_val = _get_val(vals, sat_idx)
                ecf_val = _get_val(vals, ecf_idx)
                nfce_val = _get_val(vals, nfce_idx)
                json_val = _get_val(vals, json_idx)
                minoristas_val = _get_val(vals, minoristas_idx)
                
                std = {
                    'teste': float(t_key) if '.' in t_key else int(t_key),
                    'linha_original': f"{sheet_name}!{offset}",
                    'bloco_atual': sheet_name.strip().upper(),
                    'tipo_promo': rd.get(tipo_promo_key, ''),
                    'itens_da_venda': _get_val(vals, itens_idx),
                    'pagamento': _get_val(vals, pagamento_idx),
                    'observacoes': _get_val(vals, observacoes_idx),
                    'observacao_parceiro': _get_val(vals, observacao_parceiro_idx),
                    'subtotal_esperado': _get_val(vals, subtotal_idx),
                    'desconto_esperado': _get_val(vals, desconto_idx) or '0',
                    'total_esperado': _get_val(vals, total_idx),
                    'sat': sat_val,
                    'ecf': ecf_val,
                    'nfce': nfce_val,
                    'json': json_val,
                    'minoristas': minoristas_val,
                    # Cupom: priority NFCE > SAT > ECF > explicit Cupom
                    'cupom': cupom_val,
                    # Include product catalog for downstream use
                    'product_catalog': self.product_catalog,
                }
                all_tests.append(std)

        try:
            workbook.close()
        except Exception:
            pass

        return all_tests

    def _read_excel_openpyxl(self) -> List[Dict[str, Any]]:
        """Read test script from Excel file using openpyxl.
        Supports multiple ETAPA sheets and Observacoes.1 column.
        """
        import openpyxl
        # Use data_only=True to get calculated formula values
        workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        all_tests: List[Dict[str, Any]] = []

        for sheet_name in workbook.sheetnames:
            # Filter sheets by ETAPA
            if 'ETAPA' not in sheet_name.upper():
                continue
            # Apply etapa_filter if set
            if self.etapa_filter and sheet_name.strip().upper() != self.etapa_filter:
                continue

            ws = workbook[sheet_name]
            header = None
            header_row = None
            start_idx = None
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                vals = [str(c).strip() if c is not None else '' for c in row]
                up = [v.upper() for v in vals]
                if 'TESTE' in up and any(k in up for k in ['ITENS DA VENDA', 'ARTICULOS MOVIMIENTO', 'ITENS']):
                    header = vals
                    header_row = r_idx
                    start_idx = r_idx + 1
                    break
            if not header:
                continue

            # Get column indices for key fields (use first occurrence of each)
            col_indices = {}
            for i, h in enumerate(header):
                h_lower = h.strip().lower() if h else ''
                if h_lower and h_lower not in col_indices:
                    col_indices[h_lower] = i
            
            # Map keys to indices
            def get_col_idx(*possible_names):
                for name in possible_names:
                    if name.lower() in col_indices:
                        return col_indices[name.lower()]
                return None
            
            nfce_idx = get_col_idx('nfce', 'nfc-e')
            sat_idx = get_col_idx('sat')
            ecf_idx = get_col_idx('ecf')
            cupom_idx = get_col_idx('cupom', 'numero cupom', 'numero do cupom')
            json_idx = get_col_idx('json')
            minoristas_idx = get_col_idx('minoristas')
            itens_idx = get_col_idx('itens da venda', 'articulos movimiento', 'itens')
            pagamento_idx = get_col_idx('pagamento')
            observacoes_idx = get_col_idx('observacoes')
            observacao_parceiro_idx = get_col_idx('observacoes.1')
            tipo_promo_idx = get_col_idx('tipo promo', 'tipo promocao', 'tipo promo', 'promotion_type', 'tipo')
            subtotal_idx = get_col_idx('sub-total', 'subtotal')
            total_idx = get_col_idx('total')
            desconto_idx = get_col_idx('desconto')
            
            rows = list(ws.iter_rows(min_row=start_idx, values_only=True))
            seen_tests = {}  # t_key -> row data (keep best one with valid cupom)

            def _get_val(vals_list, idx):
                """Get value from list by index, return '' if out of bounds."""
                if idx is not None and idx < len(vals_list):
                    v = vals_list[idx]
                    if v and str(v).strip():
                        val = str(v).strip()
                        # Skip placeholder values
                        if val.lower() in ['(status)', 'status', '(pendente)', 'pendente', '(aguardando)', 'aguardando', 'none', 'null']:
                            return ''
                        return val
                return ''

            # Reset rows iterator
            rows = list(ws.iter_rows(min_row=start_idx, values_only=True))

            for offset, row in enumerate(rows, start=start_idx):
                vals = [str(c).strip() if c is not None else '' for c in row]
                if len(vals) < len(header):
                    vals += [''] * (len(header) - len(vals))
                elif len(vals) > len(header):
                    vals = vals[:len(header)]
                rd = dict(zip(header, vals))
                t_raw = rd.get('Teste', '')
                if t_raw is None or str(t_raw).strip() == '':
                    continue
                
                # Handle both explicit numbers and Excel formulas (e.g., =A8+1)
                t_str = str(t_raw).strip()
                is_formula = t_str.startswith('=')
                try:
                    t_num = float(t_str) if not is_formula else None
                except ValueError:
                    t_num = None
                
                # Skip rows where Teste is not a valid number (headers, instructions, etc.)
                if t_num is None and not is_formula:
                    continue
                # Also skip if formula doesn't follow =A... pattern (not sequential test numbering)
                if is_formula and not ('A' in t_str or 'a' in t_str):
                    continue
                
                # Use explicit number if available, otherwise assign sequential
                if t_num is not None:
                    t_key = str(int(t_num) if t_num == int(t_num) else t_num)
                else:
                    # For formula rows, assign next sequential number
                    last_num = max([int(k) for k in seen_tests if k.isdigit()], default=0)
                    t_key = str(last_num + 1)
                
                if t_key in seen_tests:
                    # Check if existing one has valid cupom, if not replace with this one
                    existing_vals = seen_tests[t_key]
                    # Use the new index-based extraction
                    existing_nfce = _get_val(existing_vals, nfce_idx)
                    existing_sat = _get_val(existing_vals, sat_idx)
                    existing_ecf = _get_val(existing_vals, ecf_idx)
                    existing_cupom = _get_val(existing_vals, cupom_idx)
                    
                    # Check if this row has better cupom data
                    this_nfce = _get_val(vals, nfce_idx)
                    this_sat = _get_val(vals, sat_idx)
                    this_ecf = _get_val(vals, ecf_idx)
                    this_cupom = _get_val(vals, cupom_idx)
                    
                    existing_has_cupom = any(v and v.strip() for v in [existing_nfce, existing_sat, existing_ecf, existing_cupom])
                    this_has_cupom = any(v and v.strip() for v in [this_nfce, this_sat, this_ecf, this_cupom])
                    
                    if not existing_has_cupom and this_has_cupom:
                        seen_tests[t_key] = vals  # Replace with better row (store vals list)
                    # else keep existing
                    continue
                seen_tests[t_key] = vals  # Store vals list (maintains column order)

            # Build all_tests from seen_tests AFTER processing all rows (handles duplicates correctly)
            for t_key, vals in seen_tests.items():
                rd = dict(zip(header, vals))
                
                # Find column keys (case-insensitive)
                subtotal_key = next((k for k in rd if k and k.lower() in ['sub-total', 'subtotal']), 'Sub-Total')
                total_key = next((k for k in rd if k and k.lower() in ['total']), 'Total')
                desconto_key = next((k for k in rd if k and k.lower() in ['desconto']), 'Desconto')
                
                # Also look for alternative column names for ETAPA 2
                itens_key = next((k for k in rd if k and k.lower() in ['itens da venda', 'articulos movimiento', 'itens']), 'Itens da venda')
                pagamento_key = next((k for k in rd if k and k.lower() in ['pagamento']), 'Pagamento')
                observacoes_key = next((k for k in rd if k and k.lower() in ['observacoes']), 'Observacoes')
                observacao_parceiro_key = next((k for k in rd if k and k.lower() in ['observacoes.1']), 'Observacoes.1')
                tipo_promo_key = next((k for k in rd if k and k.lower() in ['tipo promo', 'tipo promocao', 'tipo promo', 'promotion_type', 'tipo']), 'Tipo Promo')
                
                # Cupom fields - priority: NFCE (under "Numero de cupom") > SAT > ECF > explicit Cupom columns
                nfce_key = next((k for k in rd if k and k.lower() in ['nfce', 'nfc-e']), 'NFCE')
                sat_key = next((k for k in rd if k and k.lower() in ['sat']), 'SAT')
                ecf_key = next((k for k in rd if k and k.lower() in ['ecf']), 'ECF')
                cupom_key = next((k for k in rd if k and k.lower() in ['cupom', 'numero cupom', 'numero do cupom']), 'Cupom')
                json_key = next((k for k in rd if k and k.lower() in ['json']), 'Json')
                minoristas_key = next((k for k in rd if k and k.lower() in ['minoristas']), 'Minoristas')
                
                # Cupom: priority NFCE (under "Numero de cupom") > SAT > ECF > explicit Cupom
                cupom_val = (_get_val(vals, nfce_idx) or 
                            _get_val(vals, sat_idx) or 
                            _get_val(vals, ecf_idx) or 
                            _get_val(vals, cupom_idx))
                
                # Also get individual fields using index-based extraction for consistency
                sat_val = _get_val(vals, sat_idx)
                ecf_val = _get_val(vals, ecf_idx)
                nfce_val = _get_val(vals, nfce_idx)
                json_val = _get_val(vals, json_idx)
                minoristas_val = _get_val(vals, minoristas_idx)
                
                std = {
                    'teste': float(t_key) if '.' in t_key else int(t_key),
                    'linha_original': f"{sheet_name}!{offset}",
                    'bloco_atual': sheet_name.strip().upper(),
                    'tipo_promo': rd.get(tipo_promo_key, ''),
                    'itens_da_venda': _get_val(vals, itens_idx),
                    'pagamento': _get_val(vals, pagamento_idx),
                    'observacoes': _get_val(vals, observacoes_idx),
                    'observacao_parceiro': _get_val(vals, observacao_parceiro_idx),
                    'subtotal_esperado': _get_val(vals, subtotal_idx),
                    'desconto_esperado': _get_val(vals, desconto_idx) or '0',
                    'total_esperado': _get_val(vals, total_idx),
                    'sat': sat_val,
                    'ecf': ecf_val,
                    'nfce': nfce_val,
                    'json': json_val,
                    'minoristas': minoristas_val,
                    # Cupom: priority NFCE > SAT > ECF > explicit Cupom
                    'cupom': cupom_val,
                    # Include product catalog for downstream use
                    'product_catalog': self.product_catalog,
                }
                all_tests.append(std)

        try:
            workbook.close()
        except Exception:
            pass

        return all_tests

    def _read_csv(self) -> List[Dict[str, Any]]:
        """Read test script from CSV file."""
        # Try different encodings
        encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                with open(self.file_path, 'r', encoding=encoding) as f:
                    # Try to detect delimiter
                    sample = f.read(2048)
                    f.seek(0)
                    
                    # Try common delimiters
                    delimiter = ','
                    if ';' in sample and sample.count(';') > sample.count(','):
                        delimiter = ';'
                    elif '\\t' in sample:
                        delimiter = '\\t'
                    
                    reader = csv.reader(f, delimiter=delimiter)
                    rows = list(reader)
                
                if not rows:
                    continue
                self.data = self._convert_rows_to_dicts(rows)
                break
            except UnicodeDecodeError:
                continue
            except Exception as e:
                if encoding == encodings[-1]:  # Last encoding tried
                    # If all encodings fail, raise the last error
                    raise e
                continue
        
        return self._process_data()

    def _convert_rows_to_dicts(self, rows: List[List[Any]]) -> List[Dict[str, Any]]:
        """
        Convert a list of rows (list of lists) to a list of dictionaries.
        Handles header detection and generic column naming.

        Args:
            rows: List of rows, where each row is a list of cell values

        Returns:
            List of dictionaries representing the data
        """
        if not rows:
            return []
        
        # Check if first row looks like a header
        first_row = [str(cell).strip().lower() for cell in rows[0]]
        header_indicators = ['teste', 'tipo', 'promo', 'itens', 'venda', 'pagamento', 
                           'observacoes', 'subtotal', 'desconto', 'total']
        matches = sum(1 for indicator in header_indicators 
                    if any(indicator in cell for cell in first_row))
        if matches >= 3:  # Likely a header row
            headers = [str(cell).strip() for cell in rows[0]]
            data_rows = rows[1:]
        else:
            # No header, create generic column names
            max_cols = max(len(row) for row in rows) if rows else 0
            headers = [f'col_{i}' for i in range(max_cols)]
            data_rows = rows
        
        # Convert to list of dictionaries
        result = []
        for row in data_rows:
            # Pad or truncate row to match headers
            padded_row = row + [''] * (len(headers) - len(row))
            padded_row = padded_row[:len(headers)]
            row_dict = dict(zip(headers, padded_row))
            result.append(row_dict)
        return result

    def _process_data(self) -> List[Dict[str, Any]]:
        """Process the raw data into test cases."""
        if not self.data:
            return []
        
        # Find the header row (first row that looks like a test header)
        header_index = self._find_header_row(self.data)
        
        if header_index is None:
            # If no clear header, assume data starts at index 0
            header_index = 0
        
        # Extract the data starting from the header row
        relevant_data = self.data[header_index:] if header_index < len(self.data) else []
        
        # Standardize column names
        standardized_data = []
        for row in relevant_data:
            if isinstance(row, dict):
                standardized_row = {}
                for key, value in row.items():
                    # Normalize key: lowercase, replace spaces and hyphens with underscores
                    norm_key = str(key).lower().strip().replace(' ', '_').replace('-', '_')
                    standardized_row[norm_key] = value
                standardized_data.append(standardized_row)
        
        # Map to expected columns
        expected_mapping = {
            'teste': ['teste', 'test', 'id', 'numero'],
            'tipo_promo': ['tipo_promo', 'tipo promocao', 'tipo promo', 'promotion_type', 'tipo'],
            'itens_da_venda': ['itens_da_venda', 'itens da venda', 'items', 'articulos_movimiento', 'itens'],
            'pagamento': ['pagamento', 'payment', 'forma_pagamento', 'pagamento'],
            'observacoes': ['observacoes', 'observacoes_raw', 'obs', 'observations', 'observacao'],
            'subtotal': ['sub_total', 'subtotal', 'sub-total', 'subtotal_esperado'],
            'desconto': ['desconto', 'discount', 'desconto_esperado'],
            'total': ['total', 'total_esperado']
        }
        
        # Create reverse mapping from variations to standard names
        variation_to_standard = {}
        for standard, variations in expected_mapping.items():
            for var in variations:
                variation_to_standard[var] = standard
        
        standardized_tests = []
        for idx, row in enumerate(standardized_data):
            # Skip rows that are clearly instructional or empty
            if self._is_instructional_row(row):
                continue
            
            test_dict = {
                'linha_original': idx + 1,  # Original line number (1-based)
                'bloco_atual': self._get_current_block(standardized_data, idx),
            }
            
            # Add each field, handling missing values
            for std_key, variations in expected_mapping.items():
                value = None
                # Try to find the value using any of the variation keys
                for var in variations:
                    if var in row:
                        value = row[var]
                        break
                if value is None:
                    # Try direct match with standard key
                    value = row.get(std_key, '')
                
                test_dict[std_key] = value if value is not None else '' 
            
            standardized_tests.append(test_dict)
        
        return standardized_tests

    def _find_header_row(self, data: List[Dict]) -> int:
        """
        Find the row that contains the test header.

        Returns:
            Index of the header row, or None if not found
        """
        if not data:
            return None
        
        # Common header indicators
        header_indicators = ['teste', 'tipo promo', 'itens da venda', 'pagamento', 
                           'sub total', 'desconto', 'total']
        
        for idx, row in enumerate(data):
            if not isinstance(row, dict):
                continue
            
            # Convert row values to lowercase strings for comparison
            row_values = ' '.join([str(val).lower() for val in row.values() if val])
            
            # Check if this row contains enough header indicators
            matches = sum(1 for indicator in header_indicators if indicator in row_values)
            if matches >= 4:  # At least 4 header indicators suggest this is the header row
                return idx
        
        return None

    def _is_instructional_row(self, row: Dict[str, Any]) -> bool:
        """
        Check if a row is instructional/header rather than a test case.
        
        Args:
            row: A row from the data
            
        Returns:
            True if the row is instructional, False otherwise
        """
        # Check if Teste column is non-numeric and not empty
        teste_value = row.get('teste', '')
        if not teste_value:  # Empty or None
            return True
        
        # If it's a string that doesn't look like a test number
        if isinstance(teste_value, str):
            teste_value = teste_value.strip()
            # Empty string or non-numeric instructional text
            if not teste_value:
                return True
            # Check if it's clearly instructional text
            upper_teste = teste_value.upper()
            if any(keyword in upper_teste for keyword in ['INSTRUÇÃO', 'SIGA', 'BLOCO', 'STATUS', 'NFCE', 'TIPO PROMO', 'NUMERO DE CUPOM', 'ARTICULOS MOVIMIENTO', 'OBSERVACOES', 'PREENCHIMENTO', 'SCANNTECH']):
                return True
            # Check if it's not a simple number (allow formulas like "1", "2", "3")
            if not (teste_value.replace('.', '').isdigit() or (teste_value.startswith('=') and 'A' in teste_value)):
                return True
        
        # Check if it's clearly a block header
        teste_str = str(teste_value).upper()
        if any(keyword in teste_str for keyword in ['BLOCO DE TESTE:', 'BLOCO']):
            return True
        
        return False

    def _get_current_block(self, data: List[Dict], row_index: int) -> str:
        """
        Determine the current block based on previous rows.

        Args:
            data: List of data rows
            row_index: Current row index in the data

        Returns:
            Block name string
        """
        # Look backwards for the most recent block header
        for i in range(row_index, -1, -1):
            if i < len(data):
                row = data[i]
                teste_value = row.get('teste', '')
                if isinstance(teste_value, str):
                    teste_str = teste_value.strip().upper()
                    if any(keyword in teste_str for keyword in ['BLOCO DE TESTE:', 'BLOCO']):
                        # Extract the block name
                        import re
                        match = re.search(r'BLOCO\\s+DE\\s+TESTE:\\s*(.+)', teste_str)
                        if match:
                            return match.group(1).strip()
                        else:
                            # Fallback: return the whole string after "BLOCO"
                            parts = teste_str.split('BLOCO', 1)
                            if len(parts) > 1:
                                return parts[1].strip(': ')
        return "UNKNOWN BLOCO"