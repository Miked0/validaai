#!/usr/bin/env python3
"""
Debug: Read all tests from ETAPA 1 sheet
"""

import openpyxl

wb = openpyxl.load_workbook('biblioteca/teste 2/TEMPLATE_COM_BIN_NOVO (2).xlsx')
ws = wb['ETAPA 1']

# Find header row
header = None
start_idx = None
for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    vals = [str(c).strip() if c is not None else '' for c in row]
    up = [v.upper() for v in vals]
    if 'TESTE' in up and any(k in up for k in ['ITENS DA VENDA', 'ARTICULOS MOVIMIENTO', 'ITENS']):
        header = vals
        start_idx = r_idx + 1
        print(f"Header found at row {r_idx}: {header}")
        break

# Print all rows
print(f"\nReading from row {start_idx}...")
for row_idx, row in enumerate(ws.iter_rows(min_row=start_idx, values_only=True), start=start_idx):
    vals = [str(c).strip() if c is not None else '' for c in row]
    if len(vals) < len(header):
        vals += [''] * (len(header) - len(vals))
    elif len(vals) > len(header):
        vals = vals[:len(header)]
    
    # Print all non-empty rows
    if any(v for v in vals):
        print(f"Row {row_idx}: {vals}")

wb.close()
