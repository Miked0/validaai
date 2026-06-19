#!/usr/bin/env python3
"""
ValidaAI - GUI Application v2 using Design System Components.
Modern, maintainable UI built on token-based design system.

Version: 2.0.0
Build Date: 2026-06-18
"""
import sys
import os
import threading
import json
from pathlib import Path
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Add project root to path
BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))

# Import core validation modules (inlined from run_full_validation.py)
from validaai import (
    TestScriptReader,
    ItemParser,
    PaymentNormalizer,
    TestValidator,
    ResultExporter,
    APISalesBuilder,
    API_SALES_AVAILABLE,
    get_payment_label,
)

# Import Design System
from gui.components import (
    apply_design_system,
    Card, CardSection,
    FieldRow,
    Button, ButtonGroup,
    StatusBadge, Status,
    ProgressRing,
    FileDropZone,
    LogViewer,
    create_file_field, create_folder_field, create_primary_button,
    create_success_button, create_danger_button, create_secondary_button,
)
# Import design tokens for colors
try:
    from design.tokens import (
        COLOR_BG_PRIMARY, COLOR_BG_CARD, COLOR_BG_INPUT, COLOR_BG_HOVER,
        COLOR_FG_PRIMARY, COLOR_FG_MUTED, COLOR_FG_DISABLED,
        COLOR_ACCENT_PRIMARY, COLOR_ACCENT_HOVER, COLOR_ACCENT_PRESS,
        COLOR_SUCCESS, COLOR_ERROR, COLOR_WARNING, COLOR_INFO,
        COLOR_BORDER_DEFAULT, COLOR_BORDER_FOCUS, COLOR_BORDER_ERROR,
    )
except ImportError:
    # Fallback values
    COLOR_BG_PRIMARY = "#2C2F33"
    COLOR_BG_CARD = "#23272A"
    COLOR_BG_INPUT = "#1E1F22"
    COLOR_BG_HOVER = "#2C2F33"
    COLOR_FG_PRIMARY = "#FFFFFF"
    COLOR_FG_MUTED = "#CCCCCC"
    COLOR_FG_DISABLED = "#888888"
    COLOR_ACCENT_PRIMARY = "#5865F2"
    COLOR_ACCENT_HOVER = "#4752C4"
    COLOR_ACCENT_PRESS = "#3C45A5"
    COLOR_SUCCESS = "#3BA55C"
    COLOR_ERROR = "#ED4245"
    COLOR_WARNING = "#FAA61A"
    COLOR_INFO = "#5865F2"
    COLOR_BORDER_DEFAULT = "#40444B"
    COLOR_BORDER_FOCUS = "#5865F2"
    COLOR_BORDER_ERROR = "#ED4245"

# Export Pipeline
from core.exporters import ExportPipeline, ExportConfig

# App metadata
APP_VERSION = "2.0.0"
APP_BUILD_DATE = "2026-06-18"


class ValidationController:
    """State machine for validation flow."""
    
    class State:
        IDLE = "idle"
        READING = "reading"
        PARSING = "parsing"
        NORMALIZING = "normalizing"
        VALIDATING = "validating"
        API_LOADING = "api_loading"
        API_VALIDATING = "api_validating"
        REVALIDATING = "revalidating"
        EXPORTING = "exporting"
        DONE = "done"
        ERROR = "error"
    
    def __init__(self, view: "ValidaAIAppV2"):
        self.view = view
        self.state = self.State.IDLE
        self._validated_tests: List[Dict] = []
        self._partner_jsons: Dict = {}
    
    @property
    def is_busy(self) -> bool:
        return self.state != self.State.IDLE and self.state != self.State.DONE and self.state != self.State.ERROR
    
    def transition(self, new_state: str) -> None:
        old_state = self.state
        self.state = new_state
        self.view.on_state_change(old_state, new_state)
    
    def run_validation(self, roteiro_path: str, etapa: str, audit_file: str, output_path: str, cupom_paths=None) -> None:
        """Start validation in background thread."""
        if self.is_busy:
            return
        
        self.transition(self.State.READING)
        thread = threading.Thread(
            target=self._validation_worker,
            args=(roteiro_path, etapa, audit_file, output_path, cupom_paths or []),
            daemon=True
        )
        thread.start()
    
    def _validation_worker(self, roteiro_path: str, etapa: str, audit_file: str, output_path: str, cupom_paths=None) -> None:
        cupom_paths = cupom_paths or []
        try:
            # Step 1: Read tests
            self.view.schedule_ui_update(lambda: self._log("1. Lendo roteiro de testes..."))
            if cupom_paths:
                self.view.schedule_ui_update(lambda: self._log(f"   Cupons selecionados: {len(cupom_paths)}"))
            reader = TestScriptReader(roteiro_path)
            if etapa:
                reader.set_etapa(etapa)
            raw_tests = reader.read_tests()
            self.view.schedule_ui_update(lambda: self._log(f"   Encontrados {len(raw_tests)} casos de teste.") if raw_tests else self._log("   Nenhum caso encontrado."))
            self.transition(self.State.PARSING)
            self.view.schedule_ui_update(lambda: self._set_progress(15))
            
            # Step 2: Parse items
            self.view.schedule_ui_update(lambda: self._log("2. Parseando itens..."))
            item_parser = ItemParser()
            parsed_tests = [item_parser.parse_items(t) for t in raw_tests]
            self.transition(self.State.NORMALIZING)
            self.view.schedule_ui_update(lambda: self._set_progress(30))
            
            # Step 3: Normalize payments
            self.view.schedule_ui_update(lambda: self._log("3. Normalizando pagamentos..."))
            payment_normalizer = PaymentNormalizer()
            normalized_tests = [payment_normalizer.normalize_payment(t) for t in parsed_tests]
            self.transition(self.State.VALIDATING)
            self.view.schedule_ui_update(lambda: self._set_progress(45))
            
            # Step 4: Business rule validation
            self.view.schedule_ui_update(lambda: self._log("4. Validando regras de negócio..."))
            validator = TestValidator(tolerance=0.01)
            validated_tests = [validator.validate(t) for t in normalized_tests]
            self.transition(self.State.API_LOADING)
            self.view.schedule_ui_update(lambda: self._set_progress(60))
            
            # Step 5: Load & validate partner JSONs
            self.view.schedule_ui_update(lambda: self._log("5. Carregando e validando JSON do parceiro..."))
            if API_SALES_AVAILABLE:
                api_builder = APISalesBuilder()
                partner_jsons = self._load_partner_jsons(audit_file) if audit_file else {}
                self._partner_jsons = partner_jsons
                
                self.view.schedule_ui_update(lambda: self._log(f"   Carregados {len(partner_jsons)} JSONs do parceiro."))
                if partner_jsons:
                    self.view.schedule_ui_update(lambda: self._log(f"   Testes com JSON: {sorted(partner_jsons.keys())}"))
                
                for t in validated_tests:
                    try:
                        test_cupom = self._get_test_cupom(t)
                        partner_json = partner_jsons.get(test_cupom)
                        
                        if partner_json:
                            self.view.schedule_ui_update(lambda t=t, test_cupom=test_cupom: self._log(f"   Teste {t.get('teste')}: Usando JSON DO PARCEIRO (cupom: {test_cupom})"))
                            t['sale_json'] = partner_json
                            api_check = api_builder.validate_sale_json(partner_json)
                            t['api_status'] = api_check.get('status', 'ERRO_JSON')
                            t['api_alertas'] = api_check.get('alertas', []) or []
                            self.view.schedule_ui_update(lambda t=t: self._log(f"   Teste {t.get('teste')}: Validação -> {api_check.get('status')} - {api_check.get('motivo', '')}"))
                            
                            # Validate cupom consistency
                            self._validate_cupom_consistency(t, partner_json)
                            
                            # Validate payment codes
                            self._validate_payment_codes(t, partner_json)
                        else:
                            self.view.schedule_ui_update(lambda t=t, test_cupom=test_cupom: self._log(f"   Teste {t.get('teste')}: ERRO - JSON do parceiro NÃO ENCONTRADO (cupom: {test_cupom})"))
                            t['sale_json'] = {}
                            t['api_status'] = 'ERRO'
                            t['api_alertas'] = [f"JSON do parceiro não encontrado para cupom {test_cupom} (teste {t.get('teste')})"]
                    except Exception as e:
                        t['sale_json'] = {}
                        t['api_status'] = 'ERRO_JSON'
                        t['api_alertas'] = [str(e)]
            else:
                self.view.schedule_ui_update(lambda: self._log("   api_sales indisponível; pulando etapa de API."))
            
            self.transition(self.State.REVALIDATING)
            self.view.schedule_ui_update(lambda: self._set_progress(80))
            
            # Step 6: Apply API results and re-validate
            self.view.schedule_ui_update(lambda: self._log("6. Aplicando resultados da API e re-validando..."))
            self._apply_api_results(validated_tests)
            
            revalidator = TestValidator(tolerance=0.01)
            final_tests = []
            for t in validated_tests:
                api_status = t.get('api_status', '')
                if api_status.startswith('ERRO') and not t.get('api_status') == 'IGNORADO':
                    final_tests.append(t)
                else:
                    final_tests.append(revalidator.validate(t))
            validated_tests = final_tests
            self._validated_tests = validated_tests
            
            # Step 7: Export results
            self.transition(self.State.EXPORTING)
            self.view.schedule_ui_update(lambda: self._set_progress(90))
            self.view.schedule_ui_update(lambda: self._log("7. Exportando resultados..."))
            
            self._export_results(validated_tests, output_path, audit_file, roteiro_path)
            
            # Done
            self.view.schedule_ui_update(lambda: self._set_progress(100))
            status_counts = {}
            for test in validated_tests:
                status = test.get('status_final', 'UNKNOWN')
                status_counts[status] = status_counts.get(status, 0) + 1
            summary = " | ".join(f"{k}: {v}" for k, v in status_counts.items())
            self.view.schedule_ui_update(lambda: self._log(f"Concluído. {summary}"))
            self.transition(self.State.DONE)
            self.view.schedule_ui_update(lambda s=summary, o=output_path: self._on_complete(s, o))
            
        except Exception as e:
            self.transition(self.State.ERROR)
            self.view.schedule_ui_update(lambda: self._log(f"ERRO: {e}"))
            self.view.schedule_ui_update(lambda: self._on_error(str(e)))
    
    def _load_partner_jsons(self, audit_file: str) -> Dict:
        import pandas as pd
        jsons = {}
        total_parsed = 0
        xls = pd.ExcelFile(audit_file)
        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(audit_file, sheet_name=sheet_name, dtype=str)
                test_col = None
                test_cols = [c for c in df.columns if c and ('teste' in c.lower() or 'cupom' in c.lower() or 'numero' in c.lower())]
                if test_cols:
                    for preferred in ['Número cupom', 'Numero cupom', 'Teste', 'Numero']:
                        if preferred in df.columns:
                            test_col = preferred
                            break
                        for c in test_cols:
                            if preferred.lower() in c.lower():
                                test_col = c
                                break
                    if test_col is None and test_cols:
                        test_col = test_cols[0]
                
                request_col = None
                if 'Request' in df.columns:
                    request_col = 'Request'
                else:
                    request_cols = [c for c in df.columns if c and ('request' in c.lower() or 'json' in c.lower() or 'movimiento' in c.lower()) and c.lower() != 'id request']
                    if request_cols:
                        for preferred in ['Request JSON', 'Json Request', 'JSON', 'Json']:
                            for c in request_cols:
                                if preferred.lower() == c.lower():
                                    request_col = c
                                    break
                            if request_col:
                                break
                        if request_col is None:
                            request_col = request_cols[0]
                
                if not test_col or not request_col:
                    continue
                
                for _, row in df.iterrows():
                    test_val = str(row.get(test_col, '')).strip()
                    request_val = str(row.get(request_col, '')).strip()
                    if test_val and request_val and request_val not in ['nan', 'None', '']:
                        try:
                            parsed = json.loads(request_val)
                            jsons[test_val] = parsed
                            total_parsed += 1
                        except Exception:
                            pass
            except Exception:
                pass
        return jsons
    
    def _get_test_cupom(self, test_dict: Dict) -> str:
        placeholders = {'(status)', 'status', '(status)', 'n/a', 'na', ''}
        for field in ['cupom', 'nfce', 'sat', 'ecf']:
            val = str(test_dict.get(field, '')).strip()
            if val and val.lower() not in ['nan', 'none', ''] and val.lower() not in placeholders:
                return val
        return ''
    
    def _validate_cupom_consistency(self, test_dict: Dict, partner_json: Dict) -> None:
        def _extrair_cupom_json(json_data):
            if not isinstance(json_data, dict):
                return ''
            if isinstance(json_data.get('movimiento'), dict):
                return str(json_data['movimiento'].get('numero', '')).strip()
            return str(json_data.get('numero', '')).strip()
        
        cupom_roteiro = str(test_dict.get('cupom', '')).strip()
        cupom_json = _extrair_cupom_json(partner_json)
        cupom_sat = str(test_dict.get('sat', '')).strip()
        cupom_ecf = str(test_dict.get('ecf', '')).strip()
        cupom_nfce = str(test_dict.get('nfce', '')).strip()
        
        cupom_valido = False
        if cupom_roteiro and cupom_json and cupom_roteiro == cupom_json:
            cupom_valido = True
        elif cupom_sat and cupom_json and cupom_sat == cupom_json:
            cupom_valido = True
        elif cupom_ecf and cupom_json and cupom_ecf == cupom_json:
            cupom_valido = True
        elif cupom_nfce and cupom_json and cupom_nfce == cupom_json:
            cupom_valido = True
        
        if not cupom_valido:
            if cupom_roteiro or cupom_sat or cupom_ecf or cupom_nfce:
                test_dict['api_status'] = 'ERRO'
                motivo = f"Cupom não confere: Roteiro='{cupom_roteiro}', SAT='{cupom_sat}', ECF='{cupom_ecf}', NFCe='{cupom_nfce}' vs JSON='{cupom_json}'"
                test_dict['api_alertas'].append(motivo)
                self.view.schedule_ui_update(lambda m=motivo: self._log(f"   ERRO Cupom: {m}"))
    
    def _validate_payment_codes(self, test_dict: Dict, partner_json: Dict) -> None:
        def _extrair_pagos_json(json_data):
            if not isinstance(json_data, dict):
                return []
            if isinstance(json_data.get('movimiento'), dict):
                return json_data['movimiento'].get('pagos', [])
            return json_data.get('pagos', [])
        
        pagamentos_esperados = test_dict.get('pagamentos', [])
        if pagamentos_esperados:
            expected_codes = sorted([str(p.get('codigo', '')).strip() for p in pagamentos_esperados if p.get('codigo')])
        else:
            pgto_raw = str(test_dict.get('pagamento', '')).strip().lower()
            pgto_map = {
                'dinheiro': '9', 'dinheiro com troco': '9',
                'cartao credito': '10', 'cartao crédito': '10',
                'cartao debito': '13', 'cartao débito': '13',
                'pix': '14', 'qr': '14', 'pix/qr': '14',
                'cheque': '11', 'vale': '12', 'finalizadora': '15',
            }
            expected_codes = [pgto_map.get(pgto_raw, '')] if pgto_raw else []
        
        pagos_json = _extrair_pagos_json(partner_json)
        actual_codes = sorted([str(p.get('codigoTipoPago', '')).strip() for p in pagos_json if p.get('codigoTipoPago')])
        
        if expected_codes and actual_codes:
            if expected_codes != actual_codes:
                motivo = f"Códigos de pagamento divergentes: esperado={expected_codes} vs parceiro={actual_codes}"
                test_dict['api_alertas'].append(motivo)
                if test_dict.get('api_status') == 'OK':
                    test_dict['api_status'] = 'REVISAO'
                self.view.schedule_ui_update(lambda m=motivo: self._log(f"   REVISAO Pagamento: {m}"))
                for i, p in enumerate(pagos_json):
                    codigo = p.get('codigoTipoPago')
                    label = get_payment_label(codigo) if codigo is not None else 'N/A'
                    self.view.schedule_ui_update(lambda i=i, label=label, codigo=codigo, p=p: self._log(f"   Pago {i}: {label} (codigoTipoPago={codigo}), detalleFinalizadora={p.get('detalleFinalizadora')}, importe={p.get('importe')}"))
    
    def _apply_api_results(self, validated_tests: List[Dict]) -> None:
        for t in validated_tests:
            api_status = t.get('api_status', '')
            api_alertas = t.get('api_alertas', []) or []
            api_ignorado = t.get('api_status') == 'IGNORADO'
            
            test_cupom = ''
            placeholders = {'(status)', 'status', '(status)', 'n/a', 'na', ''}
            for field in ['cupom', 'nfce', 'sat', 'ecf']:
                val = str(t.get(field, '')).strip()
                if val and val.lower() not in ['nan', 'none', ''] and val.lower() not in placeholders:
                    test_cupom = val
                    break
            
            if test_cupom and test_cupom.startswith('-'):
                t['status_final'] = 'OK'
                t['motivo_status'] = 'Venda cancelada (cupom negativo) - não validado'
                t['api_status'] = 'IGNORADO'
                t['api_alertas'] = ['Venda cancelada ignorada']
                t['sale_json'] = {}
                continue
            
            if api_ignorado:
                continue
            
            if api_status.startswith('ERRO'):
                t['status_final'] = api_status
                t['motivo_status'] = '; '.join(api_alertas) if api_alertas else 'Erro na validação do parceiro'
                t['alertas'] = list(set((t.get('alertas', []) or []) + api_alertas))
            elif api_status == 'REVISAO':
                if t.get('status_final') == 'OK':
                    t['status_final'] = 'REVISAO'
                    t['motivo_status'] = '; '.join(api_alertas) if api_alertas else 'Revisão necessária'
                t['alertas'] = list(set((t.get('alertas', []) or []) + api_alertas))
            else:
                t['alertas'] = list(set((t.get('alertas', []) or []) + api_alertas))

# ════════════════════════════════════════════════════════════
# Export Pipeline Integration
# ═════════════════════════════════════════════════════════════

    def _export_results(self, validated_tests: List[Dict], output_path: str, audit_file: str, roteiro_path: str) -> None:
        """Export validation results using the new pluggable pipeline."""
        self.view.schedule_ui_update(lambda: self._log("   Exportando resultados com pipeline multi-formato..."))
        
        output_path_obj = Path(output_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        etapa = self.etapa_var.get().strip() or "TODAS"
        
        metadata = {
            "app_version": APP_VERSION,
            "etapa": etapa,
            "roteiro": roteiro_path,
            "audit_file": audit_file,
            "generated_at": datetime.now().isoformat(),
        }
        
        # Configure pipeline with all exporters
        config = ExportConfig(
            exporters=["excel", "csv", "json_audit", "html_report"],
            excel={"sheet_name": "Resumo", "freeze_header": True},
            csv={"delimiter": ",", "encoding": "utf-8-sig"},
            json_audit={"indent": 2},
            html_report={"theme": "dark"},
        )
        pipeline = ExportPipeline(config)
        
        # Run all exporters
        base_path = output_path_obj.parent / f"validacao_{timestamp}_{etapa}_v{APP_VERSION}"
        export_results = pipeline.run(validated_tests, base_path, metadata)
        
        # Log results
        for exp_result in export_results:
            if exp_result.success:
                self.view.schedule_ui_update(
                    lambda r=exp_result: self._log(f"   ✓ {r.exporter_name.upper()}: {r.output_path.name} ({r.rows_exported} rows, {r.duration_ms}ms)")
                )
            else:
                self.view.schedule_ui_update(
                    lambda r=exp_result: self._log(f"   ✗ {r.exporter_name.upper()}: ERRO - {r.error}")
                )
        
        # Also export resumo to original roteiro location (backward compatibility)
        if roteiro_path and Path(roteiro_path).exists():
            roteiro_output = Path(roteiro_path).parent / "validacao_resultado.xlsx"
            try:
                from core.exporters import ExcelExporter
                excel_exporter = ExcelExporter({"sheet_name": "Resumo", "freeze_header": True})
                excel_exporter.export(validated_tests, roteiro_output, metadata)
                self.view.schedule_ui_update(
                    lambda p=str(roteiro_output): self._log(f"   Resumo compatível salvo em: {p}")
                )
            except Exception as e:
                self.view.schedule_ui_update(
                    lambda: self._log(f"   Aviso: falha ao salvar resumo compatível: {e}")
                )
    
    # Callbacks for UI updates
    def _log(self, msg: str) -> None:
        self.view.log(msg)
    
    def _set_progress(self, value: float) -> None:
        self.view.set_progress(value)
    
    def _on_complete(self, summary: str, output_path: str) -> None:
        self.view.on_validation_complete(summary, output_path)
    
    def _on_error(self, error: str) -> None:
        self.view.on_validation_error(error)


class ValidaAIAppV2:
    """
    Main application window using Design System Components.
    """
    
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("ValidaAI — Validação de Roteiro de Testes PDV")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)
        
        # Apply Design System
        apply_design_system(root)
        
        # State variables
        self.roteiro_path = tk.StringVar()
        self.cupom_paths: List[str] = []  # Lista de cupons PDF/JPEG
        self.audit_file = tk.StringVar()
        self.output_path = tk.StringVar(value=str(BASE_DIR / "output" / "validacao_resultado.xlsx"))
        self.etapa_var = tk.StringVar()
        self.etapa_list: List[str] = []
        
        # Controller
        self.controller = ValidationController(self)
        
        # UI state
        self._ui_updates: List[callable] = []
        self._process_ui_updates()
        
        # Build UI
        self._build_ui()
        
        # Keyboard shortcuts
        self.root.bind("<Control-o>", lambda e: self._select_roteiro())
        self.root.bind("<Control-s>", lambda e: self._select_output())
        self.root.bind("<F5>", lambda e: self._run_validation())
        self.root.bind("<Escape>", lambda e: self.root.quit())
    
    def _process_ui_updates(self) -> None:
        """Process queued UI updates from background thread."""
        while self._ui_updates:
            try:
                update = self._ui_updates.pop(0)
                update()
            except Exception:
                pass
        self.root.after(50, self._process_ui_updates)
    
    def schedule_ui_update(self, callback: callable) -> None:
        """Schedule a UI update from background thread (thread-safe)."""
        self._ui_updates.append(callback)
    
    def log(self, message: str) -> None:
        """Add message to log viewer."""
        self.schedule_ui_update(lambda: self.log_viewer.log(message))
    
    def set_progress(self, value: float) -> None:
        """Update progress ring."""
        self.schedule_ui_update(lambda: self.progress_ring.progress.__set__(value / 100))
    
    def on_state_change(self, old_state: str, new_state: str) -> None:
        """Handle controller state changes."""
        self.schedule_ui_update(lambda: self._update_ui_for_state(new_state))
    
    def on_validation_complete(self, summary: str, output_path: str) -> None:
        """Handle validation completion."""
        self.schedule_ui_update(lambda: self._on_complete(summary, output_path))
    
    def on_validation_error(self, error: str) -> None:
        """Handle validation error."""
        self.schedule_ui_update(lambda: self._on_error(error))
    
    def _update_ui_for_state(self, state: str) -> None:
        """Update UI elements based on controller state."""
        is_busy = self.controller.is_busy
        
        self.run_btn.config(state=tk.DISABLED if is_busy else tk.NORMAL)
        
        if state == ValidationController.State.READING:
            self.progress_ring.message = "Lendo..."
        elif state == ValidationController.State.PARSING:
            self.progress_ring.message = "Parseando..."
        elif state == ValidationController.State.NORMALIZING:
            self.progress_ring.message = "Normalizando..."
        elif state == ValidationController.State.VALIDATING:
            self.progress_ring.message = "Validando..."
        elif state == ValidationController.State.API_LOADING:
            self.progress_ring.message = "Carregando API..."
        elif state == ValidationController.State.API_VALIDATING:
            self.progress_ring.message = "Validando API..."
        elif state == ValidationController.State.REVALIDATING:
            self.progress_ring.message = "Re-validando..."
        elif state == ValidationController.State.EXPORTING:
            self.progress_ring.message = "Exportando..."
        elif state == ValidationController.State.DONE:
            self.progress_ring.message = "✓ Concluído"
        elif state == ValidationController.State.ERROR:
            self.progress_ring.message = "✗ Erro"
    
    def _build_ui(self) -> None:
        # Main container
        main = ttk.Frame(self.root, padding=16, style="TFrame")
        main.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header = ttk.Frame(main, style="TFrame")
        header.pack(fill=tk.X, pady=(0, 16))
        
        ttk.Label(header, text="ValidaAI", style="Title.TLabel").pack(side=tk.LEFT)
        ttk.Label(header, text=f"v{APP_VERSION}", style="Subtitle.TLabel").pack(side=tk.LEFT, padx=(8, 0))
        
        # Input Card
        input_card = CardSection(main, text="Entradas", padding=12)
        input_card.pack(fill=tk.X, pady=(0, 12))
        
        # Roteiro field
        self.roteiro_field = create_file_field(
            input_card,
            "Roteiro de testes (Excel/CSV):",
            self.roteiro_path,
            filetypes=[("Planilhas", "*.xlsx *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv")]
        )
        self.roteiro_field.pack(fill=tk.X, pady=(0, 8))
        # Override browse to detect etapas immediately after selection
        original_browse = self.roteiro_field.browse_btn.cget("command")
        def browse_with_etapa():
            from tkinter import filedialog
            path = filedialog.askopenfilename(
                title="Selecionar roteiro de testes",
                filetypes=[("Planilhas", "*.xlsx *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv")]
            )
            if path:
                self.roteiro_path.set(path)
                self._on_roteiro_selected()
        self.roteiro_field.browse_btn.config(command=browse_with_etapa)

        # Cupons field (PDF/JPEG) - multi-select
        cupom_frame = ttk.Frame(input_card, style="TFrame")
        cupom_frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(cupom_frame, text="Cupons (PDF/JPEG):").pack(side=tk.LEFT, padx=(0, 8))
        self.cupom_listbox = tk.Listbox(cupom_frame, height=4, bg=COLOR_BG_INPUT, fg=COLOR_FG_PRIMARY, selectbackground=COLOR_ACCENT_PRIMARY)
        self.cupom_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        cupom_btn_frame = ttk.Frame(cupom_frame, style="TFrame")
        cupom_btn_frame.pack(side=tk.LEFT)
        ttk.Button(cupom_btn_frame, text="Adicionar", command=self._add_cupons, style="DS.Secondary.TButton").pack(side=tk.TOP, pady=2)
        ttk.Button(cupom_btn_frame, text="Remover", command=self._remove_cupom, style="DS.Secondary.TButton").pack(side=tk.TOP, pady=2)

        # Etapa combo
        etapa_frame = ttk.Frame(input_card, style="TFrame")
        etapa_frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(etapa_frame, text="Etapa:").pack(side=tk.LEFT, padx=(0, 8))
        self.etapa_combo = ttk.Combobox(
            etapa_frame,
            textvariable=self.etapa_var,
            state="readonly",
            width=25,
            style="DS.TCombobox"
        )
        self.etapa_combo.pack(side=tk.LEFT)
        
        # Audit file field
        self.audit_field = create_file_field(
            input_card,
            "Export da auditoria (xlsx com JSONs do parceiro):",
            self.audit_file,
            filetypes=[("Excel", "*.xlsx")]
        )
        self.audit_field.pack(fill=tk.X, pady=(0, 8))
        
        # Output Card
        output_card = CardSection(main, text="Saída", padding=12)
        output_card.pack(fill=tk.X, pady=(0, 12))
        
        output_field = create_file_field(
            output_card,
            "Arquivo de resultado:",
            self.output_path,
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")]
        )
        # Override browse to use asksaveasfilename
        def browse_output():
            from tkinter import filedialog
            p = filedialog.asksaveasfilename(
                title="Salvar resultado como",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
                initialfile="validacao_resultado.xlsx"
            )
            if p:
                self.output_path.set(p)
        output_field.browse_btn.config(command=browse_output)
        output_field.pack(fill=tk.X)
        
        # Progress Card
        progress_card = CardSection(main, text="Progresso", padding=12)
        progress_card.pack(fill=tk.X, pady=(0, 12))
        
        progress_row = ttk.Frame(progress_card.content, style="TFrame")
        progress_row.pack(fill=tk.X)
        
        self.progress_ring = ProgressRing(progress_row, size=56)
        self.progress_ring.pack(side=tk.LEFT, padx=(0, 12))
        
        self.status_label = ttk.Label(progress_row, text="Pronto", style="Status.TLabel")
        self.status_label.pack(side=tk.LEFT, anchor=tk.W)
        
        # Action Buttons
        actions = ButtonGroup(main, spacing=8)
        actions.pack(fill=tk.X, pady=(8, 8))
        
        self.run_btn = create_primary_button(actions.widget, "Executar Validação (F5)", self._run_validation)
        actions.add_button("Abrir pasta", self._open_output_dir, variant="secondary")
        actions.add_button("Sair", self.root.quit, variant="danger")
        
        # Log Card
        log_card = CardSection(main, text="Log", padding=8)
        log_card.pack(fill=tk.BOTH, expand=True)
        
        self.log_viewer = LogViewer(log_card.content, height=12)
        self.log_viewer.pack(fill=tk.BOTH, expand=True)
        self.log_viewer.log("ValidaAI v2 pronto. Selecione um roteiro para iniciar.", "success")
    
    def _on_roteiro_selected(self) -> None:
        """Detect ETAPA sheets when roteiro is selected."""
        path = self.roteiro_path.get().strip()
        if not path or not Path(path).exists():
            return
        
        if path.lower().endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                etapas = [sn.strip() for sn in wb.sheetnames if 'ETAPA' in sn.upper()]
                wb.close()
                if etapas:
                    self.etapa_list = etapas
                    self.etapa_combo['values'] = etapas
                    self.etapa_var.set(etapas[0])
                else:
                    self.etapa_list = ['Etapa 1', 'Etapa 2', 'Etapa 3']
                    self.etapa_combo['values'] = self.etapa_list
            except Exception:
                self.etapa_list = ['Etapa 1', 'Etapa 2', 'Etapa 3']
                self.etapa_combo['values'] = self.etapa_list
    
    def _add_cupons(self) -> None:
        """Add coupon PDF/JPEG files."""
        from tkinter import filedialog
        paths = filedialog.askopenfilenames(
            title="Selecionar cupons (PDF/JPEG)",
            filetypes=[("Imagens/PDF", "*.pdf *.jpg *.jpeg *.png"), ("PDF", "*.pdf"), ("JPEG", "*.jpg *.jpeg"), ("Todos", "*.*")],
        )
        if paths:
            self.cupom_paths.extend(paths)
            self._refresh_cupom_list()

    def _remove_cupom(self) -> None:
        """Remove selected coupon from list."""
        sel = self.cupom_listbox.curselection()
        if sel:
            idx = sel[0]
            self.cupom_paths.pop(idx)
            self._refresh_cupom_list()

    def _refresh_cupom_list(self) -> None:
        """Refresh the coupon listbox display."""
        self.cupom_listbox.delete(0, tk.END)
        for p in self.cupom_paths:
            self.cupom_listbox.insert(tk.END, Path(p).name)

    def _run_validation(self) -> None:
        roteiro = self.roteiro_path.get().strip()
        etapa = self.etapa_var.get().strip()
        audit = self.audit_file.get().strip()
        output = self.output_path.get().strip()
        
        # Validation
        if not roteiro:
            messagebox.showerror("Erro", "Nenhum roteiro selecionado.\n\nClique em \"Selecionar\" e escolha um arquivo .xlsx ou .csv.")
            return
        if not Path(roteiro).exists():
            messagebox.showerror("Erro", f"Arquivo não encontrado:\n{roteiro}\n\nVerifique se o arquivo foi movido ou excluído.")
            return
        ext = Path(roteiro).suffix.lower()
        if ext not in ('.xlsx', '.csv'):
            messagebox.showerror("Erro", f"Extensão não suportada: {ext}\n\nUse apenas arquivos .xlsx ou .csv")
            return
        if ext == '.xlsx':
            try:
                import openpyxl
                wb = openpyxl.load_workbook(roteiro, read_only=True, data_only=True)
                has_etapa = any('ETAPA' in sn.upper() for sn in wb.sheetnames)
                wb.close()
                if not has_etapa:
                    messagebox.showwarning("Aviso", "Nenhuma aba 'ETAPA' encontrada no arquivo.\nO processamento tentará ler todas as abas.")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao ler arquivo Excel:\n{e}")
                return
        
        self.log_viewer.clear()
        self.log("Iniciando validação...")
        self.controller.run_validation(roteiro, etapa, audit, output, self.cupom_paths)
    
    def _open_output_dir(self) -> None:
        out_dir = Path(self.output_path.get()).parent
        out_dir.mkdir(parents=True, exist_ok=True)
        os.startfile(str(out_dir))
    
    def _on_complete(self, summary: str, output_path: str) -> None:
        self.log(f"Concluído. {summary}")
        messagebox.showinfo("Validação concluída", f"Resultados exportados para:\n{output_path}")
    
    def _on_error(self, error: str) -> None:
        self.log(f"ERRO: {error}")
        messagebox.showerror("Erro na validação", error)


def main():
    root = tk.Tk()
    app = ValidaAIAppV2(root)
    root.mainloop()


if __name__ == "__main__":
    main()