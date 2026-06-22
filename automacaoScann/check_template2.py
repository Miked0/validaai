import openpyxl
wb = openpyxl.load_workbook('biblioteca/TEMPLATE_COM_BIN_NOVO.xlsx')
ws = wb['ETAPA 1']
print(f'ETAPA 1: {ws.max_row} rows x {ws.max_column} cols')
for i, row in enumerate(ws.iter_rows(max_row=34, values_only=True)):
    print(f'  Row {i+1}: {list(row)}')