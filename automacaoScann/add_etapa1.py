import openpyxl

wb = openpyxl.load_workbook('C:/Users/Mike/validaai/automacaoScann/biblioteca/TEMPLATE_COM_BIN_NOVO.xlsx')

# ============================================================
# ABA: ETAPA 1 - 27 testes completos
# ============================================================
ws = wb.create_sheet('ETAPA 1')

# Cabeçalho
headers = [
    'Teste', 'Tipo Promo', 'Itens da venda', 'Pagamento', 'Observacoes',
    'SAT', 'ECF', 'NFCE', 'Sub-Total', 'Desconto', 'Total',
    'Json', 'Minoristas', 'Cupom', 'Observacoes.1'
]
for col, h in enumerate(headers, 1):
    ws.cell(row=7, column=col, value=h)

# Dados dos 27 testes (baseados nas especificações)
testes = [
    (1, 'N/A', '2 x 7891000010860 + 3.579 x PESABLE', 'Dinheiro', 'N/A', '', '', '', '149.07', '0.00', '149.07', '', '', '', ''),
    (2, 'N/A', '5 x 7894904500383 + 4 x 7894904578207 + 7894904003495 + 7894904003495', 'Dinheiro com Troco', 'N/A', '', '', '', '16.17', '0.00', '16.17', '', '', '', ''),
    (3, 'N/A', '1 x 7894904003495 + 4 x 7894904003495 + 4 x 7894904003495', 'Cartao Credito', 'N/A', '', '', '', '17.28', '0.00', '17.28', '', '', '', ''),
    (4, 'N/A', '2 x 7891149440801 + 1 x 7891149102808 + 2 x 7891991001359', 'Cartao Credito', 'Realizar o pagamento com finalizadora POS', '', '', '', '25.31', '0.00', '25.31', '', '', '', ''),
    (5, 'N/A', '3 x 7891991001359 + 1 x 7891149102808 + 2 x 789114940888888 + 2 x 7891991001359', 'Cartao Debito', 'N/A', '', '', '', '45.67', '0.00', '45.67', '', '', '', ''),
    (6, 'N/A', '5 x 7891149102808 + 1 x 7891149102808', 'PIX', 'N/A', '', '', '', '30.54', '0.00', '30.54', '', '', '', ''),
    (7, 'N/A', '2 x 7894904573394 + 3 x 7894904573387', 'Dinheiro e Cartao Credito', 'Utilizar canal de venda 2', '', '', '', '14.25', '0.00', '14.25', '', '', '', ''),
    (8, 'N/A', '4 x 7894904573387 + 2 x 7894904573394 + 2 x 7894904573387', 'Dinheiro e Cartao Debito', 'N/A', '', '', '', '22.80', '0.00', '22.80', '', '', '', ''),
    (9, 'N/A', '6 x 7896079500175', 'Dinheiro com Troco e duas vezes Cartao Credito', 'Utilizar canal de venda diferente de 1 e 2 (ex: canal 3)', '', '', '', '22.14', '0.00', '22.14', '', '', '', ''),
    (10, 'N/A', '8 x 7897511400237 + 4 x 7897511400244', 'Dinheiro + Dinheiro + Cartao Credito', 'N/A', '', '', '', '44.28', '0.00', '44.28', '', '', '', ''),
    (11, 'N/A', '1 x 7891024132906', 'Dinheiro', 'cancelar venda', '', '', '', '3.50', '0.00', '3.50', '', '', '', ''),
    (12, 'N/A', '3 x 7891024132906', 'Dinheiro com Troco', 'cancelar venda', '', '', '', '10.50', '0.00', '10.50', '', '', '', ''),
    (13, 'N/A', '1 x 7891024132906 + 1 x 7891024132906 + 1 x 7891024132906', 'Cartao Credito', 'cancelar venda', '', '', '', '10.50', '0.00', '10.50', '', '', '', ''),
    (14, 'N/A', '2 x 7891149105533 + 2 x 7891149103119', 'Cartao Debito', 'cancelar venda', '', '', '', '140.50', '0.00', '140.50', '', '', '', ''),
    (15, 'N/A', '1 x 7891149105533 + 1 x 7891149103119', 'Dinheiro + Cartao Credito', 'cancelar venda', '', '', '', '70.25', '0.00', '70.25', '', '', '', ''),
    (16, 'N/A', '3 x 7891000010860 + 2 x 7894904573394 + 1 x 7894904573387 + 3 x 7894904573387 + 3.579 x PESABLE + 1 x 7896079500175', 'Dinheiro + Cartao Debito', 'cancelar venda', '', '', '', '181.76', '0.00', '181.76', '', '', '', ''),
    (17, 'N/A', '3 x 7891991001359 + 4 x 7894904573387 + 1 x 7891149102808', 'PIX', 'cancelar venda', '', '', '', '31.76', '0.00', '31.76', '', '', '', ''),
    (18, 'N/A', '4 x 7894904573394 + 1 x 7891149103119', 'Dinheiro', 'Dar acrescimo na linha', '', '', '', '45.75', '-5.00', '50.75', '', '', '', ''),
    (19, 'N/A', '2 x 7891024132906', 'Dinheiro', 'Dar acrescimo no subtotal ou cabecalho', '', '', '', '7.00', '-6.00', '13.00', '', '', '', ''),
    (20, 'N/A', '5 x 5000329002537 + 1 x 7891024132906', 'Dinheiro', 'Dar desconto na linha', '', '', '', '501.00', '6.00', '495.00', '', '', '', ''),
    (21, 'N/A', '3 x 5000329002537 + 1 x 7891150024588 + 1 x 7891024132906', 'Dinheiro', 'Dar desconto no subtotal/cabeçalho', '', '', '', '334.53', '6.00', '328.53', '', '', '', ''),
    (22, 'N/A', '2 x 7894904573394', '', 'cancelar cupom antes de pagar', '', '', '', '5.70', '0.00', '5.70', '', '', '', ''),
    (23, 'N/A', '2 x 7891000010860 + 1 x 7891000029329', 'Dinheiro', 'cancelar item 7891000029329', '', '', '', '23.80', '0.00', '23.80', '', '', '', ''),
    (24, 'N/A', '6 x 7896079500175', 'Dinheiro', 'cancelar unidade (de 6 para 5)', '', '', '', '18.45', '0.00', '18.45', '', '', '', ''),
    (25, 'N/A', '2 x 7891000010860 + 357.9 x PESABLE', 'Dinheiro', 'peso grande 357.9', '', '', '', '12550.30', '0.00', '12550.30', '', '', '', ''),
    (26, 'N/A', '2 x 1003607622300391065', 'Dinheiro', 'EAN invalido de 19 digitos', '', '', '', '1.58', '0.00', '1.58', '', '', '', ''),
    (27, 'N/A', '1 x 7891999144485', 'Dinheiro', 'N/A', '', '', '', '6.75', '0.00', '6.75', '', '', '', ''),
]

for row_idx, test_data in enumerate(testes, 8):
    for col_idx, value in enumerate(test_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save('C:/Users/Mike/validaai/automacaoScann/biblioteca/TEMPLATE_COM_BIN_NOVO.xlsx')
print('ETAPA 1 complete with 27 tests')