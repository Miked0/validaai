import openpyxl
wb = openpyxl.load_workbook('biblioteca/TEMPLATE_COM_BIN_NOVO.xlsx')
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'  {name}: {ws.max_row} rows x {ws.max_column} cols')
    for i, row in enumerate(ws.iter_rows(max_row=3, values_only=True)):
        print(f'    Row {i+1}: {list(row)[:10]}')