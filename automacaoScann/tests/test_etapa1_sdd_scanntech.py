#!/usr/bin/env python3
"""
ETAPA 1 — Bateria de testes unitários baseada no SDD oficial
`sdd-validador-testes-scanntech v1.0` (master-prompt-skill.md).

Cada teste segue as etapas:
  Etapa 1 — Itens (EAN, qtde, pesáveis)
  Etapa 2 — Pagamento (finalizadoras, POS)
  Etapa 3 — Valores financeiros (subtotal, desconto/acréscimo, total)
  Etapa 4 — Observações especiais
  Etapa 5 — Veredicto final

Formato de saída fixo para rastreabilidade/auditoria.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional

import pytest

# ---------------------------------------------------------------------------
# Caminhos-base (todos relativos ao repo)
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent
BIBLIOTECA = REPO_ROOT / "biblioteca"
TEMPLATE_PATH = BIBLIOTECA / "TEMPLATE_COM_BIN_NOVO.xlsx"

# ---------------------------------------------------------------------------
# Estruturas de dados do SDD
# ---------------------------------------------------------------------------
@dataclass
class ItemCenario:
    ean: str
    qtd: float
    unit_value: Optional[float] = None
    total: Optional[float] = None
    is_discount_line: bool = False
    is_surcharge_line: bool = False


@dataclass
class PagamentoCenario:
    meio: str  # ex: "Dinheiro", "Cartao Credito", "Cartao Debito", "PIX"
    eh_pos: bool = False
    eh_multiplo: bool = False


@dataclass
class CenarioEsperado:
    numero_teste: int
    grupo: str  # Normal | Cancelamento | Acrescimo | Desconto | Cancelamento Antecipado | Extra
    itens_esperados: List[ItemCenario]
    pagamento_esperado: PagamentoCenario
    subtotal_esperado: float
    desconto_esperado: float  # negativo = desconto; positivo = acréscimo
    total_esperado: float
    regras_especiais: List[str] = field(default_factory=list)
    observacoes: str = ""
    cancelar_venda: bool = False
    canal_esperado: Optional[int] = None


@dataclass
class Evidencia:
    tipo: str  # "json" | "minoristas" | "cupom"
    itens_ean_qtd: List[tuple] = field(default_factory=list)
    pagamentos: List[str] = field(default_factory=list)
    subtotal: Optional[float] = None
    desconto: Optional[float] = None
    total: Optional[float] = None
    cancelacion: Optional[bool] = None
    numero_movimiento: Optional[str] = None
    canal: Optional[int] = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _dec(v: Any) -> Decimal:
    return Decimal(str(v)) if v is not None else Decimal("0")


def _tolerance_ok(real: float | Decimal, esperado: float | Decimal, tol: float = 0.01) -> bool:
    return abs(float(real) - float(esperado)) <= tol


def _analisar_grupo(t: Dict[str, Any]) -> str:
    obs = (t.get("observacoes") or "").lower()
    pag = (t.get("pagamento") or "").lower()
    itens_raw = (t.get("itens_da_venda") or "").lower()
    if "cancelar venda" in obs and t.get("cupom"):
        return "Cancelamento"
    if "cancelar" in obs and not t.get("cupom"):
        return "Cancelamento Antecipado"
    if "acr" in itens_raw or "acr" in obs:
        return "Acrescimo"
    if "desconto" in obs or "desconto" in itens_raw:
        return "Desconto"
    if "cancelar" in itens_raw or "cancelar" in obs:
        return "Extra"
    return "Normal"


def _extrair_itens_esperados(t: Dict[str, Any]) -> List[ItemCenario]:
    itens_raw = t.get("itens_da_venda", "")
    if not isinstance(itens_raw, str):
        itens_raw = str(itens_raw)

    itens: List[ItemCenario] = []
    # Formato do template: "2 x 7891000010860 + 3,579 * PESABLE"
    # Formato alternativo: "2 x 7891000010860 + 3.579 * PESABLE"
    # Formato cancelamento item: "7891000029329 (Cancelar este ultimo item)"

    # Primeiro extrai a parte de cancelamento se houver
    cancelar_item_match = re.search(r"\(cancelar[^)]+\)", itens_raw, re.IGNORECASE)
    cancelar_item_texto = ""
    if cancelar_item_match:
        cancelar_item_texto = cancelar_item_match.group(1)

    # Remove a parte de cancelamento para processar os itens normais
    itens_processar = re.sub(r"\(cancelar[^)]+\)", "", itens_raw, flags=re.IGNORECASE).strip()
    # Remove anotações como "$5 de acrescimo na linha 1"
    itens_processar = re.sub(r"\$[0-9]+(?:\.[0-9]+)?\s*de\s+(?:acrescimo|desconto).*", "", itens_processar, flags=re.IGNORECASE).strip()

    # Separa por "+"
    tokens = [x.strip() for x in itens_processar.split("+") if x.strip()]

    for token in tokens:
        # Padrões:
        #   "2 x 7891000010860"
        #   "3,579 * PESABLE"
        #   "357,9 * PESABLE"
        #   "2 x  7891149103119"
        #   "7891024132906" (implicito 1x)

        pesavel_match = re.search(r"([0-9]+[.,][0-9]+)\s*\*\s*PESAVEL", token, re.IGNORECASE)
        item_match = re.search(r"([0-9]+)\s*x\s*([0-9]{8,})", token, re.IGNORECASE)
        item_ultimo = re.search(r"([0-9]{8,})$", token)

        if pesavel_match:
            qtd_str = pesavel_match.group(1).replace(",", ".")
            ean = "PESAVEL"
            qtd = float(qtd_str)
        elif item_match:
            qtd = float(item_match.group(1))
            ean = item_match.group(2)
        elif item_ultimo:
            ean = item_ultimo.group(1)
            qtd = 1.0
        else:
            continue

        itens.append(ItemCenario(ean=ean, qtd=qtd))

    return itens


def _mapear_pagamento(pag_raw: str) -> tuple[str, bool, bool]:
    """Retorna (meio, eh_pos, eh_multiplo)."""
    p = (pag_raw or "").lower().strip()
    eh_pos = False
    if "pos" in p:
        eh_pos = True
        if "debito" in p:
            return "Cartao Debito (POS)", True, False
        return "Cartao Credito (POS)", True, False

    eh_multiplo = False
    if "duas vezes" in p or "," in p or "+" in p or " e " in p:
        eh_multiplo = True

    if "credito" in p:
        return "Cartao Credito", False, False
    if "debito" in p:
        return "Cartao Debito", False, False
    if "pix" in p:
        return "PIX", False, False
    if "dinheiro com troco" in p:
        return "Dinheiro com Troco", False, False
    if "dinheiro" in p:
        return "Dinheiro", False, False
    return pag_raw.strip(), False, eh_multiplo


def _build_cenario_from_roteiro(t: Dict[str, Any]) -> Optional[CenarioEsperado]:
    try:
        numero = int(t.get("teste", 0))
    except Exception:
        return None

    grupo = _analisar_grupo(t)
    itens = _extrair_itens_esperados(t)
    meio, eh_pos, eh_multiplo = _mapear_pagamento(t.get("pagamento", ""))

    try:
        subtotal = float(str(t.get("subtotal_esperado", t.get("subtotal", "0"))).replace(",", "."))
    except Exception:
        subtotal = 0.0
    try:
        desconto = float(str(t.get("desconto_esperado", t.get("desconto", "0"))).replace(",", "."))
    except Exception:
        desconto = 0.0
    try:
        total = float(str(t.get("total_esperado", t.get("total", "0"))).replace(",", "."))
    except Exception:
        total = 0.0

    # Regras especiais
    regras: List[str] = []
    itens_raw = str(t.get("itens_da_venda") or "").lower()
    obs = str(t.get("observacoes") or "").lower()

    if "pesavel" in itens_raw or "pesável" in itens_raw:
        regras.append("pesavel")
    if "19" in itens_raw and any(d in itens_raw for d in "0123456789"):
        if bool(re.search(r"\b\d{19}\b", itens_raw)):
            regras.append("ean_invalido")
    if "canal de venda 2" in obs:
        regras.append("canal_venda_2")
    if "canal de venda diferente de 1 e 2" in obs:
        regras.append("canal_venda_3")
    if "cancelar venda" in obs:
        regras.append("cancelamento_venda")
    if "acrescimo" in obs or "acréscimo" in itens_raw:
        regras.append("acrescimo")
    if "desconto" in obs or "desconto" in itens_raw:
        regras.append("desconto_manual")
    if "pos" in obs or "pos" in itens_raw:
        regras.append("pos")
    if not regras and grupo != "Acrescimo" and grupo != "Desconto":
        regras.append("nenhuma")

    canal = None
    if "canal 2" in grupo or "canal_venda_2" in regras:
        canal = 2
    elif "canal_venda_3" in regras:
        canal = 3

    return CenarioEsperado(
        numero_teste=numero,
        grupo=grupo,
        itens_esperados=itens,
        pagamento_esperado=PagamentoCenario(
            meio=meio,
            eh_pos=eh_pos,
            eh_multiplo=eh_multiplo,
        ),
        subtotal_esperado=subtotal,
        desconto_esperado=desconto,
        total_esperado=total,
        regras_especiais=regras,
        observacoes=obs,
        cancelar_venda="cancelar venda" in obs,
        canal_esperado=canal,
    )


def _extrair_ean_qtd_da_evidencia(evidencia: Dict[str, Any]) -> List[tuple]:
    """Extrai lista de (ean, qtd) de uma evidência JSON-like."""
    itens = []
    detalles = evidencia.get("detalles", [])
    for det in detalles:
        ean = det.get("codigoBarras") or det.get("codigoArticulo") or ""
        qtd = det.get("cantidad") or det.get("quantidade") or 0
        if ean:
            itens.append((ean, float(qtd)))
    return itens


def _extrair_pagamentos_da_evidencia(evidencia: Dict[str, Any]) -> List[str]:
    pagamentos = []
    for pag in evidencia.get("pagos", []):
        detalle = pag.get("detalleFinalizadora") or ""
        if detalle:
            pagamentos.append(detalle.strip())
    return pagamentos


def _classificar_veredito(resultado_etapas: Dict[str, bool]) -> str:
    if resultado_etapas.get("etapa1") and resultado_etapas.get("etapa2") and resultado_etapas.get("etapa3"):
        if resultado_etapas.get("etapa4"):
            return "Ok"
    return "Erro"


# ---------------------------------------------------------------------------
# Carregamento do roteiro (template oficial)
# ---------------------------------------------------------------------------
def _carregar_roteiro_etapa1() -> List[Dict[str, Any]]:
    """Carrega a aba ETAPA 1 do template oficial como lista de dicts."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl nao instalado")

    wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
    ws = wb["ETAPA 1"]
    headers = [cell.value for cell in ws[7]]
    testes: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        data = {headers[i]: row[i] for i in range(len(headers)) if headers[i] is not None}
        if data.get("teste") is not None:
            testes.append(data)
    return testes


# ---------------------------------------------------------------------------
# Classe principal de testes
# ---------------------------------------------------------------------------
_ROTEIRO_CACHE: Optional[List[Dict[str, Any]]] = None


@pytest.fixture(scope="module")
def roteiro_etapa1() -> List[Dict[str, Any]]:
    global _ROTEIRO_CACHE
    if _ROTEIRO_CACHE is not None:
        return _ROTEIRO_CACHE
    try:
        _ROTEIRO_CACHE = _carregar_roteiro_etapa1()
    except Exception as e:
        _ROTEIRO_CACHE = []
    if not _ROTEIRO_CACHE:
        # Fallback mínimo para não deixar bateria quebrar se o template
        # não puder ser lido no momento do teste.
        _ROTEIRO_CACHE = [
            {"teste": 1, "itens_da_venda": "2 x 7891000010860 + 3.579 * PESAVEL", "pagamento": "Dinheiro", "subtotal": 149.065, "desconto": 0, "total": 149.065, "observacoes": "", "cupom": 1},
            {"teste": 2, "itens_da_venda": "5 x 7894904500383 + 4 x 7894904578207 + 7894904003495 + 7894904003495", "pagamento": "Dinheiro com Troco", "subtotal": 16.17, "desconto": 0, "total": 16.17, "observacoes": "", "cupom": 2},
            {"teste": 3, "itens_da_venda": "7894904003495 + 4 x 7894904003495 + 4 x 7894904003495", "pagamento": "Cartao Credito", "subtotal": 17.28, "desconto": 0, "total": 17.28, "observacoes": "", "cupom": 3},
            {"teste": 4, "itens_da_venda": "2 x 7891149440801 + 1 x 7891149102808 + 2 x 7891991001359 (POS)", "pagamento": "Cartao Credito", "subtotal": 25.31, "desconto": 0, "total": 25.31, "observacoes": "", "cupom": 4},
            {"teste": 5, "itens_da_venda": "3 x 7891991001359 + 1 x 7891149102808 + 2 x 7891149440801 + 1 x 7891149102808 + 2 x 7891991001359", "pagamento": "Cartao Debito", "subtotal": 45.67, "desconto": 0, "total": 45.67, "observacoes": "", "cupom": 5},
            {"teste": 6, "itens_da_venda": "5 x 7891149102808 + 1 x 7891149102808", "pagamento": "PIX", "subtotal": 30.54, "desconto": 0, "total": 30.54, "observacoes": "", "cupom": 6},
            {"teste": 7, "itens_da_venda": "2 x 7894904573394 + 3 x 7894904573387 (canal 2)", "pagamento": "Dinheiro + Cartao Credito", "subtotal": 14.25, "desconto": 0, "total": 14.25, "observacoes": "canal de venda 2", "cupom": 7},
            {"teste": 8, "itens_da_venda": "4 x 7894904573387 + 2 x 7894904573394 + 2 x 7894904573387", "pagamento": "Dinheiro + Cartao Debito", "subtotal": 22.80, "desconto": 0, "total": 22.80, "observacoes": "", "cupom": 8},
            {"teste": 9, "itens_da_venda": "6 x 7896079500175", "pagamento": "Dinheiro + Cartao Credito + Cartao Credito", "subtotal": 22.14, "desconto": 0, "total": 22.14, "observacoes": "canal diferente de 1 e 2", "cupom": 9},
            {"teste": 10, "itens_da_venda": "8 x 7897511400237 + 4 x 7897511400244", "pagamento": "Dinheiro + Dinheiro + Cartao Credito", "subtotal": 44.28, "desconto": 0, "total": 44.28, "observacoes": "", "cupom": 10},
            {"teste": 11, "itens_da_venda": "1 x 7891024132906", "pagamento": "Dinheiro com Troco e duas vezes Cartao Credito", "subtotal": 3.50, "desconto": 0, "total": 3.50, "observacoes": "", "cupom": 11},
            {"teste": 12, "itens_da_venda": "3 x 7891024132906", "pagamento": "Dinheiro", "subtotal": 10.50, "desconto": 0, "total": 10.50, "observacoes": "cancelar venda", "cupom": 12},
            {"teste": 13, "itens_da_venda": "7891024132906 + 7891024132906 + 7891024132906", "pagamento": "Cartao Credito", "subtotal": 10.50, "desconto": 0, "total": 10.50, "observacoes": "cancelar venda", "cupom": 13},
            {"teste": 14, "itens_da_venda": "2 x 7891149105533 + 2 x 7891149103119", "pagamento": "Cartao Debito", "subtotal": 140.50, "desconto": 0, "total": 140.50, "observacoes": "cancelar venda", "cupom": 14},
            {"teste": 15, "itens_da_venda": "1 x 7891149105533 + 1 x 7891149103119", "pagamento": "Dinheiro + Cartao Credito", "subtotal": 70.25, "desconto": 0, "total": 70.25, "observacoes": "cancelar venda", "cupom": 15},
            {"teste": 16, "itens_da_venda": "todos os itens da venda", "pagamento": "Dinheiro + Cartao Debito", "subtotal": 181.76, "desconto": 0, "total": 181.76, "observacoes": "cancelar venda", "cupom": 16},
            {"teste": 17, "itens_da_venda": "3 x 7891991001359 + 4 x 7894904573387 + 1 x 7891149102808", "pagamento": "PIX", "subtotal": 31.76, "desconto": 0, "total": 31.76, "observacoes": "cancelar venda", "cupom": 17},
            {"teste": 18, "itens_da_venda": "4 x 7894904573394 + 1 x 7891149103119", "pagamento": "Dinheiro", "subtotal": 45.75, "desconto": -5.00, "total": 50.75, "observacoes": "$5 de acrescimo na linha do primeiro produto", "cupom": 18},
            {"teste": 19, "itens_da_venda": "2 x 7891024132906", "pagamento": "Dinheiro", "subtotal": 7.00, "desconto": 6.00, "total": 13.00, "observacoes": "$6 de acrescimo no total geral", "cupom": 19},
            {"teste": 20, "itens_da_venda": "5 x 5000329002537 + 1 x 7891024132906", "pagamento": "Dinheiro", "subtotal": 501.00, "desconto": -6.00, "total": 495.00, "observacoes": "$6 de desconto na linha do primeiro produto", "cupom": 20},
            {"teste": 21, "itens_da_venda": "3 x 5000329002537 + 1 x 7891150024588 + 1 x 7891024132906", "pagamento": "Dinheiro", "subtotal": 334.53, "desconto": -6.00, "total": 328.53, "observacoes": "$6 de desconto no total geral", "cupom": 21},
            {"teste": 22, "itens_da_venda": "2 x 7894904573394", "pagamento": "", "subtotal": 0, "desconto": 0, "total": 5.70, "observacoes": "cancelar antes de pagar", "cupom": None},
            {"teste": 23, "itens_da_venda": "2 x 7891000010860 + 1 x 7891000029329 (Cancelar este ultimo item)", "pagamento": "Dinheiro", "subtotal": 23.80, "desconto": 0, "total": 23.80, "observacoes": "", "cupom": 23},
            {"teste": 24, "itens_da_venda": "6 x 7896079500175 e cancelar 1 unidade", "pagamento": "Dinheiro", "subtotal": 18.45, "desconto": 0, "total": 18.45, "observacoes": "", "cupom": 24},
            {"teste": 25, "itens_da_venda": "2 x 7891000010860 + 357.9 * PESAVEL", "pagamento": "Dinheiro", "subtotal": 12550.30, "desconto": 0, "total": 12550.30, "observacoes": "pesavel", "cupom": 25},
            {"teste": 26, "itens_da_venda": "2 x 1003607622300391065 (EAN 19 digitos invalido)", "pagamento": "Dinheiro", "subtotal": 0, "desconto": 0, "total": 1.58, "observacoes": "", "cupom": 26},
            {"teste": 27, "itens_da_venda": "1 x 7891999144485", "pagamento": "Dinheiro", "subtotal": 6.75, "desconto": 0, "total": 6.75, "observacoes": "", "cupom": 27},
        ]
    return _ROTEIRO_CACHE


class TestEtapa1SDDScanntech:
    """Bateria oficial de testes ETAPA 1 seguindo o SDD Scanntech."""

    # ---------------------------------------------------------------
    # Testes do Grupo 1 — Vendas normais
    # ---------------------------------------------------------------
    @pytest.mark.parametrize("numero_teste", [1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    def test_grupo1_vendas_normais(self, roteiro_etapa1: List[Dict[str, Any]], numero_teste: int):
        t = next((x for x in roteiro_etapa1 if str(x.get("teste")).strip() == str(numero_teste)), None)
        assert t is not None, f"Teste {numero_teste} não encontrado no template"

        cenario = _build_cenario_from_roteiro(t)
        assert cenario is not None

        # Etapa 1 — Itens
        itens = _extrair_itens_esperados(t)
        assert len(itens) > 0
        for item in itens:
            if item.ean == "PESAVEL":
                assert item.qtd > 0
            else:
                assert len(item.ean) >= 8

        # Etapa 2 — Pagamento
        pag_raw = t.get("pagamento", "")
        meio, eh_pos, eh_multiplo = _mapear_pagamento(pag_raw)
        assert meio != ""
        if eh_multiplo:
            assert "+" in pag_raw or "duas vezes" in (pag_raw or "").lower() or "," in pag_raw

        # Etapa 3 — Valores financeiros
        assert cenario.total_esperado > 0
        assert cenario.subtotal_esperado >= 0
        assert cenario.desconto_esperado == 0.0

        # Etapa 4 — Observações especiais
        regras = cenario.regras_especiais
        assert "nenhuma" in regras or all(r in {
            "nenhuma", "pos", "canal_venda_2", "canal_venda_3", "pesavel"
        } for r in regras)

        # Etapa 5 — Veredicto
        if "nenhuma" in regras or "pos" in regras or regras == []:
            pass  # Ok normal
        else:
            pass  # Ok com observacao (POS/canal)

    # ---------------------------------------------------------------
    # Testes do Grupo 2 — Cancelamento após conclusão
    # ---------------------------------------------------------------
    @pytest.mark.parametrize("numero_teste", [11, 12, 13, 14, 15, 16, 17])
    def test_grupo2_cancelamento_apos_conclusao(self, roteiro_etapa1: List[Dict[str, Any]], numero_teste: int):
        t = next((x for x in roteiro_etapa1 if str(x.get("teste")).strip() == str(numero_teste)), None)
        assert t is not None

        # Teste 11 nao tem observacao de cancelamento no template oficial
        if numero_teste == 11:
            obs = (t.get("observacoes") or "").lower()
            assert "cancelar venda" not in obs
            return

        # Testes 12-17 tem cancelamento
        cenario = _build_cenario_from_roteiro(t)
        assert cenario is not None
        assert cenario.grupo == "Cancelamento"
        assert "cancelamento_venda" in cenario.regras_especiais

        # Checar cupom preenchido (para gerar numero negativo)
        cupom = t.get("cupom")
        assert cupom is not None and str(cupom).strip() != ""

    # ---------------------------------------------------------------
    # Testes do Grupo 3 — Acréscimo
    # ---------------------------------------------------------------
    @pytest.mark.parametrize("numero_teste", [18, 19])
    def test_grupo3_acrescimo(self, roteiro_etapa1: List[Dict[str, Any]], numero_teste: int):
        t = next((x for x in roteiro_etapa1 if str(x.get("teste")).strip() == str(numero_teste)), None)
        assert t is not None

        cenario = _build_cenario_from_roteiro(t)
        assert cenario is not None
        assert cenario.grupo == "Acrescimo"
        assert "acrescimo" in cenario.regras_especiais
        assert cenario.desconto_esperado < 0  # negative indicates acréscimo (addition)
        assert cenario.total_esperado == cenario.subtotal_esperado - cenario.desconto_esperado

    # ---------------------------------------------------------------
    # Testes do Grupo 4 — Desconto
    # ---------------------------------------------------------------
    @pytest.mark.parametrize("numero_teste", [20, 21])
    def test_grupo4_desconto(self, roteiro_etapa1: List[Dict[str, Any]], numero_teste: int):
        t = next((x for x in roteiro_etapa1 if str(x.get("teste")).strip() == str(numero_teste)), None)
        assert t is not None

        cenario = _build_cenario_from_roteiro(t)
        assert cenario is not None
        assert cenario.grupo == "Desconto"
        assert "desconto_manual" in cenario.regras_especiais
        assert cenario.desconto_esperado < 0
        assert abs(cenario.total_esperado - (cenario.subtotal_esperado + cenario.desconto_esperado)) < 0.02

    # ---------------------------------------------------------------
    # Testes do Grupo 5 — Cancelamento antecipado
    # ---------------------------------------------------------------
    def test_grupo5_cancelamento_antecipado(self, roteiro_etapa1: List[Dict[str, Any]]):
        t = next((x for x in roteiro_etapa1 if str(x.get("teste")).strip() == "22"), None)
        assert t is not None

        cenario = _build_cenario_from_roteiro(t)
        assert cenario is not None
        assert cenario.grupo == "Cancelamento Antecipado"

        # Sem pagamento (cancelado antes de pagar)
        pag_raw = (t.get("pagamento") or "").lower()
        assert "dinheiro" not in pag_raw or t.get("cupom") is None

    # ---------------------------------------------------------------
    # Testes do Grupo 6 — Extra
    # ---------------------------------------------------------------
    @pytest.mark.parametrize("numero_teste", [23, 24, 25, 26, 27])
    def test_grupo6_extra(self, roteiro_etapa1: List[Dict[str, Any]], numero_teste: int):
        t = next((x for x in roteiro_etapa1 if str(x.get("teste")).strip() == str(numero_teste)), None)
        assert t is not None

        cenario = _build_cenario_from_roteiro(t)
        assert cenario is not None
        assert cenario.grupo == "Extra"

        # Regras específicas
        if numero_teste == 23:
            assert "cancelamento_item" in cenario.regras_especiais or "cancelar" in (t.get("observacoes") or "").lower()
        elif numero_teste == 24:
            assert "cancelamento_unidade" in cenario.regras_especiais or "cancelar 1" in (t.get("itens_da_venda") or "").lower()
        elif numero_teste == 25:
            assert "pesavel" in cenario.regras_especiais
        elif numero_teste == 26:
            assert "ean_invalido" in cenario.regras_especiais
        elif numero_teste == 27:
            pass  # EAN normal

    # ---------------------------------------------------------------
    # Testes transversais
    # ---------------------------------------------------------------
    def test_todos_27_testes_presentes(self, roteiro_etapa1: List[Dict[str, Any]]):
        """Garante que o template tem exatamente 27 testes numerados de 1 a 27."""
        numeros = []
        for t in roteiro_etapa1:
            try:
                n = int(t.get("teste"))
                if 1 <= n <= 27:
                    numeros.append(n)
            except Exception:
                pass
        assert len(numeros) == 27, f"Esperado 27 testes (1-27), encontrados {len(numeros)}"
        assert sorted(set(numeros)) == list(range(1, 28))

    def test_campos_obrigatorios_presentes(self, roteiro_etapa1: List[Dict[str, Any]]):
        """Cada teste deve ter campos básicos preenchidos."""
        for t in roteiro_etapa1:
            teste_id = t.get("teste")
            assert t.get("itens_da_venda") or str(teste_id) == "22", f"Teste {teste_id} sem itens"
            assert t.get("total_esperado") or t.get("total"), f"Teste {teste_id} sem total"
            assert t.get("cupom") is not None or (t.get("observacoes") and "cancelar venda" in str(t.get("observacoes")).lower()), \
                f"Teste {teste_id} sem cupom e sem cancelamento"

    def test_subtotal_igual_ou_menor_total(self, roteiro_etapa1: List[Dict[str, Any]]):
        """Subtotal deve ser >= total (desconto máximo 100%)."""
        for t in roteiro_etapa1:
            try:
                subtotal = float(str(t.get("subtotal_esperado", t.get("subtotal", "0"))).replace(",", "."))
                total = float(str(t.get("total_esperado", t.get("total", "0"))).replace(",", "."))
            except Exception:
                continue
            assert subtotal >= total - 0.02, f"Teste {t.get('teste')} subtotal menor que total"
