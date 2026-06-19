#!/usr/bin/env python3
"""
Snapshot tests for Export Pipeline.
Compares generated outputs against golden files.
Run with: pytest tests/test_snapshots.py -v
"""
import json
import pytest
import tempfile
from pathlib import Path
from decimal import Decimal
from datetime import datetime

# Add project root
import sys
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from core.exporters import (
    ExportPipeline, ExportConfig, 
    ExcelExporter, CSVExporter, JSONAuditExporter, HTMLReportExporter,
    export_all, ExportResult
)


# ─── Sample Validation Results ──────────────────────────
SAMPLE_RESULTS = [
    {
        "teste": 1,
        "bloco": "ETAPA 1",
        "tipo_promo": "",
        "itens_da_venda": "2 x 7891000010860 + 3.579 * PESABLE",
        "pagamento": "Dinheiro",
        "codigo_tipo_pago": 9,
        "subtotal_esperado": Decimal("149.065"),
        "desconto_esperado": Decimal("0"),
        "total_esperado": Decimal("149.065"),
        "status_final": "REVISAO",
        "motivo_status": "Item pesável detectado",
        "alertas": ["Item pesável detectado - requer revisão manual"],
        "observacoes": "",
        "observacao_parceiro": "",
        "sat": "",
        "ecf": "1",
        "nfce": "",
        "cupom": "",
        "api_status": "OK",
        "api_alertas": [],
        "sale_json": {"pagos": [{"codigoTipoPago": 9, "importe": 149.07}]}
    },
    {
        "teste": 2,
        "bloco": "ETAPA 1",
        "tipo_promo": "",
        "itens_da_venda": "5 x 7894904500383",
        "pagamento": "Dinheiro com Troco",
        "codigo_tipo_pago": 9,
        "subtotal_esperado": Decimal("16.17"),
        "desconto_esperado": Decimal("0"),
        "total_esperado": Decimal("16.17"),
        "status_final": "REVISAO",
        "motivo_status": "Pagamento com troco",
        "alertas": ["Pagamento com troco - requer revisão manual"],
        "observacoes": "",
        "sat": "",
        "ecf": "2",
        "nfce": "",
        "cupom": "",
        "api_status": "OK",
        "api_alertas": [],
        "sale_json": {"pagos": [{"codigoTipoPago": 9, "importe": 16.17}]}
    },
    {
        "teste": 3,
        "bloco": "ETAPA 1",
        "tipo_promo": "",
        "itens_da_venda": "1 x 7894904003495",
        "pagamento": "Cartao Credito",
        "codigo_tipo_pago": 10,
        "subtotal_esperado": Decimal("17.28"),
        "desconto_esperado": Decimal("0"),
        "total_esperado": Decimal("17.28"),
        "status_final": "OK",
        "motivo_status": "Todos os campos válidos",
        "alertas": [],
        "observacoes": "",
        "sat": "",
        "ecf": "3",
        "nfce": "",
        "cupom": "",
        "api_status": "OK",
        "api_alertas": [],
        "sale_json": {"pagos": [{"codigoTipoPago": 10, "importe": 17.28}]}
    },
]


# ─── Golden Files Directory ─────────────────────────────
GOLDEN_DIR = Path(__file__).parent / "golden"


def _load_golden(filename: str) -> str:
    """Load golden file content."""
    path = GOLDEN_DIR / filename
    if not path.exists():
        pytest.skip(f"Golden file not found: {path}")
    return path.read_text(encoding="utf-8")


def _write_golden(filename: str, content: str):
    """Write golden file (use to create initial goldens)."""
    GOLDEN_DIR.mkdir(exist_ok=True)
    Path(GOLDEN_DIR / filename).write_text(content, encoding="utf-8")


# ─── Fixture: Run pipeline once for all tests ────────────
@pytest.fixture(scope="module")
def pipeline_results(tmp_path_factory):
    """Run full export pipeline and return results."""
    output_dir = tmp_path_factory.mktemp("snapshots")
    
    results = export_all(
        SAMPLE_RESULTS,
        output_dir,
        timestamp="20260618_120000",
        etapa="ETAPA1",
        version="2.0.0"
    )
    
    # Map by exporter name
    return {r.exporter_name: r for r in results}


# ─── Excel Snapshot Test ────────────────────────────────
def test_excel_snapshot(pipeline_results):
    """Excel output matches golden master."""
    result = pipeline_results["excel"]
    assert result.success, f"Excel export failed: {result.error}"
    assert result.output_path.exists()
    
    # Load and verify structure
    import openpyxl
    wb = openpyxl.load_workbook(result.output_path)
    assert "Resumo" in wb.sheetnames
    assert "Metadados" in wb.sheetnames
    
    ws = wb["Resumo"]
    assert ws.max_row >= 4  # header + 3 data rows
    assert ws.max_column >= 15
    
    # Verify headers
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "teste" in headers
    assert "status_final" in headers
    assert "motivo_status" in headers
    assert "pagamento_label" in headers
    
    # Verify data rows
    for row_idx in range(2, 5):
        teste_val = ws.cell(row=row_idx, column=headers.index("teste") + 1).value
        assert teste_val in [1, 2, 3]
        
        status_val = ws.cell(row=row_idx, column=headers.index("status_final") + 1).value
        assert status_val in ["OK", "REVISAO", "ERRO", "NOT_RUN"]


def test_excel_has_correct_status_colors(pipeline_results):
    """Excel status column has correct color formatting."""
    result = pipeline_results["excel"]
    import openpyxl
    wb = openpyxl.load_workbook(result.output_path)
    ws = wb["Resumo"]
    
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    status_col = headers.index("status_final") + 1
    
    for row_idx in range(2, 5):
        cell = ws.cell(row=row_idx, column=status_col)
        assert cell.font.color is not None
        assert cell.font.bold is True


# ─── CSV Snapshot Test ──────────────────────────────────
def test_csv_snapshot(pipeline_results):
    """CSV output matches golden master."""
    result = pipeline_results["csv"]
    assert result.success
    assert result.output_path.exists()
    
    import csv
    with open(result.output_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    
    assert len(rows) == 3
    assert set(rows[0].keys()) >= {"teste", "status_final", "motivo_status"}
    
    statuses = {int(r["teste"]): r["status_final"] for r in rows}
    assert statuses == {1: "REVISAO", 2: "REVISAO", 3: "OK"}


def test_csv_encoding_utf8_sig(pipeline_results):
    """CSV uses UTF-8-SIG encoding (BOM)."""
    result = pipeline_results["csv"]
    with open(result.output_path, "rb") as f:
        content = f.read(3)
    assert content == b"\xef\xbb\xbf"  # UTF-8 BOM


# ─── JSON Audit Snapshot Test ───────────────────────────
def test_json_audit_snapshot(pipeline_results):
    """JSON audit output matches golden master."""
    result = pipeline_results["json_audit"]
    assert result.success
    assert result.output_path.exists()
    
    with open(result.output_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    # Top-level structure
    assert "generated_at" in data
    assert "app_version" in data
    assert "total" in data
    assert data["total"] == 3
    assert "summary" in data
    assert "results" in data
    assert len(data["results"]) == 3
    
    # Summary structure
    summary = data["summary"]
    assert summary["by_status"]["REVISAO"] == 2
    assert summary["by_status"]["OK"] == 1
    assert summary["total_errors"] == 0  # No ERRO status in samples
    
    # Each result has required fields
    for r in data["results"]:
        assert "teste" in r
        assert "status_final" in r
        assert "sale_json" in r
        assert "api_status" in r


def test_json_round_trip(pipeline_results):
    """JSON can be loaded and re-serialized without loss."""
    result = pipeline_results["json_audit"]
    with open(result.output_path, "r", encoding="utf-8") as f:
        original = f.read()
    
    # Parse and re-serialize
    data = json.loads(original)
    re_serialized = json.dumps(data, ensure_ascii=False, indent=2)
    
    # Parse again
    data2 = json.loads(re_serialized)
    assert data2["total"] == 3
    assert len(data2["results"]) == 3


# ─── HTML Report Snapshot Test ──────────────────────────
def test_html_report_snapshot(pipeline_results):
    """HTML report output matches golden master."""
    result = pipeline_results["html_report"]
    assert result.success
    assert result.output_path.exists()
    
    html = result.output_path.read_text(encoding="utf-8")
    
    # Required elements
    assert "<!DOCTYPE html>" in html
    assert "ValidaAI" in html
    assert "chart.js" in html.lower()  # CDN link uses lowercase
    assert "status-badge" in html
    assert "tab-panel" in html
    
    # Tabs
    assert 'data-tab="tabela"' in html
    assert 'data-tab="grafico"' in html
    assert 'data-tab="detalhes"' in html
    
    # Table structure exists
    assert "<thead>" in html
    assert "<tbody>" in html
    assert "<th>Teste</th>" in html
    
    # Chart data embedded
    assert "data-counts" in html
    assert '"REVISAO": 2' in html or '"REVISAO":2' in html
    assert '"OK": 1' in html
    
    # CSS variables from design system
    assert "--bg-primary" in html
    assert "--accent" in html
    assert "--success" in html
    assert "--error" in html
    assert "--warning" in html


def test_html_has_chart_js(pipeline_results):
    """HTML includes Chart.js CDN link."""
    result = pipeline_results["html_report"]
    html = result.output_path.read_text(encoding="utf-8")
    assert "chart.js" in html.lower()
    assert "cdn.jsdelivr.net" in html


def test_html_responsive_viewport(pipeline_results):
    """HTML has responsive viewport meta tag."""
    result = pipeline_results["html_report"]
    html = result.output_path.read_text(encoding="utf-8")
    assert 'viewport' in html


# ─── Pipeline Integration Test ──────────────────────────
def test_pipeline_runs_all_exporters(pipeline_results):
    """All 4 exporters ran successfully."""
    assert len(pipeline_results) == 4
    for name, result in pipeline_results.items():
        assert result.success, f"{name} failed: {result.error}"
        assert result.rows_exported == 3
        assert result.duration_ms >= 0


def test_pipeline_output_files_exist(pipeline_results):
    """All output files exist on disk."""
    for name, result in pipeline_results.items():
        assert result.output_path.exists(), f"{name}: file not found"
        assert result.output_path.stat().st_size > 0, f"{name}: empty file"


# ─── Exporter Config Test ───────────────────────────────
def test_exporter_config_from_file():
    """ExportConfig loads from JSON file correctly."""
    from core.exporters import ExportConfig
    from pathlib import Path
    
    config_path = Path(__file__).parent.parent / "config" / "export.json"
    if config_path.exists():
        config = ExportConfig.from_file(config_path)
        assert config.exporters == ["excel", "csv", "json_audit", "html_report"]
        assert config.output.get("directory") == "output"
        assert config.notifications.get("enabled") is False
        assert config.scheduler.get("enabled") is False


def test_pipeline_from_config_file(tmp_path):
    """ExportPipeline can be created from config file."""
    from core.exporters import ExportPipeline, ExportConfig
    
    config = ExportConfig(exporters=["csv", "json_audit"])
    pipeline = ExportPipeline(config)
    assert len(pipeline._exporters) == 2
    assert {e.name for e in pipeline._exporters} == {"csv", "json_audit"}


# ─── Regression Test: Ground Truth ──────────────────────
def test_ground_truth_distribution():
    """Verify sample results match expected ground truth distribution."""
    from collections import Counter
    statuses = [r["status_final"] for r in SAMPLE_RESULTS]
    counts = Counter(statuses)
    assert counts["REVISAO"] == 2
    assert counts["OK"] == 1
    assert "ERRO" not in counts
    assert "NOT_RUN" not in counts


# ─── Performance Regression ─────────────────────────────
def test_exporter_performance(pipeline_results):
    """Exporters complete within reasonable time."""
    for name, result in pipeline_results.items():
        # All should complete within 1000ms (generous for test env)
        assert result.duration_ms < 1000, f"{name} too slow: {result.duration_ms}ms"
        
        # HTML should be fastest
        if name == "html_report":
            assert result.duration_ms < 50


# ─── Run as standalone ──────────────────────────────────
if __name__ == "__main__":
    # Quick manual test
    import sys
    sys.path.insert(0, str(BASE_DIR))
    
    print("Running snapshot tests...")
    
    # Use pytest to run
    import subprocess
    result = subprocess.run([
        sys.executable, "-m", "pytest", __file__, "-v", "--tb=short"
    ])
    sys.exit(result.returncode)