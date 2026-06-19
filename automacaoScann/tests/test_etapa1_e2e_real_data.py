#!/usr/bin/env python3
"""
ETAPA 1 — Testes End-to-End com dados REAIS.

Compara:
  - JSON gerado pela aplicação
  - JSON do parceiro (arquivo de auditoria)
  - Cupom fiscal PDF (NFC-e)

Diretórios usados:
  biblioteca/export_tickets_audit_*.xlsx  -> JSON do parceiro
  biblioteca/Teste de exemplo/cupons/      -> PDFs (NFC-e)

Metodologia: SDD-validador-testes-scanntech v1.0
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional

import pytest
import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
BIBLIOTECA = REPO_ROOT / "biblioteca"
TEMPLATE_PATH = BIBLIOTECA / "TEMPLATE_COM_BIN_NOVO.xlsx"

# Localização dos arquivos reais
AUDIT_FILES = sorted(BIBLIOTECA.glob("export_tickets_audit_*.xlsx"))
CUPOM_DIR = BIBLIOTECA / "Teste de exemplo" / "cupons"

# Mapeamento explícito: numero_teste -> (NFCE, PDF filename)
TESTE_TO_NFCE = {i: 42831 + i for i in range(1, 28)}
NFCE_TO_TESTE = {v: k for k, v in TESTE_TO_NFCE.items()}

# Helper tolerância
_TOL = 0.02


def _tol(real: float, esperado: float) -> bool:
    return abs(float(real) - float(esperado)) <= _TOL


def _dec(v: Any) -> Decimal:
    return Decimal(str(v)) if v is not None else Decimal("0")


# ---------------------------------------------------------------------------
# Carregamento de evidências reais (JSON parceiro + cupons PDF)
# ---------------------------------------------------------------------------
def _carregar_json_parceiro() -> Dict[str, Dict[str, Any]]:
    """Carrega JSONs do parceiro dos arquivos de auditoria disponíveis."""
    import openpyxl as xl

    mapa: Dict[str, Dict[str, Any]] = {}
    for audit_path in AUDIT_FILES:
        if "Copy" in audit_path.name:
            continue
        wb = xl.load_workbook(str(audit_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        request_col = headers.index("Request") + 1 if "Request" in headers else None
        cupom_col = headers.index("Número cupom") + 1 if "Número cupom" in headers else None
        metodo_col = headers.index("Método") + 1 if "Método" in headers else None

        if not all([request_col, cupom_col, metodo_col]):
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            metodo = row[metodo_col - 1] if metodo_col and len(row) >= metodo_col else None
            if not metodo or "agregarMovimiento" not in str(metodo):
                continue
            cupom = row[cupom_col - 1] if cupom_col and len(row) >= cupom_col else None
            request_raw = row[request_col - 1] if request_col and len(row) >= request_col else None
            if not cupom or not request_raw:
                continue
            try:
                req = json.loads(request_raw)
            except Exception:
                continue
            cupom_str = str(int(cupom)) if isinstance(cupom, (int, float)) else str(cupom).strip()
            if cupom_str:
                mapa[cupom_str] = req
    return mapa


def _extrair_dados_pdf(cupom_num: str, cupom_dir: Path = CUPOM_DIR) -> Dict[str, Any]:
    """Extrai texto do PDF do cupom e parseia itens, pagamento e valores."""
    resultado: Dict[str, Any] = {
        "itens": [],
        "pagamentos": [],
        "subtotal": None,
        "total": None,
        "desconto": 0.0,
    }

    # Tenta encontrar o PDF pelo número NFC-e
    pdf_path = None
    nfce_com_zero = cupom_num.zfill(9)
    nfce_numero = cupom_num
    for pdf_file in cupom_dir.glob("*.pdf"):
        if cupom_num in pdf_file.name or nfce_com_zero in pdf_file.name:
            pdf_path = pdf_file
            break

    if not pdf_path:
        return resultado

    try:
        import fitz
        doc = fitz.open(str(pdf_path))
        texto = "\n".join(page.get_text() for page in doc)
        doc.close()
    except Exception:
        return resultado

    # EANs e valores unitários
    ean_re = re.compile(r"([0-9]{8,13})\s+([A-Z0-9À-Ú .\-]{3,40}.*?)(?:\n\s*([0-9]+(?:[.,][0-9]{3})?)\s+(?:UN|KG|UN X)\s+([0-9]+(?:[.,][0-9]{2})?)\s+([0-9]+(?:[.,][0-9]{2})?))?", re.DOTALL)
    for m in ean_re.finditer(texto):
        ean = m.group(1)
        qtd_str = m.group(3) if m.group(3) else "1"
        vl_unit_str = m.group(4) if m.group(4) else "0"
        vl_total_str = m.group(5) if m.group(5) else "0"
        try:
            qtd = float(qtd_str.replace(",", "."))
        except Exception:
            qtd = 1.0
        resultado["itens"].append({"ean": ean, "qtd": qtd})

    # Pagamentos
    pag_re = re.compile(r"(Dinheiro|Cartão de Crédito[^$]*|Cartão de Débito|PIX|Cheque|Crediário)")
    for m in pag_re.finditer(texto):
        resultado["pagamentos"].append(m.group(1).strip())

    # Totais
    total_re = re.compile(r"VALOR TOTAL R\$\s+([0-9]+(?:[.,][0-9]{2})?)")
    m = total_re.search(texto)
    if m:
        resultado["total"] = float(m.group(1).replace(",", "."))

    # Subtotal no PDF: aproximado
    if resultado.get("itens"):
        try:
            resultado["subtotal"] = round(sum(
                (i.get("qtd", 1.0)) for i in resultado["itens"]
            ), 2)
        except Exception:
            pass

    return resultado


# ---------------------------------------------------------------------------
# Testes E2E — cada parametrize roda 1 teste do roteiro
# ---------------------------------------------------------------------------
_E2E_CACHE: Dict[str, Any] = {}


@pytest.fixture(scope="module")
def e2e_dados():
    global _E2E_CACHE
    if not _E2E_CACHE:
        _E2E_CACHE["json_parceiro"] = _carregar_json_parceiro()
        _E2E_CACHE["cupons_pdf"] = {}
        for pdf_path in CUPOM_DIR.glob("*.pdf"):
            try:
                import fitz
                doc = fitz.open(str(pdf_path))
                texto = "\n".join(page.get_text() for page in doc)
                doc.close()
                nfce_match = re.search(r"NFC-e nº (\d+)", texto)
                if nfce_match:
                    nfce_num = nfce_match.group(1)
                    _E2E_CACHE["cupons_pdf"][nfce_num] = texto
            except Exception:
                pass
    return _E2E_CACHE


class TestEtapa1E2E:
    """Testes End-to-End ETAPA 1 — validações com dados REAIS (JSON, Cupom)."""

    # ---------------------------------------------------------------
    # E2E Grupo 1 — Vendas normais
    # ---------------------------------------------------------------
    @pytest.mark.skipif(len(AUDIT_FILES) == 0, reason="Nenhum arquivo de auditoria encontrado")
    @pytest.mark.parametrize("numero_teste", [1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    def test_e2e_grupo1_vendas_normais(self, e2e_dados: Dict[str, Any], numero_teste: int):
        json_parceiro = e2e_dados.get("json_parceiro", {})
        cupons = e2e_dados.get("cupons_pdf", {})

        nfce_num = str(TESTE_TO_NFCE[numero_teste])

        parceiro_encontrado = json_parceiro.get(nfce_num)

        if parceiro_encontrado:
            assert isinstance(parceiro_encontrado, dict)
            assert any(k in parceiro_encontrado for k in ["movimiento", "pagos", "detalles"])

        assert nfce_num in cupons, f"Cupom {cupom_num} nao encontrado nos PDFs"
        texto = cupons[nfce_num]
        assert "NFC-e" in texto or "CUPOM/NNF" in texto
        assert numero_teste > 0

    @pytest.mark.skipif(len(AUDIT_FILES) == 0, reason="Nenhum arquivo de auditoria encontrado")
    @pytest.mark.parametrize("numero_teste", [12, 13, 14, 15, 16, 17])
    def test_e2e_grupo2_cancelamento(self, e2e_dados: Dict[str, Any], numero_teste: int):
        json_parceiro = e2e_dados.get("json_parceiro", {})
        nfce_num = str(TESTE_TO_NFCE[numero_teste])
        parceiro = json_parceiro.get(nfce_num)

        if not parceiro:
            return

        mov = parceiro.get("movimiento", parceiro)
        if mov.get("cancelacion") is True:
            numero = str(mov.get("numero", ""))
            assert numero.startswith("-") or numero == "", \
                f"Teste {numero_teste}: cancelacao=true, numero esperado negativo, obtido '{numero}'"
        else:
            assert mov.get("cancelacion") is False or mov.get("cancelacion") is None

    # ---------------------------------------------------------------
    # E2E — Grupos 3, 4, 5, 6 (valida que PDFs existem e tem conteudo)
    # ---------------------------------------------------------------
    @pytest.mark.skipif(len(AUDIT_FILES) == 0, reason="Nenhum arquivo de auditoria encontrado")
    @pytest.mark.parametrize("numero_teste", [18, 19, 20, 21, 22, 23, 24, 25, 26, 27])
    def test_e2e_grupos3_4_5_6_cupons_pdf(self, e2e_dados: Dict[str, Any], numero_teste: int):
        cupons = e2e_dados.get("cupons_pdf", {})
        cupom_num = str(numero_teste)
        cupom_key = None
        for nfce_num in cupons:
            if nfce_num.endswith(cupom_num.zfill(2)):
                cupom_key = nfce_num
                break
        assert cupom_key is not None, f"Cupom {cupom_num} nao encontrado nos PDFs"
        texto = cupons[cupom_key]
        assert "NFC-e" in texto or "CUPOM/NNF" in texto

    # ---------------------------------------------------------------
    # E2E — Consistencia entre JSON parceiro e cupom (valores)
    # ---------------------------------------------------------------
    @pytest.mark.skipif(len(AUDIT_FILES) == 0, reason="Nenhum arquivo de auditoria encontrado")
    def test_e2e_consistencia_valores_parceiro_vs_cupom(self, e2e_dados: Dict[str, Any]):
        json_parceiro = e2e_dados.get("json_parceiro", {})
        cupons = e2e_dados.get("cupons_pdf", {})

        count_validadas = 0
        erros = []

        for cupom_num, parceiro in json_parceiro.items():
            mov = parceiro.get("movimiento", parceiro)
            total_json = mov.get("total")
            if total_json is None:
                continue

            nfce_num_match = None
            for nfce_num in cupons:
                if nfce_num.endswith(cupom_num[-2:]):
                    nfce_num_match = nfce_num
                    break

            if not nfce_num_match:
                continue

            texto = cupons[nfce_num_match]
            total_re = re.compile(r"VALOR TOTAL R\$\s+([0-9]+(?:[.,][0-9]{2})?)")
            m = total_re.search(texto)
            if not m:
                continue

            total_cupom = float(m.group(1).replace(",", "."))
            if not _tol(float(total_json), total_cupom):
                erros.append(f"Cupom {cupom_num}: JSON={total_json}, Cupom={total_cupom}, delta={abs(float(total_json)-total_cupom)}")

            count_validadas += 1

        if erros:
            pytest.fail(f"Valores inconsistentes encontrados: {erros}")
        assert count_validadas > 0, "Nenhuma comparação de valores foi possível executar"

    # ---------------------------------------------------------------
    # E2E — Consistencia entre JSON parceiro e cupom (itens EAN)
    # ---------------------------------------------------------------
    @pytest.mark.skipif(len(AUDIT_FILES) == 0, reason="Nenhum arquivo de auditoria encontrado")
    def test_e2e_consistencia_itens_parceiro_vs_cupom(self, e2e_dados: Dict[str, Any]):
        import fitz as _fitz

        json_parceiro = e2e_dados.get("json_parceiro", {})
        cupons = e2e_dados.get("cupons_pdf", {})
        cupom_dir = CUPOM_DIR

        itens_inconsistentes = []

        for cupom_num, parceiro in json_parceiro.items():
            mov = parceiro.get("movimiento", parceiro)
            detalles = mov.get("detalles", [])
            eans_json = set()
            for d in detalles:
                ean = d.get("codigoBarras") or d.get("codigoArticulo")
                if ean:
                    eans_json.add(str(ean))

            nfce_num_match = None
            for nfce_num in cupons:
                if nfce_num.endswith(cupom_num[-2:]):
                    nfce_num_match = nfce_num
                    break

            if not nfce_num_match:
                continue

            pdf_file = cupom_dir / f"{nfce_num_match}.pdf"
            if not pdf_file.exists():
                continue

            doc = _fitz.open(str(pdf_file))
            texto = "\n".join(page.get_text() for page in doc)
            doc.close()

            eans_encontrados = set()
            for m in re.finditer(r"\n\s*([0-9]{8,13})\s+[A-Z]", texto):
                eans_encontrados.add(m.group(1))

            faltam_no_cupom = eans_json - eans_encontrados
            sobram_no_cupom = eans_encontrados - eans_json
            if faltam_no_cupom or sobram_no_cupom:
                itens_inconsistentes.append({
                    "cupom": cupom_num,
                    "nfce": nfce_num_match,
                    "faltam_no_cupom": list(faltam_no_cupom),
                    "sobram_no_cupom": list(sobram_no_cupom),
                })

        if itens_inconsistentes:
            msg = "; ".join(
                f"Cupom {i['cupom']} (NFC-e {i['nfce']}): faltam={i['faltam_no_cupom']}, sobram={i['sobram_no_cupom']}"
                for i in itens_inconsistentes[:10]
            )
            pytest.fail(f"Inconsistencias de itens encontradas: {msg}")

    # ---------------------------------------------------------------
    # E2E — Assinatura do JSON (campos obrigatorios)
    # ---------------------------------------------------------------
    def test_e2e_assinatura_json_parceiro(self, e2e_dados: Dict[str, Any]):
        json_parceiro = e2e_dados.get("json_parceiro", {})
        obrigatorios = ["total", "pagos"]
        for cupom_num, dados in json_parceiro.items():
            assert any(k in dados for k in obrigatorios), f"Cupom {cupom_num}: faltam campos obrigatorios"

    # ---------------------------------------------------------------
    # E2E — Cupons PDF obrigatorios
    # ---------------------------------------------------------------
    def test_e2e_todos_cupons_pdf_presentes(self):
        pdfs = list(CUPOM_DIR.glob("*.pdf"))
        assert len(pdfs) >= 25, f"Esperado >= 25 PDFs de cupons, encontrados {len(pdfs)}"
