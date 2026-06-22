#!/usr/bin/env python3
"""
ValidaAI Test 2 Validation Runner
Loads custom template + 30 audit JSONs, runs validation pipeline.
"""

import json
import sys
sys.path.insert(0, 'src')

from validaai.reader import TestScriptReader
from validaai.parser_items import ItemParser
from validaai.payments import PaymentNormalizer
from validaai.api_sales import APISalesBuilder
from validaai.validators import TestValidator

# ============================================================
# 1. LOAD AUDIT JSONS (partner JSONs from audit export)
# ============================================================
print("=" * 60)
print("LOADING AUDIT JSONS")
print("=" * 60)

import pandas as pd

df_audit = pd.read_excel('biblioteca/teste 2/export_tickets_audit_companyId=74866_auditDate=2026-06-11_a99383_18-06-2026_00-00.xlsx', header=None)

partner_jsons = {}
for i in range(1, len(df_audit)):
    row = df_audit.iloc[i]
    cupom = str(row[7]).strip()
    metodo = str(row[5]).strip()
    request_json_str = str(row[18]).strip()
    
    if metodo != 'agregarMovimiento':
        continue
    if not cupom or cupom.lower() in ['nan', 'none', '']:
        continue
    if not request_json_str or request_json_str == '{ }':
        continue
        
    try:
        request_json = json.loads(request_json_str)
        partner_jsons[cupom] = request_json
        print(f"  Cupom {cupom}: JSON loaded OK")
    except Exception as e:
        print(f"  Cupom {cupom}: ERROR parsing JSON: {e}")

print(f"\nTotal partner JSONs loaded: {len(partner_jsons)}")

# ============================================================
# 2. LOAD TEMPLATE TESTS (ETAPA 1)
# ============================================================
print("\n" + "=" * 60)
print("LOADING TEMPLATE TESTS (ETAPA 1)")
print("=" * 60)

reader = TestScriptReader('biblioteca/teste 2/TEMPLATE_COM_BIN_NOVO (2).xlsx')
reader.set_etapa('ETAPA 1')
tests = reader.read_tests()

# The reader uses 'cupom' field from _first_nonempty which may pick wrong column
# Let's fix by using the NFCE column directly
# Re-read with manual extraction for cupom
import openpyxl
wb = openpyxl.load_workbook('biblioteca/teste 2/TEMPLATE_COM_BIN_NOVO (2).xlsx')
ws = wb['ETAPA 1']

# Find header row
header = None
start_idx = None
for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    vals = [str(c).strip() if c is not None else '' for c in row]
    up = [v.upper() for v in vals]
    if 'TESTE' in up and any(k in up for k in ['ITENS DA VENDA', 'ARTICULOS MOVIMIENTO', 'ITENS']):
        header = vals
        start_idx = r_idx + 1
        break

# Find column indices
col_teste = header.index('Teste') if 'Teste' in header else 0
col_itens = header.index('Itens da venda') if 'Itens da venda' in header else 2
col_pagamento = header.index('Pagamento') if 'Pagamento' in header else 3
col_obs = header.index('Observacoes') if 'Observacoes' in header else 4
col_sat = header.index('SAT') if 'SAT' in header else 5
col_ecf = header.index('ECF') if 'ECF' in header else 6
col_nfce = header.index('NFCE') if 'NFCE' in header else 7
col_subtotal = header.index('Sub-Total') if 'Sub-Total' in header else 8
col_desconto = header.index('Desconto') if 'Desconto' in header else 9
col_total = header.index('Total') if 'Total' in header else 10
col_tipo_promo = header.index('Tipo Promo') if 'Tipo Promo' in header else 1

# Read all test rows
tests = []
seen_tests = set()
for row in ws.iter_rows(min_row=start_idx, values_only=True):
    vals = [str(c).strip() if c is not None else '' for c in row]
    if len(vals) < len(header):
        vals += [''] * (len(header) - len(vals))
    elif len(vals) > len(header):
        vals = vals[:len(header)]
    
    t_raw = vals[col_teste]
    if not t_raw:
        continue
    try:
        float(t_raw)
    except ValueError:
        continue
    t_key = t_raw.strip()
    if t_key in seen_tests:
        continue
    seen_tests.add(t_key)
    
    test = {
        'teste': float(t_key) if '.' in t_key else int(t_key),
        'tipo_promo': vals[col_tipo_promo],
        'itens_da_venda': vals[col_itens],
        'pagamento': vals[col_pagamento],
        'observacoes': vals[col_obs],
        'subtotal_esperado': vals[col_subtotal],
        'desconto_esperado': vals[col_desconto],
        'total_esperado': vals[col_total],
        'sat': vals[col_sat],
        'ecf': vals[col_ecf],
        'nfce': vals[col_nfce],
        'cupom': vals[col_nfce],  # Use NFCE as cupom
    }
    tests.append(test)

wb.close()

print(f"Total tests read: {len(tests)}")
for t in tests:
    print(f"  Teste {t['teste']}: Cupom={t['cupom']}, Sub={t['subtotal_esperado']}, Total={t['total_esperado']}")

# ============================================================
# 3. PROCESS TESTS THROUGH PIPELINE
# ============================================================
print("\n" + "=" * 60)
print("PROCESSING TESTS THROUGH PIPELINE")
print("=" * 60)

item_parser = ItemParser()
payment_normalizer = PaymentNormalizer()
api_builder = APISalesBuilder()
validator = TestValidator(partner_jsons=partner_jsons)

results = []

for test in tests:
    print(f"\n--- Teste {test['teste']} ---")
    print(f"  Itens: {test['itens_da_venda'][:80]}")
    print(f"  Pagamento: {test['pagamento'][:60]}")
    print(f"  Obs: {test['observacoes'][:60]}")
    print(f"  Cupom: {test['cupom']}")
    print(f"  Sub: {test['subtotal_esperado']}, Desc: {test['desconto_esperado']}, Total: {test['total_esperado']}")
    
    # Parse items
    test_with_items = item_parser.parse_items(test)
    
    # Normalize payment
    test_with_payments = payment_normalizer.normalize_payment(test_with_items)
    
    # Build sale_json for SDD validation
    sale_json = api_builder.build_sale_json(
        teste=test['teste'],
        itens_da_venda=test['itens_da_venda'],
        pagamento=test['pagamento'],
        subtotal=test['subtotal_esperado'],
        desconto=test['desconto_esperado'],
        total=test['total_esperado'],
        observacoes=test['observacoes'],
        numero_cupom=test['cupom'],
        tipo_promo=test['tipo_promo'],
        pagamentos=test_with_payments.get('pagamentos', []),
    )
    test_with_payments['sale_json'] = sale_json
    test_with_payments['partner_jsons'] = partner_jsons
    
    # LEGACY validation
    legacy_result = validator.validate_legacy(test_with_payments)
    legacy_status = legacy_result['status_final']
    legacy_motivo = legacy_result['motivo_status']
    
    # SDD validation
    try:
        sdd_result = validator.validate(test_with_payments)
        sdd_status = sdd_result['status_final']
        sdd_motivo = sdd_result['motivo_status']
    except Exception as e:
        sdd_status = 'ERROR'
        sdd_motivo = f'Validation error: {e}'
    
    print(f"  LEGACY: {legacy_status} - {legacy_motivo[:80]}")
    print(f"  SDD:    {sdd_status} - {sdd_motivo[:80]}")
    
    results.append({
        'teste': test['teste'],
        'cupom': test['cupom'],
        'legacy_status': legacy_status,
        'legacy_motivo': legacy_motivo,
        'sdd_status': sdd_status,
        'sdd_motivo': sdd_motivo,
        'itens': test['itens_da_venda'],
        'pagamento': test['pagamento'],
        'observacoes': test['observacoes'],
    })

# ============================================================
# 4. SUMMARY
# ============================================================
print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)

legacy_counts = {}
sdd_counts = {}
for r in results:
    legacy_counts[r['legacy_status']] = legacy_counts.get(r['legacy_status'], 0) + 1
    sdd_counts[r['sdd_status']] = sdd_counts.get(r['sdd_status'], 0) + 1

print(f"LEGACY distribution: {legacy_counts}")
print(f"SDD distribution:    {sdd_counts}")

print("\nDetailed results:")
for r in results:
    match = "✓" if r['legacy_status'] == r['sdd_status'] else "✗"
    print(f"  Teste {r['teste']:>3}: Legacy={r['legacy_status']:15} SDD={r['sdd_status']:15} {match}")

# Save detailed results
import os
os.makedirs('output', exist_ok=True)
with open('output/teste2_validation_results.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2)
print("\nResults saved to output/teste2_validation_results.json")
